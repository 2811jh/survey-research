#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
问卷分析工具 - 异动诊断 (survey_drift)
==========================================

按 周/月/天/季度/自定义区间/任意列 分桶，逐题相邻期显著性检验，
双门槛判异动，Agent 写一句话结论，导出 4-Sheet Excel。

子命令:
    analyze  分桶 + 检验 → drift_findings.json
    export   findings + conclusions → Excel
"""

import argparse
import json
import os
import re
import sys
from datetime import timedelta

import numpy as np
import pandas as pd
from scipy import stats


# ===================== 时间标签与分桶 ===================== #

def _fmt_md(d):
    return f"{d.month}.{d.day}"


def week_label(dt):
    iso = dt.isocalendar()
    monday = dt - timedelta(days=iso[2] - 1)
    sunday = monday + timedelta(days=6)
    return f"第{iso[1]}周（{_fmt_md(monday)}-{_fmt_md(sunday)}）"


def month_label(dt):
    return f"{str(dt.year)[2:]}年{dt.month}月"


def day_label(dt):
    return f"{dt.year}-{dt.month:02d}-{dt.day:02d}"


_LABELERS = {"week": week_label, "month": month_label, "day": day_label}


def bucketize(dt_series, granularity):
    """返回 (label_series, ordered_labels)。ordered_labels 按桶内最早时间升序（旧→新）。"""
    labeler = _LABELERS[granularity]
    labels = dt_series.apply(lambda d: labeler(d) if pd.notna(d) else None)
    valid = pd.DataFrame({"label": labels, "dt": dt_series}).dropna(subset=["label"])
    order = valid.groupby("label")["dt"].min().sort_values()
    return labels, list(order.index)


def quarter_label(dt):
    q = (dt.month - 1) // 3 + 1
    return f"{str(dt.year)[2:]}年Q{q}"


def _bucketize_quarter(dt_series):
    """按季度分桶。返回 (label_series, ordered_labels)。"""
    labels = dt_series.apply(lambda d: quarter_label(d) if pd.notna(d) else None)
    valid = pd.DataFrame({"label": labels, "dt": dt_series}).dropna(subset=["label"])
    order = valid.groupby("label")["dt"].min().sort_values()
    return labels, list(order.index)


def _bucketize_custom(dt_series, custom_ranges):
    """按自定义区间分桶。custom_ranges = [[label, start, end], ...]。
    区间为左闭右闭（含两端日期）。返回 (label_series, ordered_labels)。"""
    if not custom_ranges:
        raise ValueError("granularity=custom_ranges 时必须传 --custom_ranges")
    parsed = []
    for item in custom_ranges:
        if len(item) != 3:
            raise ValueError(f"custom_ranges 每项需为 [label, start, end]，实际：{item}")
        label, start, end = item
        parsed.append((label, pd.to_datetime(start), pd.to_datetime(end)))

    def _label_for(d):
        if pd.isna(d):
            return None
        for label, start, end in parsed:
            if start <= d <= end:
                return label
        return None

    labels = dt_series.apply(_label_for)
    ordered = [item[0] for item in parsed]
    # 过滤掉空桶
    present = set(labels.dropna().unique())
    ordered = [b for b in ordered if b in present]
    return labels, ordered


# ===================== 统计基元 ===================== #

def two_prop_z(c1, n1, c2, n2):
    """两比例 z 检验（双侧，pooled）。返回 (z, p)。n 为 0 时返回 (0.0, 1.0)。"""
    if n1 == 0 or n2 == 0:
        return 0.0, 1.0
    p1, p2 = c1 / n1, c2 / n2
    p_pool = (c1 + c2) / (n1 + n2)
    denom = p_pool * (1 - p_pool) * (1 / n1 + 1 / n2)
    if denom <= 0:
        return 0.0, 1.0
    z = (p1 - p2) / (denom ** 0.5)
    p = 2 * stats.norm.sf(abs(z))
    return float(z), float(p)


def compute_nps(series):
    """0~10 推荐题 → NPS。9-10 推荐者, 0-6 贬损者, 7-8 中立。"""
    vals = pd.to_numeric(series, errors="coerce").dropna()
    n = len(vals)
    if n == 0:
        return {"nps": 0.0, "promoter": 0, "detractor": 0, "n": 0}
    promoter = int((vals >= 9).sum())
    detractor = int((vals <= 6).sum())
    nps = (promoter - detractor) / n * 100
    return {"nps": float(nps), "promoter": promoter, "detractor": detractor, "n": n}


def compare_means(a_values, b_values):
    """相邻期均分比较。任一样本 n<30 用 Mann-Whitney U，否则 t 检验。
    返回 {test, p, mean_a, mean_b, delta}（delta = a - b，a 为较新期）。"""
    a = pd.to_numeric(pd.Series(a_values), errors="coerce").dropna()
    b = pd.to_numeric(pd.Series(b_values), errors="coerce").dropna()
    mean_a = float(a.mean()) if len(a) else 0.0
    mean_b = float(b.mean()) if len(b) else 0.0
    delta = mean_a - mean_b
    if len(a) < 2 or len(b) < 2:
        return {"test": "insufficient", "p": 1.0, "mean_a": mean_a, "mean_b": mean_b, "delta": delta}
    if a.std(ddof=1) == 0 and b.std(ddof=1) == 0:
        p = 1.0 if mean_a == mean_b else 0.0
        return {"test": "degenerate", "p": p, "mean_a": mean_a, "mean_b": mean_b, "delta": delta}
    if len(a) < 30 or len(b) < 30:
        _, p = stats.mannwhitneyu(a, b, alternative="two-sided")
        test = "mann_whitney_u"
    else:
        _, p = stats.ttest_ind(a, b, equal_var=False)
        test = "t_test"
    return {"test": test, "p": float(p), "mean_a": mean_a, "mean_b": mean_b, "delta": delta}


def evaluate_drift(delta, p, kind):
    """双门槛：p<0.05 且 实际差异达标（pp≥5 / mean≥0.1）。"""
    if p >= 0.05:
        return False
    if kind == "pp":
        return abs(delta) >= 5.0
    if kind == "mean":
        return abs(delta) >= 0.1
    return False


# ===================== 逐桶取数 ===================== #

def _bucket_mask(label_series, bucket):
    return label_series == bucket


def single_choice_props(df, col, label_series, ordered):
    """单选题各桶各选项占比。返回 (by_bucket={bucket:{option:prop}}, sizes={bucket:n})。"""
    by_bucket, sizes = {}, {}
    for b in ordered:
        sub = df.loc[_bucket_mask(label_series, b), col].dropna()
        sub = sub[sub.astype(str).str.strip() != ""]
        n = len(sub)
        sizes[b] = n
        if n == 0:
            by_bucket[b] = {}
            continue
        by_bucket[b] = (sub.astype(str).value_counts(normalize=True)).to_dict()
    return by_bucket, sizes


def _is_selected(v):
    if pd.isna(v):
        return False
    if isinstance(v, (int, float, np.integer, np.floating)):
        return float(v) != 0.0
    s = str(v).strip()
    return s not in ("", "0", "0.0", "nan", "NaN", "None", "否", "未选择")


def multi_choice_rates(df, subcols, root, label_series, ordered):
    """多选题各桶各选项勾选率。选项名取子列名冒号后部分（无冒号则去 root 前缀）。"""
    def opt_name(sc):
        s = str(sc)
        for sep in (":", "："):
            if sep in s:
                return s.split(sep, 1)[1].strip()
        return s[len(root):].strip() if s.startswith(root) else s

    by_bucket, sizes = {}, {}
    val_subcols = [sc for sc in subcols if "输入文本" not in str(sc)]
    for b in ordered:
        mask = _bucket_mask(label_series, b)
        block = df.loc[mask, val_subcols]
        # 选择矩阵：逐格判定是否勾选（与交叉分析一致，按"答过此题的人"为基数）
        selmat = pd.DataFrame({sc: block[sc].map(_is_selected) for sc in val_subcols},
                              index=block.index)
        base = int(selmat.any(axis=1).sum()) if val_subcols else 0  # 至少勾选一项=答过此题
        sizes[b] = base
        rates = {}
        for sc in val_subcols:
            sel = int(selmat[sc].sum())
            rates[opt_name(sc)] = float(sel / base) if base else 0.0
        by_bucket[b] = rates
    return by_bucket, sizes


def scale_means(df, col, label_series, ordered):
    """量表题各桶均分。返回 (by_bucket={bucket:mean}, sizes={bucket:n})。"""
    by_bucket, sizes = {}, {}
    for b in ordered:
        vals = pd.to_numeric(df.loc[_bucket_mask(label_series, b), col], errors="coerce").dropna()
        sizes[b] = len(vals)
        by_bucket[b] = float(vals.mean()) if len(vals) else 0.0
    return by_bucket, sizes


# ===================== 相邻期检验 ===================== #

def _adjacent_pairs(ordered):
    """返回 [(older, newer), ...]，按 ordered(旧→新) 相邻配对。"""
    return [(ordered[i - 1], ordered[i]) for i in range(1, len(ordered))]


def adjacent_prop_tests(by_bucket, sizes, ordered, min_n=30):
    """对每个选项、每对相邻期跑两比例 z 检验。返回 list[dict]。"""
    options = set()
    for b in ordered:
        options.update(by_bucket.get(b, {}).keys())
    results = []
    for older, newer in _adjacent_pairs(ordered):
        n_old, n_new = sizes.get(older, 0), sizes.get(newer, 0)
        low_n = n_old < min_n or n_new < min_n
        for opt in sorted(options):
            p_old = by_bucket.get(older, {}).get(opt, 0.0)
            p_new = by_bucket.get(newer, {}).get(opt, 0.0)
            c_old, c_new = round(p_old * n_old), round(p_new * n_new)
            z, pval = two_prop_z(c_new, n_new, c_old, n_old)
            delta_pp = (p_new - p_old) * 100
            drift = (not low_n) and evaluate_drift(delta_pp, pval, "pp")
            results.append({
                "option": opt, "from": older, "to": newer,
                "delta_pp": round(delta_pp, 1), "test": "two_prop_z",
                "p": round(pval, 4), "significant": pval < 0.05,
                "drift": drift, "low_n": low_n,
                "direction": "up" if delta_pp > 0 else ("down" if delta_pp < 0 else "flat"),
            })
    return results


def adjacent_mean_tests(df, col, label_series, ordered, min_n=30):
    """对量表题每对相邻期跑均分检验。返回 list[dict]。"""
    results = []
    for older, newer in _adjacent_pairs(ordered):
        a = df.loc[label_series == newer, col]
        b = df.loc[label_series == older, col]
        cmp = compare_means(a, b)
        low_n = len(pd.to_numeric(a, errors="coerce").dropna()) < min_n or \
                len(pd.to_numeric(b, errors="coerce").dropna()) < min_n
        drift = (not low_n) and evaluate_drift(cmp["delta"], cmp["p"], "mean")
        results.append({
            "from": older, "to": newer, "delta": round(cmp["delta"], 3),
            "test": cmp["test"], "p": round(cmp["p"], 4),
            "significant": cmp["p"] < 0.05, "drift": drift, "low_n": low_n,
            "direction": "up" if cmp["delta"] > 0 else ("down" if cmp["delta"] < 0 else "flat"),
        })
    return results


# ===================== 指标题识别 ===================== #

_NPS_KEYS = ["推荐给", "推荐给身边", "有多大可能将", "推荐给朋友", "净推荐"]
_SAT_KEYS = ["满意度", "满意程度", "整体满意", "整体体验"]


def identify_metric_cols(single_choice_cols):
    """从单选/量表列名按关键词识别 NPS 题与满意度题。返回 (nps_col|None, [sat_cols])。"""
    nps_col = None
    sat_cols = []
    for c in single_choice_cols:
        name = str(c)
        if nps_col is None and any(k in name for k in _NPS_KEYS):
            nps_col = c
            continue
        if any(k in name for k in _SAT_KEYS):
            sat_cols.append(c)
    return nps_col, sat_cols


# ===================== 数据加载 ===================== #

def _detect_csv_encoding(filepath, sample_size=8192):
    with open(filepath, "rb") as f:
        raw = f.read(sample_size)
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        raw.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        return "gbk"


def load_df(file_path):
    ext = file_path.rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xls"):
        df = pd.read_excel(file_path)
    else:
        df = pd.read_csv(file_path, encoding=_detect_csv_encoding(file_path), low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]
    return df


_TIME_COL_KEYWORDS = ("时间", "日期", "date", "time", "提交", "答题")


def _is_parseable_as_datetime(series, sample_size=50):
    """抽样检查列是否可解析为时间。纯数值列（如答题时长秒数）应被排除。"""
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    # 纯数值列（int/float，非 datetime64）在问卷场景下通常是时长/计数/编码，
    # 不应误判为时间列（pandas 会把小整数当 ns 纪元解析导致假阳性）。
    if pd.api.types.is_numeric_dtype(non_null) and not pd.api.types.is_datetime64_any_dtype(non_null):
        return False
    sample = non_null.sample(min(sample_size, len(non_null)), random_state=42)
    parsed = pd.to_datetime(sample, errors="coerce")
    # 解析成功率 ≥ 80% 视为时间列
    return parsed.notna().mean() >= 0.8


def detect_time_col(df, explicit):
    """时间列自动检测。返回 (col_name, source)。
    source 取值：explicit / default / auto_detect / not_found。"""
    if explicit:
        if explicit in df.columns and _is_parseable_as_datetime(df[explicit]):
            return explicit, "explicit"
        return None, "not_found"
    # 优先级 1：默认列
    if "结束答题时间" in df.columns and _is_parseable_as_datetime(df["结束答题时间"]):
        return "结束答题时间", "default"
    # 优先级 2：关键词扫描
    candidates = []
    for col in df.columns:
        s = str(col).lower()
        if any(kw in s for kw in _TIME_COL_KEYWORDS):
            if _is_parseable_as_datetime(df[col]):
                candidates.append(col)
    if candidates:
        # 取第一个命中的（按列顺序）
        return candidates[0], "auto_detect"
    return None, "not_found"


# ===================== findings 组装 ===================== #

def _five_point_scale_series(series):
    """判断某列是否五点量表（取值均为 1~5 的整数编码）。用于把五点量表题
    自动纳入 📊 指标总览 做均分显著性检验。NPS(0~10)、连续/非整数、二元题被排除。"""
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    vals = set()
    for v in non_null.unique():
        try:
            fv = float(v)
        except (ValueError, TypeError):
            return False
        iv = int(fv)
        if fv != iv:
            return False
        vals.add(iv)
    return bool(vals) and vals.issubset({1, 2, 3, 4, 5}) and max(vals) == 5 and len(vals) >= 4


def _multi_choice_label(root, subcols):
    """从多选子列还原完整题干：取首个子列冒号前的部分（含 root）。
    如 'Q6.为什么回归？:游戏版本更新' → 'Q6.为什么回归？'；失败回退 root。"""
    for sc in subcols:
        s = str(sc)
        idxs = [i for i in (s.find(":"), s.find("：")) if i != -1]
        if idxs:
            stem = s[:min(idxs)].strip()
            if stem:
                return stem
    return root


def build_findings(df, classification, granularity, time_col,
                   nps_col, satisfaction_cols, min_n=30,
                   bucket_col=None, bucket_order=None,
                   custom_ranges=None, time_col_source="default"):
    # 模式 B：列分桶
    if bucket_col:
        if bucket_col not in df.columns:
            raise KeyError(f"分桶列不存在：{bucket_col}")
        labels = df[bucket_col].astype(str)
        if bucket_order:
            ordered = [b for b in bucket_order if b in labels.unique()]
            # 补上 order 里没有的桶（避免丢数据）
            extras = [b for b in labels.unique() if b not in ordered]
            ordered = ordered + extras
        else:
            # 按出现顺序去重
            ordered = list(dict.fromkeys(labels.tolist()))
        bucket_mode = "column"
        granularity_out = None
        time_col_out = None
    else:
        # 模式 A：时间分桶
        if time_col not in df.columns:
            raise KeyError(f"时间列不存在：{time_col}")
        dt = pd.to_datetime(df[time_col], errors="coerce")
        if granularity == "custom_ranges":
            labels, ordered = _bucketize_custom(dt, custom_ranges)
        elif granularity == "quarter":
            labels, ordered = _bucketize_quarter(dt)
        else:
            labels, ordered = bucketize(dt, granularity)
        bucket_mode = "time"
        granularity_out = granularity
        time_col_out = time_col

    sizes_all = {b: int((labels == b).sum()) for b in ordered}
    low_n_buckets = [b for b, n in sizes_all.items() if n < min_n]

    metrics = []
    # 指标集合 = 关键词识别的满意度题 ∪ 全部五点量表单选题（都要做均分显著性检验）
    scale_cols = [c for c in classification.get("single_choice", [])
                  if c in df.columns and _five_point_scale_series(df[c])]
    metric_cols = list(dict.fromkeys((satisfaction_cols or []) + scale_cols))
    for col in metric_cols:
        if col not in df.columns:
            continue
        by_bucket, sizes = scale_means(df, col, labels, ordered)
        adj = adjacent_mean_tests(df, col, labels, ordered, min_n)
        metrics.append({
            "name": f"{col} 均分", "type": "satisfaction_mean", "source_col": col,
            "by_bucket": {b: round(by_bucket[b], 2) for b in ordered},
            "sizes": {b: int(sizes.get(b, 0)) for b in ordered}, "adjacent": adj,
        })
    if nps_col and nps_col in df.columns:
        by_bucket, sizes = {}, {}
        for b in ordered:
            r = compute_nps(df.loc[labels == b, nps_col])
            by_bucket[b] = round(r["nps"], 1)
            sizes[b] = r["n"]
        adj = []
        for older, newer in _adjacent_pairs(ordered):
            r_o = compute_nps(df.loc[labels == older, nps_col])
            r_n = compute_nps(df.loc[labels == newer, nps_col])
            z, pval = two_prop_z(r_n["promoter"], r_n["n"], r_o["promoter"], r_o["n"])
            delta_pp = by_bucket[newer] - by_bucket[older]
            low_n = r_o["n"] < min_n or r_n["n"] < min_n
            adj.append({
                "from": older, "to": newer, "delta_pp": round(delta_pp, 1),
                "test": "two_prop_z", "p": round(pval, 4), "significant": pval < 0.05,
                "drift": (not low_n) and evaluate_drift(delta_pp, pval, "pp"),
                "low_n": low_n,
                "direction": "up" if delta_pp > 0 else ("down" if delta_pp < 0 else "flat"),
            })
        metrics.append({
            "name": "NPS", "type": "nps", "source_col": nps_col,
            "by_bucket": by_bucket, "sizes": {b: int(sizes.get(b, 0)) for b in ordered},
            "adjacent": adj,
        })

    questions = []
    for col in classification.get("single_choice", []):
        by_bucket, sizes = single_choice_props(df, col, labels, ordered)
        opt_tests = adjacent_prop_tests(by_bucket, sizes, ordered, min_n)
        overall = _overall_chi_square(by_bucket, sizes, ordered)
        drift = any(t["drift"] for t in opt_tests)
        questions.append({
            "question": col, "type": "single_choice",
            "question_label": col,
            "options": sorted({o for b in ordered for o in by_bucket.get(b, {})}),
            "by_bucket": {b: {k: round(v, 4) for k, v in by_bucket.get(b, {}).items()} for b in ordered},
            "sizes": sizes, "overall_test": overall,
            "adjacent_option_tests": opt_tests, "drift": drift,
            "low_n": any(sizes.get(b, 0) < min_n for b in ordered),
        })
    for root, subcols in classification.get("multi_choice", {}).items():
        by_bucket, sizes = multi_choice_rates(df, subcols, root, labels, ordered)
        opt_tests = adjacent_prop_tests(by_bucket, sizes, ordered, min_n)
        drift = any(t["drift"] for t in opt_tests)
        questions.append({
            "question": root, "type": "multi_choice",
            "question_label": _multi_choice_label(root, subcols),
            "options": sorted({o for b in ordered for o in by_bucket.get(b, {})}),
            "by_bucket": {b: {k: round(v, 4) for k, v in by_bucket.get(b, {}).items()} for b in ordered},
            "sizes": sizes, "overall_test": None,
            "adjacent_option_tests": opt_tests, "drift": drift,
            "low_n": any(sizes.get(b, 0) < min_n for b in ordered),
        })

    return {
        "granularity": granularity_out, "time_col": time_col_out,
        "time_col_source": time_col_source,
        "bucket_mode": bucket_mode,
        "bucket_col": bucket_col,
        "custom_ranges": custom_ranges,
        "buckets": ordered, "bucket_sizes": sizes_all, "low_n_buckets": low_n_buckets,
        "metrics": metrics, "questions": questions,
        "nps_col": nps_col, "satisfaction_cols": metric_cols,
    }


def _overall_chi_square(by_bucket, sizes, ordered):
    """最新相邻期两桶 × 选项 的整体卡方。"""
    if len(ordered) < 2:
        return None
    older, newer = ordered[-2], ordered[-1]
    options = sorted({o for b in (older, newer) for o in by_bucket.get(b, {})})
    if not options:
        return None
    table = []
    for b in (older, newer):
        n = sizes.get(b, 0)
        table.append([round(by_bucket.get(b, {}).get(o, 0.0) * n) for o in options])
    arr = np.array(table)
    if arr.sum() == 0 or (arr.sum(axis=0) == 0).any():
        return None
    try:
        chi2, p, _, _ = stats.chi2_contingency(arr)
    except ValueError:
        return None
    return {"test": "chi_square", "from": older, "to": newer,
            "p": round(float(p), 4), "significant": bool(p < 0.05)}


# ===================== Excel 导出 ===================== #

def _overall_props(q, buckets):
    """全样本合并的整体占比（各桶按样本量加权），作为对比基线。
    返回 ({option: prop}, overall_n)。多选题同样成立（勾选率×n 求和 / 总n）。"""
    sizes = q.get("sizes", {})
    by = q.get("by_bucket", {})
    total = sum(sizes.get(b, 0) for b in buckets)
    res = {}
    for opt in q.get("options", []):
        if total > 0:
            s = sum(by.get(b, {}).get(opt, 0.0) * sizes.get(b, 0) for b in buckets)
            res[opt] = s / total
        else:
            res[opt] = 0.0
    return res, total


def _five_point_scale_opts(options):
    """判断是否五点量表题（选项均为 1~5 的整数）。是则返回 [(opt_str, 分值), ...]，否则 None。
    NPS(0~10) 因 max>5 被排除，二元题因 max<5/取值不足被排除。"""
    pairs = []
    for o in options:
        try:
            iv = int(float(str(o)))
        except (ValueError, TypeError):
            return None
        pairs.append((o, iv))
    vals = {iv for _, iv in pairs}
    if vals and vals.issubset({1, 2, 3, 4, 5}) and max(vals) == 5 and len(vals) >= 4:
        return pairs
    return None


def _weighted_satisfaction(prop_map, scale_pairs):
    """加权满意度（国际通用均分口径）：Σ(分值 × 该选项占比) = Σ(分值×人数)/总样本量。"""
    return sum(score * prop_map.get(opt, 0.0) for opt, score in scale_pairs)


_DEMOGRAPHIC_KEYS = ("性别", "年龄", "职业")


def _is_demographic(label):
    """人口统计题（性别/年龄/职业）——排序时保持选项原顺序，不按占比重排。"""
    s = str(label)
    return any(k in s for k in _DEMOGRAPHIC_KEYS)


def _qnum(key):
    """从题目键提取题号用于排序（如 'Q35.职业' → 35；'Q8.' → 8）。无题号排最后。"""
    m = re.match(r"\s*Q(\d+)", str(key))
    return int(m.group(1)) if m else 10 ** 6


def _trend_mark(delta, significant, is_pp):
    unit = "pp" if is_pp else "分"
    prec = 1 if is_pp else 2
    if not significant:
        return f"— {delta:+.{prec}f}{unit}（不显著）"
    arrow = "▲" if delta > 0 else "▼"
    return f"{arrow} {delta:+.{prec}f}{unit}"


def export_excel(findings, conclusions, output_path, summary_scope="latest",
                 value_labels=None):
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from openpyxl import Workbook

    conclusions = conclusions or {}
    value_labels = value_labels or {}
    buckets = findings["buckets"]
    wb = Workbook()

    # ---- Sheet 1: 指标总览 ----
    ws1 = wb.active
    ws1.title = "📊 指标总览"
    header = ["指标"] + buckets + ["最新vs上期", "是否显著"]
    ws1.append(header)
    for m in findings["metrics"]:
        is_pp = m["type"] == "nps"
        row = [m["name"]] + [m["by_bucket"].get(b, "") for b in buckets]
        if m["adjacent"]:
            last = m["adjacent"][-1]
            d = last.get("delta_pp", last.get("delta", 0.0))
            row += [_trend_mark(d, last["significant"], is_pp),
                    f"✅ p={last['p']}" if last["significant"] else f"✗ p={last['p']}"]
        else:
            row += ["—", "—"]
        ws1.append(row)
        # 每个指标下方紧跟「样本量」行（该指标各期有效作答人数，格式同逐题异动明细）
        m_sizes = m.get("sizes", {})
        ws1.append(["样本量"] + [m_sizes.get(b, "") for b in buckets] + ["", ""])
    _format_overview_sheet(ws1, len(buckets))

    # ---- Sheet 2: 逐题异动明细 ----
    ws2 = wb.create_sheet("📈 逐题异动明细")
    ws2.append(["题目", "选项", "整体"] + buckets + ["异动周", "AI 结论"])
    block_ranges = []          # (start_row, end_row) 每题一块（含样本量行），供分块斑马纹
    week_marks = {}            # (row, col) -> (kind, direction) 逐周环比标注
    sample_rows = set()        # 各题"样本量"行行号
    weighted_rows = set()      # 五点量表题"加权满意度"行行号
    section_rows = set()       # 归一化子区块小标题行行号
    concl_col = 3 + len(buckets) + 2  # AI结论列号（题目/选项/整体 + 各桶 + 异动周 + AI）
    for q in sorted(findings["questions"], key=lambda x: (_qnum(x["question"]), str(x["question"]))):
        qkey = q["question"]
        lmap = value_labels.get(qkey) or value_labels.get(q.get("question_label", qkey))

        def _disp(o, _m=lmap):
            return _m.get(str(o), o) if _m else o

        opts = q["options"]
        overall, overall_n = _overall_props(q, buckets)
        scale_pairs = _five_point_scale_opts(opts)
        is_demo = _is_demographic(q.get("question_label", qkey))
        if lmap and not scale_pairs:
            # 有编码→标签映射（人口题）：按编码数字升序，保持问卷逻辑顺序
            def _codekey(o):
                try:
                    return (0, int(float(o)))
                except (ValueError, TypeError):
                    return (1, str(o))
            opts = sorted(opts, key=_codekey)
        elif (q["type"] in ("single_choice", "multi_choice") and not scale_pairs
                and not is_demo):
            # 单选/多选按「整体」占比降序排；五点量表题、人口题保持原顺序
            opts = sorted(opts, key=lambda o: overall.get(o, 0.0), reverse=True)
        start_row = ws2.max_row + 1
        # 逐选项、逐"到达桶"的相邻期检验：{option: {to_bucket: test}}
        tests_by_opt = {}
        for t in q["adjacent_option_tests"]:
            tests_by_opt.setdefault(t["option"], {})[t["to"]] = t
        for i, opt in enumerate(opts):
            row_idx = ws2.max_row + 1
            drift_weeks = []
            for bi in range(1, len(buckets)):  # 从第 2 桶起才有环比
                t = tests_by_opt.get(opt, {}).get(buckets[bi])
                if not t:
                    continue
                col = 4 + bi  # 各桶列：桶0=D(4)，桶bi=4+bi
                if t.get("drift"):
                    week_marks[(row_idx, col)] = ("drift", t["direction"])
                    arrow = "▲" if t["delta_pp"] > 0 else "▼"
                    drift_weeks.append(f"{buckets[bi]}{arrow}{t['delta_pp']:+.1f}pp")
                elif t.get("significant"):
                    week_marks[(row_idx, col)] = ("sig", t["direction"])
            ws2.append([
                q.get("question_label", qkey) if i == 0 else "", _disp(opt),
                overall.get(opt, 0.0),
                *[q["by_bucket"].get(b, {}).get(opt, 0.0) for b in buckets],
                "；".join(drift_weeks), "",
            ])
        # 每题末行：样本量（整体 + 各桶有效样本量 n）
        sizes_q = q.get("sizes", {})
        ws2.append([
            "", "样本量", overall_n,
            *[sizes_q.get(b, 0) for b in buckets], "", "",
        ])
        sample_rows.add(ws2.max_row)
        # 五点量表题：样本量下再加一行加权满意度（1~5 均分，整体 + 各期）
        if scale_pairs:
            ws2.append([
                "", "加权满意度",
                _weighted_satisfaction(overall, scale_pairs),
                *[_weighted_satisfaction(q["by_bucket"].get(b, {}), scale_pairs) for b in buckets],
                "", "",
            ])
            weighted_rows.add(ws2.max_row)
        # 人口题：若含「不愿意透露」，补一版剔除后归一化占比 + 剔除后样本量（同样做逐期异动检验）
        refuse = {o for o in q["options"] if _disp(o) == "不愿意透露"} if lmap else set()
        if is_demo and refuse:
            keep = [o for o in opts if o not in refuse]
            keep_sum_ov = sum(overall.get(o, 0.0) for o in keep) or 1.0
            keep_sum_b = {b: (sum(q["by_bucket"].get(b, {}).get(o, 0.0) for o in keep) or 1.0)
                          for b in buckets}
            overall_refuse_p = sum(overall.get(o, 0.0) for o in refuse)
            overall_ex = int(round(overall_n * (1.0 - overall_refuse_p)))
            base_ex = {b: int(round(sizes_q.get(b, 0)
                                    * (1.0 - sum(q["by_bucket"].get(b, {}).get(o, 0.0) for o in refuse))))
                       for b in buckets}
            # 归一化后的相邻期两比例 z 检验（基数=剔除「不愿意透露」后的样本量）
            norm_by_bucket = {b: {o: q["by_bucket"].get(b, {}).get(o, 0.0) / keep_sum_b[b]
                                  for o in keep} for b in buckets}
            norm_tests = adjacent_prop_tests(norm_by_bucket, base_ex, buckets)
            norm_tests_by_opt = {}
            for t in norm_tests:
                norm_tests_by_opt.setdefault(t["option"], {})[t["to"]] = t
            ws2.append(["", "剔除「不愿意透露」后归一化", "", *["" for _ in buckets], "", ""])
            section_rows.add(ws2.max_row)
            for opt in keep:
                row_idx = ws2.max_row + 1
                drift_weeks = []
                for bi in range(1, len(buckets)):
                    t = norm_tests_by_opt.get(opt, {}).get(buckets[bi])
                    if not t:
                        continue
                    col = 4 + bi
                    if t.get("drift"):
                        week_marks[(row_idx, col)] = ("drift", t["direction"])
                        arrow = "▲" if t["delta_pp"] > 0 else "▼"
                        drift_weeks.append(f"{buckets[bi]}{arrow}{t['delta_pp']:+.1f}pp")
                    elif t.get("significant"):
                        week_marks[(row_idx, col)] = ("sig", t["direction"])
                ov = overall.get(opt, 0.0) / keep_sum_ov
                row_vals = [norm_by_bucket[b].get(opt, 0.0) for b in buckets]
                ws2.append(["", _disp(opt), ov, *row_vals, "；".join(drift_weeks), ""])
            ws2.append(["", "样本量", overall_ex, *[base_ex.get(b, 0) for b in buckets], "", ""])
            sample_rows.add(ws2.max_row)
        end_row = ws2.max_row
        block_ranges.append((start_row, end_row))
        concl = conclusions.get(qkey, "")
        if end_row >= start_row:
            ws2.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws2.merge_cells(start_row=start_row, start_column=concl_col,
                            end_row=end_row, end_column=concl_col)
            ws2.cell(row=start_row, column=concl_col, value=concl)
    _format_detail_sheet(ws2, len(buckets), block_ranges, week_marks,
                         sample_rows, weighted_rows, section_rows)

    # ---- Sheet 3: 异动汇总 ----
    ws3 = wb.create_sheet("⚠️ 异动汇总")
    ws3.append(["题目/指标", "时段", "变化项", "方向", "幅度", "显著性", "AI结论"])

    def _in_scope(t):
        return True if summary_scope == "all" else (t["to"] == buckets[-1])

    any_drift = False
    for m in findings["metrics"]:
        for t in m["adjacent"]:
            if t.get("drift") and _in_scope(t):
                any_drift = True
                d = t.get("delta_pp", t.get("delta", 0.0))
                arrow = "▲" if d > 0 else "▼"
                ws3.append([m["name"], t["to"], "整体", arrow, f"{d:+.2f}",
                            f"p={t['p']}", conclusions.get(m["source_col"], "")])
    for q in findings["questions"]:
        for t in q["adjacent_option_tests"]:
            if t.get("drift") and _in_scope(t):
                any_drift = True
                arrow = "▲" if t["delta_pp"] > 0 else "▼"
                ws3.append([q.get("question_label", q["question"]), t["to"], t["option"], arrow,
                            f"{t['delta_pp']:+.1f}pp", f"p={t['p']}",
                            conclusions.get(q["question"], "")])
    if not any_drift:
        msg = "本期各指标/题目均无显著异动" if summary_scope == "latest" \
            else "全时间线各指标/题目均无显著异动"
        ws3.append([msg, "", "", "", "", "", ""])
    _format_summary_sheet(ws3, no_drift=(not any_drift))

    # ---- Sheet 4: 方法与样本 ----
    ws4 = wb.create_sheet("ℹ️ 方法与样本")
    ws4.append(["项", "说明"])
    ws4.append(["分桶粒度", {"week": "按周", "month": "按月", "day": "按天"}.get(findings["granularity"])])
    ws4.append(["时间列", findings["time_col"]])
    ws4.append(["各桶样本量", "; ".join(f"{b}={findings['bucket_sizes'].get(b, 0)}" for b in buckets)])
    ws4.append(["样本不足桶(n<30)", "; ".join(findings["low_n_buckets"]) or "无"])
    ws4.append(["判异动门槛",
                "双门槛（须同时满足）：①统计显著 p<0.05；②效应量达标——占比类 |Δ|≥5pp，"
                "均分类 |Δ|≥0.1 分。仅显著不达幅度、或达幅度不显著，均只记「显著」不判「异动」。"
                "样本不足桶(任一期 n<30)不判异动。"])
    ws4.append(["检验方法总览",
                "占比类(单选各选项占比、多选各选项勾选率、人口题及其归一化占比)：两比例 z 检验；"
                "单选题整体分布：卡方检验；均分类(五点量表加权满意度)：Welch t 检验(n≥30)/Mann-Whitney U(n<30)；"
                "NPS：净推荐值差 + 推荐者比例两比例 z 检验。"])
    ws4.append(["① 两比例 z 检验",
                "适用：相邻两期某选项占比是否显著变化。理论：大样本下比例近似正态，"
                "H0 两期总体比例相等。合并比例 p̂=(x1+x2)/(n1+n2)，"
                "z=(p̂1−p̂2)/√(p̂(1−p̂)(1/n1+1/n2))，双尾取 p 值。"
                "用于：所有单选/多选选项、人口题(Q33/34/35)及其剔除「不愿意透露」归一化后占比、NPS 推荐者比例。"])
    ws4.append(["② 卡方检验(独立性)",
                "适用：相邻两期「选项×期」列联表整体分布是否改变。理论：χ²=Σ(O−E)²/E，"
                "E 为独立假设下期望频数，自由度=(选项数−1)。用于：每道单选题的整体分布是否异动"
                "(明细表该题的整体判定)。"])
    ws4.append(["③ Welch t 检验",
                "适用：相邻两期均分(1~5 分)差异。理论：不假设两组方差相等的 t 检验，"
                "t=(x̄1−x̄2)/√(s1²/n1+s2²/n2)，Welch–Satterthwaite 近似自由度。"
                "用于：五点量表题(如 Q1/Q4/Q13/Q14/Q20)的加权满意度均分，样本量 n≥30 时采用。"])
    ws4.append(["④ Mann-Whitney U 检验",
                "适用：小样本(n<30)或非正态时的均分差异，作为 t 检验的稳健替代。理论：基于秩和的"
                "非参数检验，比较两组分布位置。用于：五点量表题均分且某期 n<30 时自动改用。"])
    ws4.append(["⑤ NPS(净推荐值)",
                "定义：NPS=推荐者%(打 9~10 分)−贬损者%(打 0~6 分)，7~8 分为中立者。"
                "异动检验：对相邻期推荐者比例做两比例 z 检验，幅度以 NPS 差(pp)判双门槛。"
                "用于：指定了 --nps_col 的推荐意愿题。"])
    ws4.append(["加权满意度(均分)公式",
                "国际通用均分口径：加权满意度=Σ(分值×该分值人数)/总样本量=(1×n1+2×n2+3×n3+4×n4+5×n5)/N，"
                "结果落在 1~5 分。用于：选项恰为 1~5 的五点量表单选题，明细表在样本量行下单列一行，"
                "并纳入指标总览做均分显著性检验。"])
    ws4.append(["整体列", "明细表 C 列为全样本合并后的整体占比（各期按样本量加权），作为各周/月对比的基线"])
    ws4.append(["样本量行", "指标总览每个指标下方、明细表每题末行均标注该指标/题各期及整体的有效样本量 n（实际作答人数）；"
                          "多选题基数=答过此题(至少勾选一项)的人数，与交叉分析一致，逻辑门控题不计未触达者"])
    ws4.append(["加权满意度行", "五点量表题(选项1~5)在样本量行下再加一行加权满意度=Σ(分值×人数)/总样本量(即1~5均分)"])
    ws4.append(["人口题归一化子区块", "性别/年龄/职业(Q33/34/35)在样本量行下追加「剔除「不愿意透露」后归一化」子区块："
                              "对其余选项占比按剔除后基数重新归一(各期+整体)，样本量取剔除「不愿意透露」后的人数，"
                              "并同样对每个选项做相邻期两比例 z 异动检验(热力标注+异动周)"])
    ws4.append(["明细表颜色", "逐题明细中，某周单元格相对前一周显著变化会着色：琥珀底+加粗=大幅异动(双门槛)，红/绿字=一般显著(升绿/降红)，灰字=无显著环比变化"])
    ws4.append(["免责", "样本不足桶仅供参考，不判异动"])
    _format_method_sheet(ws4)

    wb.save(output_path)
    return {"status": "success", "output_path": output_path, "sheets": wb.sheetnames}


def _format_detail_sheet(ws, n_buckets, block_ranges=None, week_marks=None,
                         sample_rows=None, weighted_rows=None, section_rows=None):
    """逐题异动明细：Slate + Indigo 设计系统（对齐文本分析 Excel 风格）。
    深色表头 + C列整体基线 + 按题分块斑马纹 + 占比 DataBar + 逐周环比热力标注
    + 每题末行样本量（五点量表再加一行加权满意度）+ 异动周列 + 结论靛蓝卡片。
    week_marks: {(row, col): (kind, direction)}，kind ∈ {'drift','sig'}，标注某周相对前一周的显著变化。
    sample_rows: set(行号)，各题"样本量"行（整数计数、不参与 DataBar）。
    weighted_rows: set(行号)，五点量表"加权满意度"行（1~5 均分、不参与 DataBar）。
    section_rows: set(行号)，归一化子区块小标题行（如"剔除「不愿意透露」后归一化"）。"""
    import _styles as st
    from _styles import TextReportTheme as TR, Theme
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule

    week_marks = week_marks or {}
    sample_rows = sample_rows or set()
    weighted_rows = weighted_rows or set()
    section_rows = section_rows or set()
    special_rows = sample_rows | weighted_rows | section_rows
    border = st.thin_border()
    max_col = ws.max_column
    max_row = ws.max_row
    overall_col = 3
    b_first, b_last = 4, 3 + n_buckets
    drift_weeks_col, concl_col = b_last + 1, b_last + 2

    DRIFT_BG = "FEF3C7"   # amber-100 异动周高亮
    UP_FONT = "1E7D32"    # green-800 升
    DOWN_FONT = "C0392B"  # red-700 降
    center = st.ALIGN_CENTER
    left = st.ALIGN_LEFT
    top_left = st.ALIGN_TOP_LEFT

    # ---- 表头 ----
    ws.row_dimensions[1].height = 40
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = st.make_fill(TR.TITLE_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.WHITE)
        cell.alignment = center
        cell.border = border

    # ---- 分块斑马纹底色（每题一色，块间交替）----
    if not block_ranges:
        block_ranges = [(r, r) for r in range(2, max_row + 1)]
    row_base = {}
    for bi, (s, e) in enumerate(block_ranges):
        base = TR.WHITE if bi % 2 == 0 else TR.ZEBRA_ALT
        for r in range(s, e + 1):
            row_base[r] = base

    # ---- 数据行 ----
    for r in range(2, max_row + 1):
        base = row_base.get(r, TR.WHITE)
        is_sample = r in sample_rows
        is_weighted = r in weighted_rows
        is_section = r in section_rows
        ws.row_dimensions[r].height = 22
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if is_section and c not in (1, concl_col):  # 归一化子区块小标题行
                cell.fill = st.make_fill(TR.NOTE_BG)
                if c == 2:
                    cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True,
                                     italic=True, color=TR.INDIGO_MAIN)
                    cell.alignment = left
                else:
                    cell.font = Font(name=Theme.FONT_NAME, size=9, color=TR.TEXT_MUTE)
                    cell.alignment = center
                continue
            if c == 1:  # 题目（合并列）：垂直居中，跨整块显示
                cell.fill = st.make_fill(TR.INDIGO_ACCENT_BG)
                cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.INDIGO_DEEP)
                cell.alignment = left
            elif c == 2:  # 选项 / "样本量" / "加权满意度" 标签
                cell.fill = st.make_fill(base)
                if is_weighted:
                    cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True, color=TR.INDIGO_MAIN)
                elif is_sample:
                    cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True, color=TR.TEXT_MUTE)
                else:
                    cell.font = Font(name=Theme.FONT_NAME, size=10, color=TR.TEXT_MAIN)
                cell.alignment = left
            elif c == overall_col:  # 整体基线列
                cell.alignment = center
                if is_weighted:
                    cell.number_format = "0.00"
                    cell.fill = st.make_fill(base)
                    cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.INDIGO_MAIN)
                elif is_sample:
                    cell.number_format = "#,##0"
                    cell.fill = st.make_fill(base)
                    cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True, color=TR.TEXT_MUTE)
                else:
                    cell.number_format = "0.0%"
                    cell.fill = st.make_fill(TR.INDIGO_ACCENT_BG)
                    cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.INDIGO_DEEP)
            elif b_first <= c <= b_last:  # 各周占比 + 逐周环比热力标注 / 样本量 / 加权满意度
                cell.alignment = center
                if is_weighted:
                    cell.number_format = "0.00"
                    cell.fill = st.make_fill(base)
                    cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.INDIGO_MAIN)
                    continue
                if is_sample:
                    cell.number_format = "#,##0"
                    cell.fill = st.make_fill(base)
                    cell.font = Font(name=Theme.FONT_NAME, size=9, color=TR.TEXT_MUTE)
                    continue
                cell.number_format = "0.0%"
                mark = week_marks.get((r, c))
                if mark:
                    kind, direction = mark
                    color = UP_FONT if direction == "up" else DOWN_FONT
                    if kind == "drift":  # 大幅异动：琥珀底 + 加粗彩字
                        cell.fill = st.make_fill(DRIFT_BG)
                        cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=color)
                    else:                # 一般显著：彩字，不加底
                        cell.fill = st.make_fill(base)
                        cell.font = Font(name=Theme.FONT_NAME, size=10, color=color)
                else:                    # 无显著环比变化
                    cell.fill = st.make_fill(base)
                    cell.font = Font(name=Theme.FONT_NAME, size=10, color=TR.TEXT_SUB)
            elif c == drift_weeks_col:  # 异动周
                cell.fill = st.make_fill(base)
                has = bool(cell.value)
                cell.font = Font(name=Theme.FONT_NAME, size=9, bold=has,
                                 color=("B45309" if has else TR.TEXT_MUTE))
                cell.alignment = st.ALIGN_TOP_LEFT
            elif c == concl_col:  # AI 结论（合并卡片）
                cell.fill = st.make_fill(TR.INDIGO_BG)
                cell.font = Font(name=Theme.FONT_NAME, size=10, color=TR.INDIGO_DEEP)
                cell.alignment = top_left

    # ---- 占比列 DataBar（固定 0~1 刻度，跨题可比；每题占比行的连续区段，排除样本量/加权/小标题行）----
    for s, e in block_ranges:
        run_start = None
        for r in range(s, e + 2):  # e+1 作哨兵，冲刷末段连续区
            is_data = (r <= e) and (r not in special_rows)
            if is_data and run_start is None:
                run_start = r
            elif not is_data and run_start is not None:
                for c in range(overall_col, b_last + 1):
                    col = get_column_letter(c)
                    rng = f"{col}{run_start}:{col}{r - 1}"
                    rule = DataBarRule(start_type="num", start_value=0,
                                       end_type="num", end_value=1,
                                       color=TR.INDIGO_CHIP, showValue=True,
                                       minLength=0, maxLength=100)
                    ws.conditional_formatting.add(rng, rule)
                run_start = None

    # ---- 列宽 / 冻结 ----
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions[get_column_letter(overall_col)].width = 11
    for c in range(b_first, b_last + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.column_dimensions[get_column_letter(drift_weeks_col)].width = 30
    ws.column_dimensions[get_column_letter(concl_col)].width = 46
    ws.freeze_panes = "D2"
    ws.sheet_view.showGridLines = False


# ===================== Slate + Indigo 皮肤（其余 3 Sheet） ===================== #

_UP_FONT = "1E7D32"    # green-800 升
_DOWN_FONT = "C0392B"  # red-700 降
_DRIFT_BG = "FEF3C7"   # amber-100


def _slate_header(ws, height=38):
    """统一深色表头（slate-800 + 白色粗体）。"""
    import _styles as st
    from _styles import TextReportTheme as TR, Theme
    from openpyxl.styles import Font
    ws.row_dimensions[1].height = height
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = st.make_fill(TR.TITLE_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.WHITE)
        cell.alignment = st.ALIGN_CENTER
        cell.border = st.thin_border()


def _format_overview_sheet(ws, n_buckets):
    """指标总览：深色表头 + 靛蓝指标列 + 斑马纹 + 趋势/显著性颜色编码。"""
    import _styles as st
    from _styles import TextReportTheme as TR, Theme
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    border = st.thin_border()
    _slate_header(ws)
    trend_col = 1 + n_buckets + 1
    sig_col = 1 + n_buckets + 2
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 26
        is_sample = str(ws.cell(r, 1).value or "").strip() == "样本量"
        zebra = TR.WHITE if r % 2 == 0 else TR.ZEBRA_ALT
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c == 1:  # 指标 / 样本量 标签
                if is_sample:
                    cell.fill = st.make_fill(zebra)
                    cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True, color=TR.TEXT_MUTE)
                    cell.alignment = st.ALIGN_LEFT
                else:
                    cell.fill = st.make_fill(TR.INDIGO_ACCENT_BG)
                    cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.INDIGO_DEEP)
                    cell.alignment = st.ALIGN_LEFT
            elif c <= 1 + n_buckets:  # 各期数值 / 样本量
                cell.fill = st.make_fill(zebra)
                if is_sample:
                    cell.font = Font(name=Theme.FONT_NAME, size=9, color=TR.TEXT_MUTE)
                    cell.alignment = st.ALIGN_CENTER
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
                else:
                    cell.font = Font(name=Theme.FONT_NAME, size=11, bold=True, color=TR.INDIGO_MAIN)
                    cell.alignment = st.ALIGN_CENTER
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"
            elif c == trend_col:  # 最新vs上期
                cell.fill = st.make_fill(zebra)
                cell.alignment = st.ALIGN_CENTER
                v = str(cell.value or "")
                color = _UP_FONT if v.startswith("▲") else (_DOWN_FONT if v.startswith("▼") else TR.TEXT_MUTE)
                cell.font = Font(name=Theme.FONT_NAME, size=10, bold=v[:1] in ("▲", "▼"), color=color)
            elif c == sig_col:  # 是否显著
                cell.fill = st.make_fill(zebra)
                cell.alignment = st.ALIGN_CENTER
                sig = str(cell.value or "").startswith("✅")
                cell.font = Font(name=Theme.FONT_NAME, size=9, bold=sig,
                                 color=(TR.INDIGO_MAIN if sig else TR.TEXT_MUTE))
    ws.column_dimensions["A"].width = 40
    for c in range(2, 2 + n_buckets):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.column_dimensions[get_column_letter(trend_col)].width = 18
    ws.column_dimensions[get_column_letter(sig_col)].width = 14
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False


def _format_summary_sheet(ws, no_drift=False):
    """异动汇总：深色表头 + 方向/幅度颜色编码 + AI 结论靛蓝卡片。
    列：题目/指标 | 时段 | 变化项 | 方向 | 幅度 | 显著性 | AI结论。"""
    import _styles as st
    from _styles import TextReportTheme as TR, Theme
    from openpyxl.styles import Font
    border = st.thin_border()
    ncol = ws.max_column  # 7
    _slate_header(ws)
    if no_drift:  # 单行提示：整行合并、居中弱化
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        cell = ws.cell(row=2, column=1)
        cell.fill = st.make_fill(TR.NOTE_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=11, color=TR.TEXT_SUB)
        cell.alignment = st.ALIGN_CENTER
        ws.row_dimensions[2].height = 40
        for c in range(1, ncol + 1):
            ws.cell(row=2, column=c).border = border
    else:
        for r in range(2, ws.max_row + 1):
            zebra = TR.WHITE if r % 2 == 0 else TR.ZEBRA_ALT
            arrow = str(ws.cell(r, 4).value or "")
            dcolor = _UP_FONT if arrow.startswith("▲") else (_DOWN_FONT if arrow.startswith("▼") else TR.TEXT_MAIN)
            ws.row_dimensions[r].height = 30
            for c in range(1, ncol + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if c == 1:  # 题目/指标
                    cell.fill = st.make_fill(TR.INDIGO_ACCENT_BG)
                    cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.INDIGO_DEEP)
                    cell.alignment = st.ALIGN_LEFT
                elif c == 2:  # 时段
                    cell.fill = st.make_fill(zebra)
                    cell.font = Font(name=Theme.FONT_NAME, size=9, color=TR.TEXT_SUB)
                    cell.alignment = st.ALIGN_CENTER
                elif c == 3:  # 变化项
                    cell.fill = st.make_fill(zebra)
                    cell.font = Font(name=Theme.FONT_NAME, size=10, color=TR.TEXT_MAIN)
                    cell.alignment = st.ALIGN_LEFT
                elif c in (4, 5):  # 方向 / 幅度
                    cell.fill = st.make_fill(zebra)
                    cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=dcolor)
                    cell.alignment = st.ALIGN_CENTER
                elif c == 6:  # 显著性
                    cell.fill = st.make_fill(zebra)
                    cell.font = Font(name=Theme.FONT_NAME, size=9, color=TR.TEXT_SUB)
                    cell.alignment = st.ALIGN_CENTER
                else:  # AI 结论
                    cell.fill = st.make_fill(TR.INDIGO_BG)
                    cell.font = Font(name=Theme.FONT_NAME, size=10, color=TR.INDIGO_DEEP)
                    cell.alignment = st.ALIGN_TOP_LEFT
    for col, w in zip("ABCDEFG", (32, 20, 24, 8, 12, 12, 48)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _format_method_sheet(ws):
    """方法与样本：深色表头 + 靛蓝项名列 + 说明列浅底斑马纹。"""
    import _styles as st
    from _styles import TextReportTheme as TR, Theme
    from openpyxl.styles import Font
    border = st.thin_border()
    _slate_header(ws)
    col_b_width = 96
    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1)
        c1.fill = st.make_fill(TR.INDIGO_ACCENT_BG)
        c1.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.INDIGO_DEEP)
        c1.alignment = st.ALIGN_LEFT
        c1.border = border
        c2 = ws.cell(row=r, column=2)
        zebra = TR.WHITE if r % 2 == 0 else TR.ZEBRA_ALT
        c2.fill = st.make_fill(zebra)
        c2.font = Font(name=Theme.FONT_NAME, size=10, color=TR.TEXT_MAIN)
        c2.alignment = st.ALIGN_LEFT
        c2.border = border
        # 依说明文本长度估算换行行数，设置行高，避免长条文本被截断
        text = str(c2.value or "")
        lines = max(1, -(-len(text) // (col_b_width - 6)))  # 每行约 col_b_width-6 个字符
        ws.row_dimensions[r].height = max(28, lines * 16 + 6)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = col_b_width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ===================== CLI ===================== #

def default_output_filename(granularity, bucket_col=None):
    if bucket_col:
        # 列分桶模式：用列名简化作 label
        label = _short_col_label(bucket_col)
    else:
        label = {"week": "按周", "month": "按月", "day": "按天",
                 "quarter": "按季度", "custom_ranges": "按自定义区间"}.get(granularity, granularity)
    from datetime import datetime
    return f"问卷异动诊断_{label}_{datetime.now():%Y%m%d_%H%M}.xlsx"


def _short_col_label(col):
    """从列名提取简短 label：Q35.用户版本号 → 用户版本号；无前缀取整列。"""
    s = str(col)
    # 去掉 Q\d+. 前缀
    m = re.match(r"Q\d+\.\s*(.+)", s)
    return m.group(1) if m else s


def _cmd_analyze(args):
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from load_and_classify import classify_columns
    df = load_df(args.file_path)
    # 时间列自动检测（Task 3 会加 bucket_col 分支，此处先只处理时间分桶）
    time_col, time_col_source = detect_time_col(df, args.time_col)
    if time_col is None:
        return {"status": "need_input", "reason": "time_col_missing",
                "message": f"未找到时间列，可用列：{list(df.columns[:20])}；请用 --time_col 指定"}
    classification = classify_columns(df)
    single = classification["single_choice"]
    nps_col = args.nps_col or identify_metric_cols(single)[0]
    sat_cols = args.satisfaction_cols or identify_metric_cols(single)[1]
    if not nps_col and not sat_cols:
        return {"status": "need_input", "reason": "no_metric",
                "message": "未能自动识别 NPS/满意度题，请用 --nps_col / --satisfaction_cols 指定"}
    findings = build_findings(df, classification, args.granularity, time_col,
                              nps_col, sat_cols, args.min_n)
    out = args.findings_out or os.path.join(
        os.path.dirname(os.path.abspath(args.file_path)), "drift_findings.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(findings, f, ensure_ascii=False, indent=2)
    return {
        "status": "success", "granularity": args.granularity,
        "time_col": time_col, "time_col_source": time_col_source,
        "buckets": findings["buckets"], "bucket_sizes": findings["bucket_sizes"],
        "low_n_buckets": findings["low_n_buckets"],
        "questions_total": len(findings["questions"]),
        "questions_with_drift": sum(1 for q in findings["questions"] if q["drift"]),
        "metrics_total": len(findings["metrics"]),
        "findings_out": out, "nps_col": nps_col,
        "satisfaction_cols": findings["satisfaction_cols"],
    }


def _cmd_export(args):
    with open(args.findings, encoding="utf-8") as f:
        findings = json.load(f)
    conclusions = None
    if args.conclusions:
        with open(args.conclusions, encoding="utf-8") as f:
            conclusions = json.load(f)
    # 编码→标签映射：显式 --value-labels 优先，否则自动探测 findings 同目录的 value_labels.json
    value_labels = None
    vl_path = args.value_labels
    if not vl_path:
        cand = os.path.join(os.path.dirname(os.path.abspath(args.findings)), "value_labels.json")
        if os.path.exists(cand):
            vl_path = cand
    if vl_path and os.path.exists(vl_path):
        with open(vl_path, encoding="utf-8") as f:
            value_labels = json.load(f)
    out = args.output_path or os.path.join(
        os.path.dirname(os.path.abspath(args.findings)),
        default_output_filename(findings["granularity"]))
    return export_excel(findings, conclusions, out, summary_scope=args.summary_scope,
                        value_labels=value_labels)


def main():
    parser = argparse.ArgumentParser(description="问卷时间异动诊断")
    sub = parser.add_subparsers(dest="cmd", required=True)

    pa = sub.add_parser("analyze", help="分桶 + 检验 → findings JSON")
    pa.add_argument("--file_path", required=True)
    pa.add_argument("--granularity", required=True, choices=["week", "month", "day"])
    pa.add_argument("--time_col", default=None,
                    help="时间列名；缺省自动检测（默认列 + 关键词扫描）")
    pa.add_argument("--nps_col", default=None)
    pa.add_argument("--satisfaction_cols", nargs="*", default=None)
    pa.add_argument("--min_n", type=int, default=30)
    pa.add_argument("--findings_out", default=None)

    pe = sub.add_parser("export", help="findings + conclusions → Excel")
    pe.add_argument("--findings", required=True)
    pe.add_argument("--conclusions", default=None)
    pe.add_argument("--output_path", default=None)
    pe.add_argument("--value-labels", dest="value_labels", default=None,
                    help="编码→标签映射 JSON（{题目:{编码:标签}}）；缺省自动探测 findings 同目录 value_labels.json")
    pe.add_argument("--summary-scope", dest="summary_scope",
                    choices=["latest", "all"], default="latest",
                    help="异动汇总范围：latest=仅最新相邻期（默认）；all=全时间线任意相邻期")

    args = parser.parse_args()
    if args.cmd == "analyze":
        result = _cmd_analyze(args)
    else:
        result = _cmd_export(args)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if result.get("status") not in ("success", None):
        sys.exit(0 if result.get("status") == "need_input" else 1)


if __name__ == "__main__":
    main()
