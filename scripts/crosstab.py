#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
问卷分析工具 - 交叉分析
========================

完整的交叉分析流水线：
合并选项 → 交叉计算 → 得分计算 → 差异摘要 → 导出 Excel

输出专业格式化的 Excel + stdout JSON（交叉摘要 + 差异 + 得分）。

用法:
    python crosstab.py \
        --file_path "C:/xxx/data.xlsx" \
        --row_questions '["all"]' \
        --col_questions '["Q17.性别"]' \
        [--merge_rules '{"Q1.满意度": {"不满意": [1,2,3], "满意": [4,5]}}'] \
        [--calc_scores auto] \
        [--output_path "C:/xxx/data_交叉分析.xlsx"]
"""

import argparse
import json
import sys
import os
import re
import warnings
import pandas as pd
import numpy as np
from collections import defaultdict
from typing import Optional

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from load_and_classify import classify_columns
from _styles import (
    Theme, TextReportTheme as TR,
    thin_border, make_fill,
    ALIGN_CENTER, ALIGN_LEFT, ALIGN_RIGHT, ALIGN_TOP_LEFT,
)
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule


# ---- 漂移分析视觉系统色板（与 survey_drift.py 一致）----
_UP_FONT = "1E7D32"    # green-800 升 / max 高亮字色
_DOWN_FONT = "C0392B"  # red-700 降 / min 弱化字色
_DRIFT_BG = "FEF3C7"   # amber-100 max 单元格高亮底
_MIN_BG = "F8FAFC"     # slate-50 min 单元格弱化底
DIFF_THRESHOLD = 0.05  # 5pp 视为显著差异


# ========================================================================= #
#                           辅助函数
# ========================================================================= #

def _detect_csv_encoding(filepath, sample_size=8192):
    """检测 CSV 文件编码"""
    with open(filepath, 'rb') as f:
        raw = f.read(sample_size)
    if raw.startswith(b'\xef\xbb\xbf'):
        return 'utf-8-sig'
    try:
        raw.decode('utf-8')
        return 'utf-8'
    except UnicodeDecodeError:
        return 'gbk'


# ========================================================================= #
#                    文件命名 + 五点量表识别（通用化）                    #
# ========================================================================= #

_DEMOGRAPHIC_SHORT_KEYWORDS = ["性别", "年龄", "职业", "付费", "充值", "会员", "渠道", "地区", "城市", "设备"]


def _short_col_label(col):
    """从列名提取简短 label：Q33.请问您的性别是？ → 性别。"""
    s = str(col)
    m = re.match(r"Q\d+\.\s*(.+)", s)
    s = m.group(1) if m else s
    for kw in _DEMOGRAPHIC_SHORT_KEYWORDS:
        if kw in s:
            return kw
    return s[:8]


def default_output_filename(col_questions, file_path):
    """多分组用 _按{简称1}_{简称2}_{简称3}"""
    short_names = [_short_col_label(c) for c in col_questions]
    base = os.path.splitext(os.path.basename(file_path))[0]
    return f"{base}_交叉分析_按{'_'.join(short_names)}.xlsx"


def _five_point_scale_series(series):
    """判断某列是否五点量表（取值均为 1~5 的整数编码，且至少出现 4 个不同值）。"""
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    vals = set()
    for v in non_null.unique():
        try:
            fv = float(v)
        except (ValueError, TypeError):
            return False
        iv = int(fv)
        if fv != iv:
            return False
        vals.add(iv)
    return bool(vals) and vals.issubset({1, 2, 3, 4, 5}) and max(vals) == 5 and len(vals) >= 4


def _extract_subcol_number(subcol: str, prefix: str) -> int:
    """从多选子列名中提取选项序号"""
    suffix = subcol.split(prefix)[1].strip()
    match = re.search(r'^(\d+)', suffix)
    return int(match.group(1)) if match else 0


def _extract_score_from_option(option) -> Optional[float]:
    """从选项文本中提取数值分数"""
    if option is None:
        return None
    s = str(option).strip()
    if not s:
        return None
    match = re.match(r'^-?\d+(?:\.\d+)?', s)
    if not match:
        match = re.search(r'-?\d+(?:\.\d+)?', s)
    return float(match.group(0)) if match else None


# ========================================================================= #
#                     职业默认重编码预设 (Occupation Recode)
# ========================================================================= #

# 默认职业分类方案（按序输出）：
#   小学生 / 初中生 / 高中生 / 大学生 / 工作人群 / 其他 / 不愿意透露
# 每个类别用关键词子串匹配原始选项文本，以适配不同问卷的措辞差异。
# 匹配优先级 = 列表顺序（先匹配到者归入该类，学生类优先于工作人群）。
OCCUPATION_RECODE_PRESET = [
    ("小学生", ["小学"]),
    ("初中生", ["初中"]),
    ("高中生", ["高中", "中专", "职高", "技校"]),
    ("大学生", ["大学", "专科", "本科", "研究生", "硕士", "博士", "在校"]),
    ("工作人群", [
        "自由职业", "自雇", "无固定工作", "兼职", "国企", "事业单位", "公务员",
        "专业技术", "民营", "私企", "外企", "个体户", "私营", "企业主",
        "车间", "制造业", "生产", "商场", "餐饮", "运输", "服务业",
        "农林牧渔", "工作人员", "劳动者", "在职",
    ]),
    ("其他", ["其他"]),
    ("不愿意透露", ["不愿", "保密", "拒绝"]),
]


def build_occupation_recode(df: pd.DataFrame, column: str) -> dict:
    """
    针对职业列，依据 OCCUPATION_RECODE_PRESET 关键词匹配，
    自动构建 {新标签: [原始值...]} 的 merge_rules。

    未匹配到任何关键词的原始值归入「其他」。
    返回的 dict 仅包含实际出现的类别，顺序遵循预设。
    """
    if column not in df.columns:
        raise ValueError(f"列 '{column}' 不存在")

    raw_values = [v for v in df[column].dropna().unique()]
    rules = {label: [] for label, _ in OCCUPATION_RECODE_PRESET}

    for val in raw_values:
        text = str(val)
        matched = None
        for label, keywords in OCCUPATION_RECODE_PRESET:
            if any(kw in text for kw in keywords):
                matched = label
                break
        if matched is None:
            matched = "其他"
        rules[matched].append(val)

    # 去掉空类别，保持预设顺序
    return {label: vals for label, vals in rules.items() if vals}


# ========================================================================= #
#                        合并选项 (Merge / Recode)
# ========================================================================= #

def merge_options(
    df: pd.DataFrame,
    column: str,
    merge_rules: dict,
    new_column_name: Optional[str] = None,
) -> str:
    """
    合并指定列的选项值，在 df 上原地添加新列。

    Args:
        df: 数据 DataFrame
        column: 原始列名
        merge_rules: {"不满意": [1,2,3], "满意": [4,5]}
        new_column_name: 新列名（默认自动生成）

    Returns:
        新列名
    """
    if column not in df.columns:
        raise ValueError(f"列 '{column}' 不存在")

    mapping = {}
    for label, values in merge_rules.items():
        for v in values:
            mapping[v] = label

    if new_column_name is None:
        short_name = re.sub(r'^Q\d+\.', '', column).strip()
        if len(short_name) > 20:
            short_name = short_name[:20]
        new_column_name = f"recode_{short_name}"

    df[new_column_name] = df[column].map(mapping)
    return new_column_name


# ========================================================================= #
#                        交叉分析核心
# ========================================================================= #

def run_crosstab(
    df: pd.DataFrame,
    classification: dict,
    row_questions: list,
    col_questions: list,
) -> dict:
    """
    执行交叉分析。

    Args:
        df: 数据 DataFrame
        classification: 列分类信息
        row_questions: 行变量列表（支持 "all"、具体列名、多选题根如 "Q8."）
        col_questions: 列变量列表（分组维度）

    Returns:
        {
            "freq_df": DataFrame,       # 频数表
            "percent_df": DataFrame,    # 列百分比
            "col_totals": dict,         # 列合计
            "col_labels": list,         # 列标签
            "valid_rows_map": dict,     # 行变量类型映射
        }
    """
    multi_dict = classification["multi_choice"]

    # --- 处理 "all" ---
    if row_questions == ["all"] or row_questions == "all":
        row_questions = list(classification["valid_for_crosstab"])
        col_sources = set()
        for cq in col_questions:
            col_sources.add(cq)
        row_questions = [q for q in row_questions if q not in col_sources]

    # --- 验证并分类问题 ---
    def validate_and_classify(questions):
        valid = []
        invalid = []
        for q in questions:
            q_clean = str(q).strip()
            if q_clean in multi_dict:
                valid.append(("multi", q_clean))
            elif re.match(r'^Q\d+\.$', q_clean) and q_clean in multi_dict:
                valid.append(("multi", q_clean))
            elif q_clean in df.columns:
                valid.append(("single", q_clean))
            else:
                invalid.append(q_clean)
        return valid, invalid

    valid_rows, invalid_rows = validate_and_classify(row_questions)
    valid_cols, invalid_cols = validate_and_classify(col_questions)

    if invalid_rows:
        warnings.warn(f"无效行问题将被跳过：{invalid_rows}")
    if invalid_cols:
        warnings.warn(f"无效列问题将被跳过：{invalid_cols}")

    # --- 列条件生成 ---
    col_conditions = []
    col_totals = {}
    seen_cols = defaultdict(int)

    for q_type, q in valid_cols:
        q_clean = str(q).strip()
        seen_cols[q_clean] += 1
        instance_id = seen_cols[q_clean]

        if q_type == "multi":
            root = q_clean
            subcols = multi_dict[root]
            example_subcol = subcols[0]
            rest_part = example_subcol.split(root)[1].strip()
            if ':' in rest_part:
                question_text = rest_part.split(':', 1)[0].strip()
            elif '：' in rest_part:
                question_text = rest_part.split('：', 1)[0].strip()
            else:
                question_text = rest_part
            full_question = f"{root}{question_text}"
            if instance_id > 1:
                full_question += f" #{instance_id}"

            for subcol in subcols:
                rest_subcol = subcol.split(root)[1].strip()
                if ':' in rest_subcol:
                    option_text = rest_subcol.split(':', 1)[1].strip()
                elif '：' in rest_subcol:
                    option_text = rest_subcol.split('：', 1)[1].strip()
                else:
                    option_text = rest_subcol
                label = f"{full_question}\n{option_text}"
                cond = df[subcol] == 1
                col_conditions.append((label, cond))
                col_totals[label] = int(cond.sum())

            total_label = f"{full_question}\n总计"
            total_cond = (df[subcols] == 1).any(axis=1)
            col_conditions.append((total_label, total_cond))
            col_totals[total_label] = int(total_cond.sum())

        else:
            values = df[q_clean].dropna().unique()
            try:
                sorted_values = sorted(
                    values,
                    key=lambda x: int(re.match(r'^(\d+)', str(x)).group(1))
                )
            except Exception:
                sorted_values = sorted(values, key=str)

            unique_question = q_clean
            if instance_id > 1:
                unique_question += f" #{instance_id}"

            for value in sorted_values:
                label = f"{unique_question}\n{value}"
                cond = df[q_clean] == value
                col_conditions.append((label, cond))
                col_totals[label] = int(cond.sum())

            total_label = f"{unique_question}\n总计"
            total_cond = df[q_clean].notna()
            col_conditions.append((total_label, total_cond))
            col_totals[total_label] = int(total_cond.sum())

    # --- 行条件生成 ---
    row_conditions = []
    for q_type, q in valid_rows:
        if q_type == "multi":
            root = q
            subcols = multi_dict[root]
            first_rest = subcols[0].split(root)[1].strip()
            if ':' in first_rest:
                q_text = first_rest.split(':', 1)[0].strip()
            elif '：' in first_rest:
                q_text = first_rest.split('：', 1)[0].strip()
            else:
                q_text = first_rest
            full_question = f"{root}{q_text}"

            for subcol in subcols:
                rest = subcol.split(root)[1].strip()
                if ':' in rest:
                    option_text = rest.split(':', 1)[1].strip()
                elif '：' in rest:
                    option_text = rest.split('：', 1)[1].strip()
                else:
                    option_text = rest
                cond = df[subcol] == 1
                row_conditions.append(((full_question, option_text), cond))
            total_cond = (df[subcols] == 1).any(axis=1)
            row_conditions.append(((full_question, "总计"), total_cond))
        else:
            values = df[q].dropna().unique()
            try:
                sorted_values = sorted(
                    values,
                    key=lambda x: int(re.match(r'^(\d+)', str(x)).group(1))
                )
            except Exception:
                sorted_values = sorted(values, key=str)
            for value in sorted_values:
                cond = df[q] == value
                row_conditions.append(((q, str(value)), cond))
            total_cond = df[q].notna()
            row_conditions.append(((q, "总计"), total_cond))

    # --- 交叉统计计算 ---
    freq_results = []
    for (r_question, r_option), r_cond in row_conditions:
        row_data = {}
        for c_label, c_cond in col_conditions:
            count = int((r_cond & c_cond).sum())
            row_data[c_label] = count
        freq_results.append(row_data)

    index = pd.MultiIndex.from_tuples(
        [(rl[0], rl[1]) for rl, _ in row_conditions],
        names=["问题", "选项"]
    )
    col_labels = [cl for cl, _ in col_conditions]

    freq_df = pd.DataFrame(freq_results, index=index, columns=col_labels)

    # --- 列百分比 ---
    percent_df = freq_df.astype(float).copy()
    for question in percent_df.index.get_level_values(0).unique():
        q_mask = percent_df.index.get_level_values(0) == question
        total_idx = (question, "总计")
        if total_idx in freq_df.index:
            denom = freq_df.loc[total_idx].replace(0, np.nan)
        else:
            denom = pd.Series(col_totals).reindex(percent_df.columns).replace(0, np.nan)
        percent_df.loc[q_mask] = freq_df.loc[q_mask].div(denom, axis=1)
    percent_df = percent_df.fillna(0)

    return {
        "freq_df": freq_df,
        "percent_df": percent_df,
        "col_totals": col_totals,
        "col_labels": col_labels,
        "valid_rows_map": {q: q_type for q_type, q in valid_rows},
        "invalid_rows": invalid_rows,
        "invalid_cols": invalid_cols,
    }


# ========================================================================= #
#                      满意度 / NPS 得分计算
# ========================================================================= #

def _detect_score_type(question_name: str, df: pd.DataFrame) -> str:
    """自动识别题目是满意度还是 NPS"""
    q_lower = question_name.lower()
    if "nps" in q_lower or "推荐" in question_name:
        return "nps"
    if "满意度" in question_name or "满意" in question_name:
        return "satisfaction"
    if question_name in df.columns:
        values = df[question_name].dropna().unique()
        numeric_vals = []
        for v in values:
            score = _extract_score_from_option(v)
            if score is not None:
                numeric_vals.append(score)
        if numeric_vals:
            min_val, max_val = min(numeric_vals), max(numeric_vals)
            if min_val >= 0 and max_val >= 9:
                return "nps"
    return "satisfaction"


def _is_scoreable_question(question_name: str, df: pd.DataFrame) -> Optional[str]:
    """判断题目是否适合计算得分"""
    q_str = str(question_name)

    satisfaction_keywords = ["满意度", "满意", "评价如何", "评价是", "体验感受"]
    nps_keywords = ["NPS", "nps", "推荐"]

    has_satisfaction = any(kw in q_str for kw in satisfaction_keywords)
    has_nps = any(kw in q_str or kw.lower() in q_str.lower() for kw in nps_keywords)

    if not has_satisfaction and not has_nps:
        return None

    if question_name not in df.columns:
        return None

    values = df[question_name].dropna().unique()
    numeric_vals = []
    for v in values:
        score = _extract_score_from_option(v)
        if score is not None:
            numeric_vals.append(score)

    if len(numeric_vals) < 2:
        return None

    min_val, max_val = min(numeric_vals), max(numeric_vals)

    if has_nps:
        if min_val >= 0 and max_val >= 9 and max_val <= 10:
            return "nps"
        if has_satisfaction and min_val >= 1 and max_val <= 7:
            return "satisfaction"
        return None

    if has_satisfaction:
        if min_val >= 1 and max_val <= 10 and (max_val - min_val) >= 2:
            return "satisfaction"
        return None

    return None


def auto_detect_score_questions(df: pd.DataFrame, ct_result: dict) -> list:
    """自动识别可计算得分的题目：关键词（满意/NPS/推荐）∪ 五点量表自动检测。"""
    scoreable = []
    valid_rows = ct_result["valid_rows_map"]
    for q_name in valid_rows:
        if valid_rows[q_name] != "single":
            continue
        # 关键词识别
        if _is_scoreable_question(q_name, df) is not None:
            scoreable.append(q_name)
            continue
        # 五点量表自动检测
        if q_name in df.columns and _five_point_scale_series(df[q_name]):
            scoreable.append(q_name)
    return scoreable


def calc_scores(df: pd.DataFrame, ct_result: dict, score_questions: list) -> Optional[pd.DataFrame]:
    """
    计算满意度得分或 NPS。

    Returns:
        score_df（得分 DataFrame）或 None
    """
    freq_df = ct_result["freq_df"]
    row_type_map = ct_result["valid_rows_map"]

    score_results = []
    score_index = []
    score_type_info = {}

    for q in score_questions:
        q = str(q).strip()

        if q not in freq_df.index.get_level_values(0).unique():
            warnings.warn(f"题目 '{q}' 不在行变量中，已跳过")
            continue
        if row_type_map.get(q) != "single":
            warnings.warn(f"得分计算仅支持单选/量表题，已跳过：{q}")
            continue

        score_type = _detect_score_type(q, df)
        score_type_info[q] = score_type

        q_slice = freq_df.xs(q, level=0)

        if score_type == "satisfaction":
            value_map = {}
            for opt in q_slice.index:
                opt_str = str(opt).strip()
                if opt_str in ("总计", "合计", "Total"):
                    continue
                score_val = _extract_score_from_option(opt_str)
                if score_val is not None:
                    value_map[opt] = score_val

            if not value_map:
                continue

            q_counts = q_slice.loc[list(value_map.keys())]
            weights = pd.Series(value_map)
            numerator = (q_counts.T * weights).T.sum(axis=0)
            denominator = q_counts.sum(axis=0).replace(0, np.nan)
            score = numerator / denominator

            score_results.append(score)
            score_index.append((q, "满意度得分(加权均值)"))

        else:
            value_map = {}
            for opt in q_slice.index:
                opt_str = str(opt).strip()
                if opt_str in ("总计", "合计", "Total"):
                    continue
                score_val = _extract_score_from_option(opt_str)
                if score_val is not None:
                    value_map[opt] = score_val

            if not value_map:
                continue

            promoter_opts = [opt for opt, s in value_map.items() if s >= 9]
            detractor_opts = [opt for opt, s in value_map.items() if s <= 6]

            q_counts = q_slice.loc[list(value_map.keys())]
            total_count = q_counts.sum(axis=0).replace(0, np.nan)

            promoter_count = q_counts.loc[promoter_opts].sum(axis=0) if promoter_opts else 0
            detractor_count = q_counts.loc[detractor_opts].sum(axis=0) if detractor_opts else 0

            nps_score = (promoter_count - detractor_count) / total_count

            score_results.append(nps_score)
            score_index.append((q, "NPS得分(%)"))

        # 样本量行：该列分组下该题的有效作答数（紧跟得分行下方）
        non_total_slice = q_slice[~q_slice.index.astype(str).isin(["总计", "合计", "Total"])]
        sample_data = {}
        for col_label in freq_df.columns:
            col_str = str(col_label)
            if col_str.endswith("\n总计") or col_str == "总计":
                # 总计列样本量 = 该题全局有效作答数
                sample_data[col_label] = int(df[q].notna().sum())
            else:
                # 分组列样本量 = 该分组下该题有效作答数（排除总计行）
                sample_data[col_label] = int(non_total_slice[col_label].sum())
        score_results.append(pd.Series(sample_data, index=freq_df.columns))
        score_index.append((q, "样本量"))

    if not score_results:
        return None

    score_df = pd.DataFrame(
        score_results,
        index=pd.MultiIndex.from_tuples(score_index, names=["问题", "指标"]),
    )
    score_df = score_df.reindex(columns=freq_df.columns)
    return score_df


# ========================================================================= #
#                  显著性检验（分组 vs 分组维度总计）                      #
# ========================================================================= #

def two_prop_z(c1, n1, c2, n2):
    """两比例 z 检验（双侧，pooled）。返回 (z, p)。n 为 0 时返回 (0.0, 1.0)。"""
    if n1 == 0 or n2 == 0:
        return 0.0, 1.0
    p1 = c1 / n1
    p2 = c2 / n2
    p_pool = (c1 + c2) / (n1 + n2)
    if p_pool == 0 or p_pool == 1:
        return 0.0, 1.0
    se = (p_pool * (1 - p_pool) * (1 / n1 + 1 / n2)) ** 0.5
    if se == 0:
        return 0.0, 1.0
    z = (p1 - p2) / se
    from scipy import stats
    p = 2 * (1 - stats.norm.cdf(abs(z)))
    return round(z, 4), round(p, 4)


def _extract_dim_from_label(label):
    """从列标签提取分组维度名。'Q33.性别\\n男' → 'Q33.性别'。无 \\n 返回原值。"""
    s = str(label)
    if "\n" in s:
        return s.split("\n")[0]
    return s


def _extract_col_dimensions(ct_result: dict) -> list:
    """提取各分组维度信息。

    Returns:
        [{"question": dim, "values": [分组列标签...], "total_label": 总计列标签}]
    """
    col_labels = ct_result["col_labels"]
    dim_map = {}
    for label in col_labels:
        dim = _extract_dim_from_label(label)
        if dim not in dim_map:
            dim_map[dim] = {"question": dim, "values": [], "total_label": None}
        if str(label).endswith("\n总计"):
            dim_map[dim]["total_label"] = label
        else:
            dim_map[dim]["values"].append(label)
    return list(dim_map.values())


def calc_significance(ct_result: dict) -> dict:
    """对每个分组维度的各分组值 vs 该维度总计列，逐选项做两比例 z 检验。

    Returns:
        {分组维度列名: {分组值: {选项: {p, delta_pp, significant, direction}}}}
        direction: "up" (分组 > 总计) / "down" (分组 < 总计)
    """
    freq_df = ct_result["freq_df"]
    col_labels = ct_result["col_labels"]
    col_totals = ct_result["col_totals"]

    # 按分组维度归类列
    dim_cols = {}
    for label in col_labels:
        dim = _extract_dim_from_label(label)
        if dim not in dim_cols:
            dim_cols[dim] = {"total_col": None, "group_cols": []}
        if str(label).endswith("\n总计"):
            dim_cols[dim]["total_col"] = label
        else:
            dim_cols[dim]["group_cols"].append(label)

    result = {}
    for dim, info in dim_cols.items():
        total_col = info["total_col"]
        if total_col is None:
            continue
        total_n = col_totals[total_col]
        result[dim] = {}

        for group_col in info["group_cols"]:
            group_n = col_totals[group_col]
            group_value = str(group_col).split("\n")[-1]
            result[dim][group_value] = {}

            for idx in freq_df.index:
                option = idx[1] if isinstance(idx, tuple) else idx
                if str(option) in ("总计", "合计", "Total"):
                    continue
                c_group = int(freq_df.loc[idx, group_col])
                c_total = int(freq_df.loc[idx, total_col])

                z, p = two_prop_z(c_group, group_n, c_total, total_n)
                p_group = c_group / group_n if group_n else 0
                p_total = c_total / total_n if total_n else 0
                delta_pp = round((p_group - p_total) * 100, 1)
                significant = (p < 0.05) and (abs(delta_pp) >= 5)
                direction = "up" if delta_pp > 0 else "down"

                result[dim][group_value][str(option)] = {
                    "p": p, "delta_pp": delta_pp,
                    "significant": significant, "direction": direction,
                }
    return result


# ========================================================================= #
#                      差异摘要
# ========================================================================= #

def get_crosstab_summary(ct_result: dict, significance_matrix: dict = None) -> dict:
    """生成差异摘要：基于 vs 分组维度总计的显著性。

    每个分组维度取差异最大的显著选项，输出
    {dim: {max_diff_option, max_delta_pp, direction, significant, group}}。
    无显著项时该维度不出现。
    """
    if significance_matrix is None:
        significance_matrix = calc_significance(ct_result)

    diff_summary = {}
    for dim, groups in significance_matrix.items():
        for group_col, options in groups.items():
            for option, info in options.items():
                if not info["significant"]:
                    continue
                # 以维度为 key，找该维度下差异最大的显著项
                if dim not in diff_summary or abs(info["delta_pp"]) > abs(diff_summary[dim].get("max_delta_pp", 0)):
                    diff_summary[dim] = {
                        "max_diff_option": option,
                        "max_delta_pp": info["delta_pp"],
                        "direction": info["direction"],
                        "significant": True,
                        "group": group_col,
                    }
    return diff_summary


# ========================================================================= #
#                      Excel 导出 (Slate + Indigo 设计系统)
# ========================================================================= #

def _find_total_rows(ws, max_row, n_index_cols):
    """识别索引列含"总计/合计/Total"的行号集合。"""
    total_rows = set()
    for row_idx in range(2, max_row + 1):
        for ic in range(1, n_index_cols + 1):
            val = ws.cell(row=row_idx, column=ic).value
            if val and str(val).strip() in ("总计", "合计", "Total"):
                total_rows.add(row_idx)
                break
    return total_rows


def _format_crosstab_sheet(ws, is_percent=False, n_index_cols=2, has_total_col=True):
    """
    交叉分析 sheet 格式化（Slate + Indigo 风格，对齐 survey_drift 明细表）。

    - 表头：slate-800 底 + 白字粗体，行高 38
    - 索引列（题目/选项）：indigo-100 底 + indigo-900 粗体左对齐
    - 总计行：indigo-100 底 + indigo-900 粗体（与索引列同视觉层级）
    - 数据行：斑马纹（white / slate-100），TEXT_SUB 字色
    - DataBar：indigo-600，percent 固定 0~1 / freq 固定 0~max
    - 百分比格式 0.0% / 频数 0
    - freeze_panes C2，showGridLines False
    """
    max_row = ws.max_row
    max_col = ws.max_column
    border = thin_border()
    total_rows = _find_total_rows(ws, max_row, n_index_cols)

    # ---- 表头（第1行）----
    ws.row_dimensions[1].height = 38
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = make_fill(TR.TITLE_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.WHITE)
        cell.alignment = ALIGN_CENTER
        cell.border = border

    # ---- 数据行 ----
    data_seq = 0
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 22
        is_total_row = row_idx in total_rows
        if not is_total_row:
            data_seq += 1
        zebra = TR.WHITE if data_seq % 2 == 1 else TR.ZEBRA_ALT

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            if is_total_row:
                # 总计行：与索引列同视觉层级（indigo-100 底 + indigo-900 字）
                cell.fill = make_fill(TR.INDIGO_ACCENT_BG)
                cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.INDIGO_DEEP)
                if col_idx <= n_index_cols:
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_CENTER
                    if is_percent:
                        cell.number_format = '0.0%'
                    else:
                        cell.number_format = '#,##0'
            elif col_idx <= n_index_cols:
                # 索引列（题目/选项）
                cell.fill = make_fill(TR.INDIGO_ACCENT_BG)
                cell.font = Font(name=Theme.FONT_NAME, size=10,
                                bold=(col_idx == 1), color=TR.INDIGO_DEEP)
                cell.alignment = ALIGN_LEFT
            else:
                # 数据列
                cell.fill = make_fill(zebra)
                cell.font = Font(name=Theme.FONT_NAME, size=10, color=TR.TEXT_SUB)
                cell.alignment = ALIGN_CENTER
                if is_percent:
                    cell.number_format = '0.0%'
                else:
                    cell.number_format = '#,##0'

    # ---- DataBar（indigo-600，排除总计行）----
    non_total_rows = [r for r in range(2, max_row + 1) if r not in total_rows]
    if non_total_rows:
        for col_idx in range(n_index_cols + 1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            data_range = f"{col_letter}{min(non_total_rows)}:{col_letter}{max(non_total_rows)}"
            if is_percent:
                rule = DataBarRule(
                    start_type='num', start_value=0,
                    end_type='num', end_value=1,
                    color=TR.INDIGO_CHIP, showValue=True,
                    minLength=0, maxLength=100,
                )
            else:
                rule = DataBarRule(
                    start_type='num', start_value=0,
                    end_type='max',
                    color=TR.INDIGO_CHIP, showValue=True,
                    minLength=0, maxLength=100,
                )
            ws.conditional_formatting.add(data_range, rule)

    # ---- 列宽 ----
    ws.column_dimensions['A'].width = 34
    if n_index_cols >= 2:
        ws.column_dimensions['B'].width = 26
    for col_idx in range(n_index_cols + 1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ---- 冻结 + 隐藏网格线 ----
    freeze_col = get_column_letter(n_index_cols + 1)
    ws.freeze_panes = f"{freeze_col}2"
    ws.sheet_view.showGridLines = False


def _apply_significance_heatmap(ws, percent_df, col_labels, significance_matrix,
                                 start_row=2, n_index_cols=2):
    """列百分比 sheet 显著性着色 + DataBar。

    对每个分组值列的每个选项单元格：
    - 显著且 up: amber-100 底 + green-800 字 ↑ + DataBar
    - 显著且 down: amber-100 底 + red-700 字 ↓ + DataBar
    - 非显著: 保持斑马 + slate-600 字
    - 总计列: indigo-100 底（基准标识）
    """
    if not significance_matrix or percent_df is None or percent_df.empty:
        return

    border = thin_border()
    max_row = ws.max_row
    max_col = ws.max_column
    total_rows = _find_total_rows(ws, max_row, n_index_cols)

    # 总计列标 indigo-100 底
    for col_idx, label in enumerate(col_labels, start=n_index_cols + 1):
        if str(label).endswith("\n总计"):
            for r in range(1, max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.fill = make_fill(TR.INDIGO_ACCENT_BG)

    # 建行映射：ws 行号 -> (dim, group_col_value, option)
    # significance_matrix 结构: {dim: {group_col_value: {option: info}}}
    # group_col_value 是去掉 \n 前缀的分组值（如 "男"），需与 col_labels 匹配
    df_rows = list(percent_df.index)
    ws_row_to_info = {}
    for i, idx in enumerate(df_rows):
        ws_row = start_row + i
        if ws_row in total_rows:
            continue
        option = idx[1] if isinstance(idx, tuple) else idx
        option_str = str(option)
        if option_str in ("总计", "合计", "Total"):
            continue
        # 遍历 significance_matrix 找匹配的 (dim, group_col_value)
        for dim, groups in significance_matrix.items():
            for group_col_value, options in groups.items():
                if option_str in options:
                    ws_row_to_info[ws_row] = (dim, group_col_value, option_str)

    # 着色
    amber_fill = make_fill(_DRIFT_BG)  # FEF3C7
    for ws_row, (dim, group_col_value, option) in ws_row_to_info.items():
        info = significance_matrix[dim][group_col_value][option]
        # 找 group_col_value 对应的列号
        # col_labels 可能是 "Q33.性别\n男" 或裸 "男"——匹配 \n 后的部分
        for col_idx, label in enumerate(col_labels, start=n_index_cols + 1):
            label_str = str(label)
            label_value = label_str.split("\n")[-1] if "\n" in label_str else label_str
            if label_value == group_col_value or label_str == group_col_value:
                cell = ws.cell(row=ws_row, column=col_idx)
                if info["significant"]:
                    cell.fill = amber_fill
                    current_val = cell.value
                    if info["direction"] == "up":
                        cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=_UP_FONT)
                        if current_val and "↑" not in str(current_val):
                            cell.value = f"{current_val} ↑"
                    else:
                        cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=_DOWN_FONT)
                        if current_val and "↓" not in str(current_val):
                            cell.value = f"{current_val} ↓"
                break

    # DataBar：对所有非总计数据单元格加 indigo DataBar（刻度 0~1）
    non_total_rows = [r for r in range(start_row, max_row + 1) if r not in total_rows]
    if non_total_rows:
        for col_idx, label in enumerate(col_labels, start=n_index_cols + 1):
            if str(label).endswith("\n总计"):
                continue
            col_letter = get_column_letter(col_idx)
            data_range = f"{col_letter}{min(non_total_rows)}:{col_letter}{max(non_total_rows)}"
            rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='num', end_value=1,
                color=TR.INDIGO_CHIP, showValue=True,
                minLength=0, maxLength=100,
            )
            ws.conditional_formatting.add(data_range, rule)


def _format_score_sheet_v2(ws, col_labels, n_index_cols=2):
    """
    得分分析 sheet 格式化（Slate + Indigo 风格 + 最大差异标注）。

    - 表头 slate-800 + 白字，行高 38
    - 索引列（题目/指标）indigo-100 + indigo-900
    - 得分值 size 11 粗体 indigo-700 居中
    - 末尾追加"最大差异"列：▲ +X.X（绿字，>=5pp）/ — +X.X（slate-400，<5pp）
    - freeze_panes C2，showGridLines False
    """
    max_row = ws.max_row
    max_col = ws.max_column
    border = thin_border()

    # 先写表头：原 max_col 列 + 追加"最大差异"列
    diff_col = max_col + 1

    # 表头行
    ws.row_dimensions[1].height = 38
    for col_idx in range(1, diff_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = make_fill(TR.TITLE_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.WHITE)
        cell.alignment = ALIGN_CENTER
        cell.border = border
    # "最大差异"表头（如该列原无内容则写入标题）
    if ws.cell(row=1, column=diff_col).value is None:
        ws.cell(row=1, column=diff_col, value="最大差异")

    # 非总计列（用于计算得分跨组 max-min）
    non_total_cols = [c for c in col_labels if not str(c).endswith("\n总计")]
    col_to_ws_col = {}
    for ci, label in enumerate(col_labels, start=1):
        col_to_ws_col[label] = n_index_cols + ci

    # 数据行
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 26
        indicator_val = str(ws.cell(row=row_idx, column=2).value or "")
        is_nps = "NPS" in indicator_val
        zebra = TR.WHITE if row_idx % 2 == 0 else TR.ZEBRA_ALT

        # 计算该得分行跨分组的 max-min
        score_vals = []
        for label in non_total_cols:
            cidx = col_to_ws_col.get(label)
            if cidx is None:
                continue
            v = ws.cell(row=row_idx, column=cidx).value
            try:
                score_vals.append(float(v))
            except (TypeError, ValueError):
                pass

        if score_vals:
            delta = max(score_vals) - min(score_vals)
        else:
            delta = 0.0

        for col_idx in range(1, diff_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx <= n_index_cols:
                cell.fill = make_fill(TR.INDIGO_ACCENT_BG)
                cell.font = Font(name=Theme.FONT_NAME, size=10,
                                bold=(col_idx == 1), color=TR.INDIGO_DEEP)
                cell.alignment = ALIGN_LEFT
            elif col_idx == diff_col:
                # 最大差异标注列
                cell.fill = make_fill(zebra)
                if delta >= DIFF_THRESHOLD:
                    cell.value = f"▲ +{delta:.2f}"
                    cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=_UP_FONT)
                else:
                    cell.value = f"— +{delta:.2f} (不显著)"
                    cell.font = Font(name=Theme.FONT_NAME, size=9, color=TR.TEXT_MUTE)
                cell.alignment = ALIGN_CENTER
            else:
                cell.fill = make_fill(zebra)
                cell.font = Font(name=Theme.FONT_NAME, size=11, bold=True, color=TR.INDIGO_MAIN)
                cell.alignment = ALIGN_CENTER
                cell.number_format = '0.0%' if is_nps else '0.00'

    # 列宽
    ws.column_dimensions['A'].width = 34
    if n_index_cols >= 2:
        ws.column_dimensions['B'].width = 26
    for col_idx in range(n_index_cols + 1, diff_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    freeze_col = get_column_letter(n_index_cols + 1)
    ws.freeze_panes = f"{freeze_col}2"
    ws.sheet_view.showGridLines = False


def export_crosstab_excel(
    ct_result: dict,
    output_path: str,
    score_df: Optional[pd.DataFrame] = None,
    significance_matrix: dict = None,
    col_dimensions: list = None,
) -> str:
    """导出交叉分析 Excel 报告（Slate + Indigo 视觉风格）。

    significance_matrix / col_dimensions 为后续可视化任务预留（Tasks 7-9）。
    """
    freq_df = ct_result["freq_df"]
    percent_df = ct_result["percent_df"]
    col_labels = ct_result["col_labels"]

    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except PermissionError:
            raise PermissionError(f"请关闭正在使用的文件：{output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: 交叉分析（频数）
        freq_df.to_excel(writer, sheet_name='交叉分析', merge_cells=True)
        _format_crosstab_sheet(writer.sheets['交叉分析'], is_percent=False)
        writer.sheets['交叉分析'].sheet_properties.tabColor = TR.TITLE_BG

        # Sheet 2: 列百分比 — 显著性着色 + DataBar
        percent_df.to_excel(writer, sheet_name='列百分比', merge_cells=True)
        _format_crosstab_sheet(writer.sheets['列百分比'], is_percent=True)
        _apply_significance_heatmap(writer.sheets['列百分比'], percent_df, col_labels,
                                     significance_matrix, start_row=2, n_index_cols=2)
        writer.sheets['列百分比'].sheet_properties.tabColor = TR.INDIGO_CHIP

        # Sheet 3: 得分分析（如有）
        if score_df is not None and not score_df.empty:
            score_df.to_excel(writer, sheet_name='得分分析', merge_cells=True)
            _format_score_sheet_v2(writer.sheets['得分分析'], col_labels)
            writer.sheets['得分分析'].sheet_properties.tabColor = TR.INDIGO_DEEP

    return output_path


# ========================================================================= #
#                      JSON 输出生成
# ========================================================================= #

def _generate_output_json(
    ct_result: dict,
    diff_summary: dict,
    score_df: Optional[pd.DataFrame],
    output_path: str,
    significance_matrix: dict = None,
    col_dimensions: list = None,
) -> dict:
    """生成 stdout JSON 输出"""
    freq_df = ct_result["freq_df"]
    percent_df = ct_result["percent_df"]

    # 百分比表摘要
    percent_summary = {}
    for (q, opt) in percent_df.index:
        if q not in percent_summary:
            percent_summary[q] = {}
        percent_summary[q][opt] = {
            col: round(float(percent_df.loc[(q, opt), col]), 4)
            for col in percent_df.columns
        }

    # 得分摘要
    score_summary = None
    if score_df is not None and not score_df.empty:
        score_summary = {}
        non_total_cols = [c for c in ct_result["col_labels"] if not c.endswith("\n总计")]
        for (q, indicator) in score_df.index:
            scores_by_col = {}
            for col in non_total_cols:
                if col in score_df.columns:
                    scores_by_col[col] = round(float(score_df.loc[(q, indicator), col]), 4)
            score_summary[f"{q} - {indicator}"] = scores_by_col

    return {
        "status": "success",
        "output_path": output_path,
        "row_questions_count": len(ct_result["valid_rows_map"]),
        "col_conditions_count": len(ct_result["col_labels"]),
        "invalid_rows": ct_result.get("invalid_rows", []),
        "invalid_cols": ct_result.get("invalid_cols", []),
        "percent_table": percent_summary,
        "diff_summary": diff_summary,
        "score_summary": score_summary,
        "significant_matrix": significance_matrix,
        "col_dimensions": col_dimensions,
    }


# ========================================================================= #
#                        主函数
# ========================================================================= #

def run_crosstab_pipeline(
    file_path: str,
    row_questions: list,
    col_questions: list,
    sheet_name=0,
    merge_rules: dict = None,
    calc_scores_mode: str = None,
    output_path: str = None,
) -> dict:
    """
    完整的交叉分析流水线。

    Args:
        file_path: 数据文件路径
        row_questions: 行变量列表
        col_questions: 列变量列表
        sheet_name: 工作表名或编号
        merge_rules: {"列名": {"新标签": [原始值]}} 合并规则
        calc_scores_mode: "auto" 自动检测 / None 不计算
        output_path: 输出路径

    Returns:
        JSON 输出
    """
    # 加载数据
    ext = file_path.rsplit('.', 1)[-1].lower()
    if ext == 'csv':
        df = pd.read_csv(file_path, encoding=_detect_csv_encoding(file_path))
    else:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    classification = classify_columns(df)

    # auto 模式：识别候选分组维度
    if col_questions == ["auto"]:
        from load_and_classify import identify_demographic_cols
        candidates = identify_demographic_cols(df, classification)
        if not candidates:
            return {"status": "need_input", "reason": "no_demographic",
                    "message": "未识别到人口学题，请用 --col_questions 指定分组列"}
        return {"status": "need_input", "reason": "col_candidates",
                "candidates": candidates,
                "message": "识别到以下候选分组维度，请选择"}

    # 合并选项
    if merge_rules:
        for col_name, rules in merge_rules.items():
            # 支持职业默认预设：值为字符串 "occupation_default" 时自动构建分类
            if isinstance(rules, str) and rules == "occupation_default":
                rules = build_occupation_recode(df, col_name)
            new_col = merge_options(df, col_name, rules)
            # 将合并后的列替换到 col_questions 中
            if col_name in col_questions:
                idx = col_questions.index(col_name)
                col_questions[idx] = new_col
            # 更新分类信息
            classification["single_choice"].append(new_col)
            classification["valid_for_crosstab"].append(new_col)

    # 交叉分析
    ct_result = run_crosstab(df, classification, row_questions, col_questions)

    # 得分计算
    score_df = None
    if calc_scores_mode == "auto":
        score_questions = auto_detect_score_questions(df, ct_result)
        if score_questions:
            score_df = calc_scores(df, ct_result, score_questions)
    elif calc_scores_mode and calc_scores_mode != "none":
        try:
            score_questions = json.loads(calc_scores_mode)
            score_df = calc_scores(df, ct_result, score_questions)
        except json.JSONDecodeError:
            pass

    # 显著性检验（vs 分组维度总计）
    significance_matrix = calc_significance(ct_result)

    # 分组维度信息
    col_dimensions = _extract_col_dimensions(ct_result)

    # 差异摘要（基于显著性 vs 分组维度总计）
    diff_summary = get_crosstab_summary(ct_result, significance_matrix)

    # 输出路径（未指定时用 default_output_filename）
    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(file_path)),
            default_output_filename(col_questions, file_path),
        )

    # 导出 Excel
    export_crosstab_excel(ct_result, output_path, score_df, significance_matrix, col_dimensions)

    # 生成 JSON 输出
    return _generate_output_json(ct_result, diff_summary, score_df, output_path,
                                 significance_matrix, col_dimensions)


# ========================================================================= #
#                        CLI 入口
# ========================================================================= #

def main():
    parser = argparse.ArgumentParser(description="问卷交叉分析")
    parser.add_argument("--file_path", required=True, help="数据文件的绝对路径")
    parser.add_argument("--row_questions", required=True, help='行变量 JSON，如 \'["all"]\'')
    parser.add_argument("--col_questions", required=True, help='列变量 JSON，如 \'["Q17.性别"]\'')
    parser.add_argument("--sheet_name", default="0", help="工作表名或编号")
    parser.add_argument("--merge_rules", default=None, help='合并规则 JSON')
    parser.add_argument("--calc_scores", default=None, help='"auto" 或题目列表 JSON')
    parser.add_argument("--output_path", default=None, help="输出 Excel 路径")
    args = parser.parse_args()

    sheet_name = args.sheet_name
    try:
        sheet_name = int(sheet_name)
    except ValueError:
        pass

    try:
        row_questions = json.loads(args.row_questions)
    except json.JSONDecodeError as e:
        print(json.dumps({"error": f"row_questions JSON 解析失败: {e}"}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)

    try:
        col_questions = json.loads(args.col_questions)
    except json.JSONDecodeError as e:
        print(json.dumps({"error": f"col_questions JSON 解析失败: {e}"}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)

    merge_rules = None
    if args.merge_rules:
        try:
            merge_rules = json.loads(args.merge_rules)
        except json.JSONDecodeError as e:
            print(json.dumps({"error": f"merge_rules JSON 解析失败: {e}"}, ensure_ascii=False), file=sys.stderr)
            sys.exit(1)

    try:
        result = run_crosstab_pipeline(
            file_path=args.file_path,
            row_questions=row_questions,
            col_questions=col_questions,
            sheet_name=sheet_name,
            merge_rules=merge_rules,
            calc_scores_mode=args.calc_scores,
            output_path=args.output_path,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
