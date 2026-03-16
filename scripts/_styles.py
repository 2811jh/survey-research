"""
问卷分析工具 - 共享 Excel 样式模块
===================================

提供统一的专业级 Excel 格式化能力，包括：
- 配色方案（蓝灰系专业配色）
- 标题行、索引列、数据行、总计行样式
- DataBar 条件格式
- 报告专用样式

被 basic_stats.py / crosstab.py / text_export.py 等脚本共享使用。
"""

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.formatting.rule import DataBarRule


# ========================================================================= #
#                         配色方案 (专业蓝灰系)
# ========================================================================= #

class Theme:
    """统一配色主题"""
    # 标题行
    HEADER_BG = "2F5496"
    HEADER_FONT = "FFFFFF"
    # 索引列（问题 / 选项）
    INDEX_BG = "D6E4F0"
    INDEX_FONT = "1F3864"
    # 斑马纹
    ROW_EVEN_BG = "F2F2F2"
    ROW_ODD_BG = "FFFFFF"
    # 总计行
    TOTAL_BG = "E2EFDA"
    TOTAL_FONT = "375623"
    # DataBar
    FREQ_BAR = "5B9BD5"
    PCT_BAR = "ED7D31"
    # 边框
    BORDER_COLOR = "B4C6E7"
    # 得分
    SCORE_HEADER_BG = "4472C4"
    # 报告
    REPORT_HEADER_BG = "2F5496"
    REPORT_FINDING_BG = "FFF2CC"
    # 字体
    FONT_NAME = "微软雅黑"
    FONT_NAME_FALLBACK = "Arial"


class ReportTheme:
    """报告专用配色"""
    MODULE_TITLE_BG = "2F5496"
    MODULE_TITLE_FONT = "FFFFFF"
    QUESTION_TITLE_BG = "4472C4"
    QUESTION_TITLE_FONT = "FFFFFF"
    DATA_HEADER_BG = "D6E4F0"
    DATA_HEADER_FONT = "1F3864"
    FINDING_BG = "FFF2CC"
    FINDING_FONT = "7F6000"
    KEY_FINDING_BG = "E2EFDA"
    KEY_FINDING_TITLE_BG = "375623"
    RECOMMEND_BG = "FBE5D6"
    RECOMMEND_TITLE_BG = "C55A11"
    SUMMARY_BG = "E8E0F0"
    SUMMARY_TITLE_BG = "7030A0"
    DIFF_HIGH_FONT = "C00000"
    DIFF_LOW_FONT = "666666"


# ========================================================================= #
#                         基础样式工厂函数
# ========================================================================= #

def thin_border():
    """细边框"""
    side = Side(style='thin', color=Theme.BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def header_fill():
    """标题行填充"""
    return PatternFill(start_color=Theme.HEADER_BG, end_color=Theme.HEADER_BG, fill_type="solid")


def header_font(size=11):
    """标题行字体"""
    return Font(name=Theme.FONT_NAME, size=size, bold=True, color=Theme.HEADER_FONT)


def index_fill():
    """索引列填充"""
    return PatternFill(start_color=Theme.INDEX_BG, end_color=Theme.INDEX_BG, fill_type="solid")


def index_font(bold=False):
    """索引列字体"""
    return Font(name=Theme.FONT_NAME, size=10, bold=bold, color=Theme.INDEX_FONT)


def total_fill():
    """总计行填充"""
    return PatternFill(start_color=Theme.TOTAL_BG, end_color=Theme.TOTAL_BG, fill_type="solid")


def total_font():
    """总计行字体"""
    return Font(name=Theme.FONT_NAME, size=10, bold=True, color=Theme.TOTAL_FONT)


def body_font():
    """数据行字体"""
    return Font(name=Theme.FONT_NAME, size=10)


def even_fill():
    """偶数行填充"""
    return PatternFill(start_color=Theme.ROW_EVEN_BG, end_color=Theme.ROW_EVEN_BG, fill_type="solid")


def odd_fill():
    """奇数行填充"""
    return PatternFill(start_color=Theme.ROW_ODD_BG, end_color=Theme.ROW_ODD_BG, fill_type="solid")


def make_fill(color):
    """生成任意颜色填充"""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


# ========================================================================= #
#                    常用 Alignment 预设
# ========================================================================= #

ALIGN_CENTER = Alignment(wrap_text=True, vertical='center', horizontal='center')
ALIGN_LEFT = Alignment(wrap_text=True, vertical='center', horizontal='left')
ALIGN_RIGHT = Alignment(wrap_text=True, vertical='center', horizontal='right')
ALIGN_TOP_LEFT = Alignment(wrap_text=True, vertical='top', horizontal='left')


# ========================================================================= #
#              格式化交叉分析 / 百分比 / 频率统计 sheet
# ========================================================================= #

def format_data_sheet(ws, is_percent=False, index_cols=2):
    """
    通用数据表格式化：
    - 标题行：深蓝底 + 白字
    - 索引列（前 index_cols 列）：浅蓝底
    - 总计行：浅绿底 + 加粗
    - 斑马纹
    - DataBar
    - 百分比格式（如适用）

    Args:
        ws: openpyxl worksheet
        is_percent: 是否为百分比格式
        index_cols: 索引列数量（默认 2，即问题+选项）
    """
    max_row = ws.max_row
    max_col = ws.max_column
    border = thin_border()

    # ---- 标题行（第1行）----
    ws.row_dimensions[1].height = 50
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill()
        cell.font = header_font(size=10)
        cell.alignment = ALIGN_CENTER
        cell.border = border

    # ---- 找出总计行 ----
    total_rows = set()
    for row_idx in range(2, max_row + 1):
        # 检查索引列是否含"总计"
        for ic in range(1, index_cols + 1):
            val = ws.cell(row=row_idx, column=ic).value
            if val and str(val).strip() in ("总计", "合计", "Total"):
                total_rows.add(row_idx)
                break

    # ---- 数据行 ----
    data_row_count = 0
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 22
        is_total = row_idx in total_rows
        data_row_count += 1

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            if is_total:
                cell.fill = total_fill()
                cell.font = total_font()
                if col_idx <= index_cols:
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_RIGHT
                    if is_percent:
                        cell.number_format = '0%'
            elif col_idx <= index_cols:
                cell.fill = index_fill()
                cell.font = index_font(bold=(col_idx == 1))
                cell.alignment = ALIGN_LEFT
            else:
                if data_row_count % 2 == 0:
                    cell.fill = even_fill()
                else:
                    cell.fill = odd_fill()
                cell.font = body_font()
                cell.alignment = ALIGN_RIGHT
                if is_percent:
                    cell.number_format = '0%'

    # ---- DataBar ----
    bar_color = Theme.PCT_BAR if is_percent else Theme.FREQ_BAR
    non_total_rows = [r for r in range(2, max_row + 1) if r not in total_rows]

    if non_total_rows:
        for col_idx in range(index_cols + 1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            data_range = f"{col_letter}{min(non_total_rows)}:{col_letter}{max(non_total_rows)}"
            rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='max',
                color=bar_color,
                showValue="None",
                minLength=0,
                maxLength=100,
            )
            ws.conditional_formatting.add(data_range, rule)

    # ---- 列宽 ----
    for ic in range(1, index_cols + 1):
        ws.column_dimensions[get_column_letter(ic)].width = 30 if ic == 1 else 22
    for col_idx in range(index_cols + 1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ---- 冻结 + 隐藏网格线 ----
    freeze_col = get_column_letter(index_cols + 1)
    ws.freeze_panes = f"{freeze_col}2"
    ws.sheet_view.showGridLines = False


# ========================================================================= #
#                  格式化得分分析 sheet
# ========================================================================= #

def format_score_sheet(ws):
    """格式化得分分析 sheet"""
    max_row = ws.max_row
    max_col = ws.max_column
    border = thin_border()

    score_fill = make_fill(Theme.SCORE_HEADER_BG)
    ws.row_dimensions[1].height = 50
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = score_fill
        cell.font = header_font(size=10)
        cell.alignment = ALIGN_CENTER
        cell.border = border

    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 28
        indicator_val = str(ws.cell(row=row_idx, column=2).value or "")
        is_nps = "NPS" in indicator_val

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx <= 2:
                cell.fill = index_fill()
                cell.font = index_font(bold=(col_idx == 1))
                cell.alignment = ALIGN_LEFT
            else:
                cell.font = Font(name=Theme.FONT_NAME, size=12, bold=True, color="C00000")
                cell.alignment = ALIGN_CENTER
                cell.number_format = '0.0%' if is_nps else '0.00'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    for col_idx in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False


# ========================================================================= #
#              基础统计专用格式化
# ========================================================================= #

def format_basic_stats_sheet(ws, index_cols=1):
    """
    格式化基础统计 sheet（单索引列）。
    适用于样本概况、频率统计等。
    """
    format_data_sheet(ws, is_percent=False, index_cols=index_cols)


# ========================================================================= #
#              结构化分析报告 sheet
# ========================================================================= #

def write_structured_report(ws, report_data, percent_df=None, col_labels=None):
    """
    写入结构化分析报告 sheet。

    Args:
        ws: openpyxl worksheet
        report_data: 结构化 JSON:
            - per_question: [{question, finding}, ...]
            - key_findings: [str, ...]
            - recommendations: [str, ...]
            - summary: str
        percent_df: 交叉分析的百分比 DataFrame（可选）
        col_labels: 列标签列表（可选）
    """
    border = thin_border()
    non_total_cols = []
    if col_labels:
        non_total_cols = [c for c in col_labels if not c.endswith("\n总计")]
    data_col_count = 1 + len(non_total_cols) + 1
    total_width = max(data_col_count, 6)

    row = 1

    def write_module_title(title_text, bg_color):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
        cell = ws.cell(row=row, column=1, value=title_text)
        cell.fill = make_fill(bg_color)
        cell.font = Font(name=Theme.FONT_NAME, size=14, bold=True, color="FFFFFF")
        cell.alignment = ALIGN_CENTER
        cell.border = border
        ws.row_dimensions[row].height = 36
        for c in range(2, total_width + 1):
            ws.cell(row=row, column=c).border = border
        row += 1

    def write_question_title(q_text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
        cell = ws.cell(row=row, column=1, value=q_text)
        cell.fill = make_fill(ReportTheme.QUESTION_TITLE_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=12, bold=True, color=ReportTheme.QUESTION_TITLE_FONT)
        cell.alignment = ALIGN_LEFT
        cell.border = border
        ws.row_dimensions[row].height = 30
        for c in range(2, total_width + 1):
            ws.cell(row=row, column=c).border = border
        row += 1

    def write_data_table(question_name):
        nonlocal row
        if percent_df is None:
            return
        try:
            q_data = percent_df.xs(question_name, level=0)
        except KeyError:
            return

        option_rows = [opt for opt in q_data.index if str(opt).strip() not in ("总计", "合计", "Total")]
        if not option_rows:
            return

        h_fill = make_fill(ReportTheme.DATA_HEADER_BG)
        h_font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=ReportTheme.DATA_HEADER_FONT)

        headers = ["选项"] + [c.replace("\n", " ") for c in non_total_cols] + ["差异(pp)"]
        for ci, h in enumerate(headers):
            cell = ws.cell(row=row, column=ci + 1, value=h)
            cell.fill = h_fill
            cell.font = h_font
            cell.alignment = ALIGN_CENTER
            cell.border = border
        ws.row_dimensions[row].height = 26
        row += 1

        for ri, opt in enumerate(option_rows):
            pct_values = []
            for col in non_total_cols:
                try:
                    val = float(q_data.loc[opt, col])
                except (KeyError, ValueError, TypeError):
                    val = 0.0
                pct_values.append(val)

            diff_pp = (max(pct_values) - min(pct_values)) * 100 if pct_values else 0.0

            cell = ws.cell(row=row, column=1, value=str(opt))
            cell.font = Font(name=Theme.FONT_NAME, size=10, color=Theme.INDEX_FONT)
            cell.fill = even_fill() if ri % 2 == 0 else odd_fill()
            cell.alignment = ALIGN_LEFT
            cell.border = border

            for ci, pct in enumerate(pct_values):
                cell = ws.cell(row=row, column=ci + 2, value=pct)
                cell.number_format = '0.0%'
                cell.font = body_font()
                cell.fill = even_fill() if ri % 2 == 0 else odd_fill()
                cell.alignment = ALIGN_RIGHT
                cell.border = border

            diff_cell = ws.cell(row=row, column=len(pct_values) + 2, value=round(diff_pp, 1))
            diff_cell.number_format = '0.0'
            if abs(diff_pp) >= 5:
                diff_cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=ReportTheme.DIFF_HIGH_FONT)
            else:
                diff_cell.font = Font(name=Theme.FONT_NAME, size=10, color=ReportTheme.DIFF_LOW_FONT)
            diff_cell.fill = even_fill() if ri % 2 == 0 else odd_fill()
            diff_cell.alignment = ALIGN_CENTER
            diff_cell.border = border

            ws.row_dimensions[row].height = 22
            row += 1

    def write_finding(finding_text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
        cell = ws.cell(row=row, column=1, value=f"发现：{finding_text}")
        cell.fill = make_fill(ReportTheme.FINDING_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=10, color=ReportTheme.FINDING_FONT)
        cell.alignment = ALIGN_TOP_LEFT
        cell.border = border
        line_count = max(1, len(finding_text) // 60 + 1)
        ws.row_dimensions[row].height = max(40, line_count * 20)
        for c in range(2, total_width + 1):
            ws.cell(row=row, column=c).border = border
        row += 1

    def write_blank_row():
        nonlocal row
        ws.row_dimensions[row].height = 10
        row += 1

    def write_list_items(items, bg_color):
        nonlocal row
        item_fill = make_fill(bg_color)
        for i, item in enumerate(items, 1):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
            cell = ws.cell(row=row, column=1, value=f"  {i}. {item}")
            cell.fill = item_fill
            cell.font = Font(name=Theme.FONT_NAME, size=11)
            cell.alignment = ALIGN_LEFT
            cell.border = border
            line_count = max(1, len(item) // 70 + 1)
            ws.row_dimensions[row].height = max(28, line_count * 18)
            for c in range(2, total_width + 1):
                ws.cell(row=row, column=c).border = border
            row += 1

    # ---- 渲染报告 ----
    per_question = report_data.get("per_question", [])
    if per_question:
        write_module_title("逐题差异分析", ReportTheme.MODULE_TITLE_BG)
        write_blank_row()
        for q_item in per_question:
            q_name = q_item.get("question", "")
            q_finding = q_item.get("finding", "")
            write_question_title(q_name)
            write_data_table(q_name)
            if q_finding:
                write_finding(q_finding)
            write_blank_row()

    key_findings = report_data.get("key_findings", [])
    if key_findings:
        write_module_title("关键发现", ReportTheme.KEY_FINDING_TITLE_BG)
        write_list_items(key_findings, ReportTheme.KEY_FINDING_BG)
        write_blank_row()

    recommendations = report_data.get("recommendations", [])
    if recommendations:
        write_module_title("策略建议", ReportTheme.RECOMMEND_TITLE_BG)
        write_list_items(recommendations, ReportTheme.RECOMMEND_BG)
        write_blank_row()

    summary = report_data.get("summary", "")
    if summary:
        write_module_title("总结", ReportTheme.SUMMARY_TITLE_BG)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
        cell = ws.cell(row=row, column=1, value=summary)
        cell.fill = make_fill(ReportTheme.SUMMARY_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=11)
        cell.alignment = ALIGN_TOP_LEFT
        cell.border = border
        line_count = max(1, len(summary) // 70 + 1)
        ws.row_dimensions[row].height = max(60, line_count * 18)
        for c in range(2, total_width + 1):
            ws.cell(row=row, column=c).border = border
        row += 1

    # 列宽
    ws.column_dimensions['A'].width = 28
    for ci in range(2, total_width + 1):
        col_letter = get_column_letter(ci)
        if ci <= len(non_total_cols) + 1:
            ws.column_dimensions[col_letter].width = 20
        else:
            ws.column_dimensions[col_letter].width = 12

    ws.sheet_view.showGridLines = False


# ========================================================================= #
#              文本分析报告专用格式化
# ========================================================================= #

class TextReportTheme:
    """文本分析报告配色"""
    CONCLUSION_BG = "E2EFDA"
    CONCLUSION_FONT = "375623"
    DIMENSION_HEADER_BG = "4472C4"
    DIMENSION_HEADER_FONT = "FFFFFF"
    EXAMPLE_BG = "FFF2CC"
    EXAMPLE_FONT = "7F6000"
    DETAIL_HEADER_BG = "2F5496"
    DETAIL_HEADER_FONT = "FFFFFF"


def format_text_summary_sheet(ws):
    """格式化文本分析总结概览 sheet"""
    max_row = ws.max_row
    max_col = ws.max_column
    border = thin_border()

    # 标题行
    ws.row_dimensions[1].height = 40
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = make_fill(TextReportTheme.DIMENSION_HEADER_BG)
        cell.font = header_font(size=11)
        cell.alignment = ALIGN_CENTER
        cell.border = border

    # 数据行（斑马纹）
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 28
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx == 1:
                cell.fill = index_fill()
                cell.font = index_font(bold=True)
                cell.alignment = ALIGN_LEFT
            else:
                cell.fill = even_fill() if row_idx % 2 == 0 else odd_fill()
                cell.font = body_font()
                cell.alignment = ALIGN_CENTER if col_idx <= 3 else ALIGN_LEFT

    # 列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    if max_col >= 4:
        ws.column_dimensions['D'].width = 60

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def format_text_detail_sheet(ws):
    """格式化文本分析逐条明细 sheet"""
    max_row = ws.max_row
    max_col = ws.max_column
    border = thin_border()

    # 标题行
    ws.row_dimensions[1].height = 35
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = make_fill(TextReportTheme.DETAIL_HEADER_BG)
        cell.font = header_font(size=10)
        cell.alignment = ALIGN_CENTER
        cell.border = border

    # 数据行
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 40
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.font = body_font()
            if col_idx == 1:
                cell.alignment = ALIGN_TOP_LEFT
                cell.fill = even_fill() if row_idx % 2 == 0 else odd_fill()
            else:
                cell.alignment = ALIGN_CENTER
                cell.fill = even_fill() if row_idx % 2 == 0 else odd_fill()

    # 列宽
    ws.column_dimensions['A'].width = 80
    if max_col >= 2:
        ws.column_dimensions['B'].width = 30

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
