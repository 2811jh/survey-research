#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
问卷分析工具 - 文本分析结果导出
================================

将 AI 的文本分析结果导出为专业 Excel 报告。
每道题生成两个 sheet：总结概览 + 逐条明细。

用法:
    python text_export.py --output_path "C:/xxx/data_文本分析.xlsx" --results_file "C:/xxx/text_results.json"

    # 也可以直接传入 JSON 字符串
    python text_export.py --output_path "C:/xxx/data_文本分析.xlsx" --results_json '[{...}]'

results JSON 格式:
    [
        {
            "question": "Q10.您还有什么建议？",
            "conclusion": "核心结论（2-3句话）",
            "dimensions": [
                {
                    "name": "维度名（如：性能问题）",
                    "count": 100,
                    "percentage": "20.0%",
                    "examples": ["用户原文1", "用户原文2", ...]
                }
            ],
            "details": [
                {"text": "用户原文", "labels": "维度A, 维度B"}
            ]
        }
    ]
"""

import argparse
import json
import sys
import os
import re
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _styles import (
    Theme, TextReportTheme,
    format_text_summary_sheet, format_text_detail_sheet,
    thin_border, header_fill, header_font, index_fill, index_font,
    body_font, even_fill, odd_fill, make_fill,
    ALIGN_CENTER, ALIGN_LEFT, ALIGN_RIGHT, ALIGN_TOP_LEFT,
)
from text_extract import clean_column_texts, _clean_text
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule


def _clean_filename_part(text: str, max_len: int = 24) -> str:
    """生成安全的文件名片段。"""
    cleaned = re.sub(r'[\\/:*?"<>|]', '', text)
    cleaned = re.sub(r'\s+', '', cleaned)
    return cleaned[:max_len].strip() or "文本分析"


def _summarize_question_for_filename(question: str) -> str:
    """根据题目生成简短文件名语义。"""
    q_match = re.match(r'^(Q\d+)', question.strip())
    qid = q_match.group(1) if q_match else "文本题"

    activity_match = re.search(r'[“"]([^”"]+)[”"]', question)
    activity = activity_match.group(1) if activity_match else ""
    if activity:
        activity = activity.replace("活动", "").replace("节", "")

    if "MC移动版" in question and "比较满意" in question:
        topic = "MC满意原因"
    elif "外挂" in question:
        topic = "外挂表现"
    elif "模组" in question and "搜不到" in question:
        topic = "想玩缺失模组"
    elif "BUG" in question.upper():
        topic = "近期BUG"
    elif "丢失" in question and "存档" in question:
        topic = "存档丢失场景"
    elif "美术" in question or "画面" in question:
        topic = "美术不满原因"
    elif "当前版本" in question and "建议" in question:
        topic = "版本建议"
    elif "性能问题" in question and "哪些模组" in question:
        topic = "性能问题模组"
    elif "性能问题" in question:
        topic = "性能问题场景"
    elif "不太愿意" in question and "推荐" in question:
        topic = "不推荐原因"
    elif "更愿意" in question and "推荐" in question:
        topic = "推荐改进建议"
    elif "推荐给他人" in question and "打动" in question:
        topic = "推荐打动原因"
    elif "模组售后" in question or "问题反馈" in question:
        topic = "模组售后诉求"
    elif activity:
        if "比较满意" in question or "较为满意" in question:
            topic = f"{activity}满意原因"
        elif "不太满意" in question or "不满" in question or "一般" in question:
            topic = f"{activity}不满原因"
        elif "建议" in question or "期待" in question or "意见" in question:
            topic = f"{activity}建议"
        else:
            topic = activity
    elif "建议" in question or "期待" in question or "意见" in question:
        topic = "建议"
    elif "满意" in question:
        topic = "满意原因"
    elif "不满" in question or "一般" in question:
        topic = "不满原因"
    else:
        topic = "文本分析"

    return f"{qid}_{_clean_filename_part(topic)}.xlsx"


def default_output_filename(results: list) -> str:
    """根据文本分析结果生成默认 Excel 文件名。"""
    if not results:
        return "文本分析.xlsx"
    question = str(results[0].get("question", "文本分析"))
    return _summarize_question_for_filename(question)


# ========================================================================= #
#                        Excel 导出核心
# ========================================================================= #

def _detect_csv_encoding(filepath, sample_size=8192):
    """检测 CSV 文件编码"""
    with open(filepath, 'rb') as f:
        raw = f.read(sample_size)
    if raw.startswith(b'\xef\xbb\xbf'):
        return 'utf-8-sig'
    try:
        raw.decode('utf-8')
        return 'utf-8'
    except UnicodeDecodeError:
        return 'gbk'


def _safe_sheet_name(name: str, max_len: int = 28) -> str:
    """生成安全的 sheet 名称（去除非法字符，截断长度）"""
    # 去除 Excel sheet 名非法字符
    cleaned = re.sub(r'[\\/:*?"<>|]', '', name)
    if len(cleaned) > max_len:
        cleaned = cleaned[:max_len]
    return cleaned.strip()


def _auto_label_texts(texts: list, dimensions: list) -> list:
    """
    基于维度名称中的关键词，自动为每条文本标注所属维度。

    逻辑：
    1. 从每个维度名中提取关键词（按 / ( ) 、 分割）
    2. 同时从 examples 中提取高频词作为辅助关键词
    3. 逐条文本匹配，命中关键词最多的维度即为标签
    4. 支持多标签（一条文本可属于多个维度）

    Args:
        texts: 清洗后的全量文本列表
        dimensions: 维度列表，每项含 name, examples 等

    Returns:
        details 列表: [{"text": "...", "labels": "维度A, 维度B"}, ...]
    """
    if not dimensions or not texts:
        return []

    # 为每个维度构建关键词集合
    dim_keywords = []
    for dim in dimensions:
        name = dim.get("name", "")
        # 从维度名提取关键词：按常见分隔符拆分
        keywords = set()
        parts = re.split(r'[/（）()、,，\s]+', name)
        for p in parts:
            p = p.strip()
            if len(p) >= 2:  # 至少2个字的词才有效
                keywords.add(p)

        # 从 examples 中提取辅助关键词（取每条前5个2-4字词）
        for ex in dim.get("examples", []):
            ex_clean = ex.strip()
            if len(ex_clean) >= 2:
                short_words = re.findall(r'[\u4e00-\u9fff]{2,4}', ex_clean)
                for w in short_words[:5]:
                    keywords.add(w)

        dim_keywords.append({
            "name": name,
            "keywords": keywords,
        })

    # 逐条文本匹配
    details = []
    for text in texts:
        matched_dims = []
        for dk in dim_keywords:
            # 统计匹配到的关键词数量
            hit_count = sum(1 for kw in dk["keywords"] if kw in text)
            if hit_count > 0:
                matched_dims.append((dk["name"], hit_count))

        if matched_dims:
            # 按命中数降序，取所有命中的维度
            matched_dims.sort(key=lambda x: -x[1])
            labels = ", ".join(d[0] for d in matched_dims)
        else:
            labels = "其他"

        details.append({"text": text, "labels": labels})

    return details




OTHER_CANON = "其他/未归类"
_OTHER_ALIASES = {
    "其他", "其它", "无效", "模糊", "无效/模糊", "模糊/无效",
    "其他/无效", "无效/其他", "未归类", "其他/未归类", "无",
}


def _split_labels(s):
    """把 labels 字符串拆成标签列表，兼容中英逗号、顿号分隔。"""
    return [x.strip() for x in re.split(r'[,，、]', str(s or "")) if x.strip()]


def _canon_label(lab):
    """归一化标签：所有"其他/无效"类别名合并为统一的 OTHER_CANON。"""
    lab = (lab or "").strip()
    if not lab or lab in _OTHER_ALIASES:
        return OTHER_CANON
    return lab


def _rebuild_dimensions_from_details(details: list, ai_dimensions: list) -> list:
    """
    以逐条 details 为唯一数据源，重算各维度 count/percentage。

    - 合并所有"其他/无效"类别为统一桶（OTHER_CANON），并显式保留展示（不再隐藏）；
    - "其他/未归类"永远排在最后；
    - 维度顺序优先沿用 AI 给出的顺序，details 中出现的新标签追加其后；
    - examples 优先用 AI 提供的典型原声，缺失时回退为该标签下真实命中的原文（可验证）。

    多标签文本会分别计入各标签，故占比之和可能 > 100%（正常现象）。
    """
    total = len(details)
    counts = {}
    ex_by = {}
    seen_order = []
    for d in details:
        seen = set()
        for lab in _split_labels(d.get("labels", "")):
            canon = _canon_label(lab)
            if canon in seen:
                continue
            seen.add(canon)
            if canon not in counts:
                counts[canon] = 0
                ex_by[canon] = []
                seen_order.append(canon)
            counts[canon] += 1
            txt = (d.get("text") or "").strip()
            if txt and len(ex_by[canon]) < 5:
                ex_by[canon].append(txt)

    ai_examples = {}
    for dim in ai_dimensions or []:
        nm = dim.get("name", "")
        if _canon_label(nm) == OTHER_CANON:
            continue
        ai_examples[nm] = dim.get("examples", []) or []

    # 排序：非「其他」按条数降序（同数按名称稳定），「其他/未归类」固定最后一行
    non_other = [nm for nm in counts if nm != OTHER_CANON]
    non_other.sort(key=lambda nm: (-counts[nm], nm))
    final_order = non_other + ([OTHER_CANON] if OTHER_CANON in counts else [])

    # 归一化占比基数：剔除「其他/未归类」后的有效样本数
    other_count = counts.get(OTHER_CANON, 0)
    valid_n = max(total - other_count, 0)

    dims = []
    for nm in final_order:
        c = counts.get(nm, 0)
        pct = f"{c / total * 100:.1f}%" if total > 0 else "0%"
        # 归一化占比：剔除「其他/未归类」后按有效样本数重算；其他行留空（None）
        if nm == OTHER_CANON:
            norm_pct = None
        else:
            norm_pct = (c / valid_n) if valid_n > 0 else 0.0
        # 典型原文：优先 AI 精选，再用真实命中的原文补足到 5 条
        examples = []
        for e in ai_examples.get(nm, []):
            if e and e not in examples:
                examples.append(e)
            if len(examples) >= 5:
                break
        for e in ex_by.get(nm, []):
            if len(examples) >= 5:
                break
            if e and e not in examples:
                examples.append(e)
        dims.append({"name": nm, "count": c, "percentage": pct,
                     "norm_percentage": norm_pct, "examples": examples})
    return dims


_CIRCLED = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"


def _augment_conclusion(conclusion: str, dims: list, total: int) -> str:
    """
    组织「结论先行 + 分点描述」的可视化结论文本（数字由脚本回填，杜绝手写幻觉）。

    结构：
      核心结论：<AI 总述一句话>
      （分点按占比降序，总样本 N=…）
      ① 维度A  占比（条数）
      ② 维度B  占比（条数）
      …
      ⊕ 其他/未归类  占比（条数）  ← 固定最后
    """
    non_other = [d for d in dims if d["name"] != OTHER_CANON]
    other = next((d for d in dims if d["name"] == OTHER_CANON), None)

    lines = []
    lead = (conclusion or "").strip()
    if lead:
        lines.append(lead)
    lines.append(f"—— 分点明细（按占比降序，总样本 N={total} 条有效反馈）——")
    for i, d in enumerate(non_other):
        mark = _CIRCLED[i] if i < len(_CIRCLED) else f"({i + 1})"
        lines.append(f"{mark} {d['name']}  {d['percentage']}（{d['count']}条）")
    if other:
        lines.append(f"⊕ {OTHER_CANON}  {other['percentage']}（{other['count']}条）")
    return "\n".join(lines)


# ---- 可视化辅助（自适应行高 / 占比数值解析） ----


def _visual_len(s):
    """中日文字算 2 宽，其余算 1 宽，用于估算换行后的视觉行数。"""
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in str(s or ""))


def _estimate_row_height(text, col_width, line_px=16, min_px=24, max_px=409, pad=6):
    """按文本视觉长度 + 显式换行估算自适应行高（上限 409，Excel 行高极限）。"""
    if not text:
        return min_px
    lines = 0
    for seg in str(text).split("\n"):
        vlen = _visual_len(seg)
        lines += max(1, -(-vlen // max(1, int(col_width))))  # ceil
    return max(min_px, min(max_px, lines * line_px + pad))


def _pct_to_float(pct):
    """"20.0%" → 0.20；解析失败返回 0.0。"""
    try:
        return float(str(pct).replace("%%", "%").replace("%", "").strip()) / 100.0
    except (TypeError, ValueError):
        return 0.0


# ---- 逐条明细附加字段（从原始问卷数据回填 uid / 引擎版本 / 满意度等）----

def _resolve_column(df, token):
    """把用户给的列 token 解析为原始数据里的真实列名（精确 → 互相包含 → 忽略大小写包含）。"""
    if token in df.columns:
        return token
    for c in df.columns:
        cs = str(c)
        if token in cs or cs in token:
            return c
    tl = str(token).lower()
    for c in df.columns:
        if tl in str(c).lower():
            return c
    return None


def _auto_attach_header(col: str) -> str:
    """为附加列自动取一个简洁中文表头。"""
    cl = str(col).lower()
    if "uid" in cl:
        return "用户UID"
    if "engine" in cl or "引擎" in col:
        return "引擎版本"
    if "整体满意度" in col:
        return "整体满意度"
    base = re.sub(r'^[QYqy]\d+\.', '', str(col))
    base = re.split(r'[*:：\[]', base)[0].strip()
    return base[:14] or str(col)


def _attach_meta_columns(details, source_df, question_col, attach_columns, attach_headers=None):
    """
    为每条 detail 从原始数据回填附加字段（uid / 引擎版本 / 满意度等）。

    按「文本题原文」把逐条明细关联回原始行；同一原文出现多次时按出现顺序依次取值。
    返回附加列的表头列表；每条 detail 写入 detail["_meta"] = {表头: 值}。
    """
    from collections import defaultdict, deque

    text_col = _resolve_column(source_df, question_col)
    resolved = []
    for i, tok in enumerate(attach_columns or []):
        col = _resolve_column(source_df, tok)
        if col is None:
            continue
        if attach_headers and i < len(attach_headers) and attach_headers[i]:
            hdr = attach_headers[i]
        else:
            hdr = _auto_attach_header(col)
        resolved.append((col, hdr))

    if not resolved or text_col is None:
        return []

    lookup = defaultdict(deque)
    for _, r in source_df.iterrows():
        raw = r.get(text_col)
        if pd.isna(raw):
            continue
        key = _clean_text(str(raw))
        vals = {}
        for col, hdr in resolved:
            v = r.get(col)
            vals[hdr] = "" if pd.isna(v) else str(v)
        lookup[key].append(vals)

    headers = [hdr for _, hdr in resolved]
    for d in details:
        key = _clean_text(str(d.get("text", "")))
        picked = lookup[key].popleft() if lookup.get(key) else {}
        d["_meta"] = {h: picked.get(h, "") for h in headers}
    return headers


def _write_summary_sheet(writer, question_data: dict, sheet_idx: int):
    """
    写入单题的总结概览 sheet（纯 openpyxl 手写竖排布局）。

    布局：
      第1行：大标题 "文本分析报告"
      第2行：题目名称
      第3行：核心结论
      第4行：空行
      第5行：维度表表头（序号 | 问题类别 | 反馈条数 | 占比 | 典型用户原文）
      第6-N行：维度数据行，examples 用 bullet list 换行展示
    """
    question = question_data.get("question", f"题目{sheet_idx}")
    conclusion = question_data.get("conclusion", "")
    dimensions = question_data.get("dimensions", [])
    total_n = len(question_data.get("details", []))

    TR = TextReportTheme
    sheet_name = "总结概览"

    # 创建 sheet（不用 pandas）
    ws = writer.book.create_sheet(sheet_name)
    border = thin_border()
    total_width = 6  # 序号 | 问题类别 | 反馈条数 | 占比 | 归一化占比 | 典型用户原文

    row = 1

    # ---- 第1行：大标题（含副标） ----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
    cell = ws.cell(row=row, column=1, value="  文本分析报告   ·   开放题主题归纳")
    cell.fill = make_fill(TR.TITLE_BG)
    cell.font = Font(name=Theme.FONT_NAME, size=18, bold=True, color=TR.WHITE)
    cell.alignment = ALIGN_LEFT
    cell.border = border
    ws.row_dimensions[row].height = 50
    for c in range(2, total_width + 1):
        ws.cell(row=row, column=c).border = border
    row += 1

    # ---- 第2行：题目（右端标注总样本量） ----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
    cell = ws.cell(row=row, column=1, value=f"  题目  ·  {question}     ｜     总样本量 N = {total_n} 条有效反馈")
    cell.fill = make_fill(TR.SUBTITLE_BG)
    cell.font = Font(name=Theme.FONT_NAME, size=11, bold=True, color=TR.WHITE)
    cell.alignment = ALIGN_LEFT
    cell.border = border
    ws.row_dimensions[row].height = 34
    for c in range(2, total_width + 1):
        ws.cell(row=row, column=c).border = border
    row += 1

    # ---- 第3行：核心结论（冷靛底，结论先行 + 分点）----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
    conclusion_clean = conclusion.replace("%%", "%") if conclusion else ""
    conclusion_display = f"  核心结论    {conclusion_clean}" if conclusion_clean else ""
    cell = ws.cell(row=row, column=1, value=conclusion_display)
    cell.fill = make_fill(TR.INDIGO_BG)
    cell.font = Font(name=Theme.FONT_NAME, size=11, bold=False, color=TR.INDIGO_DEEP)
    cell.alignment = ALIGN_TOP_LEFT
    cell.border = border
    ws.row_dimensions[row].height = _estimate_row_height(conclusion_display, 126, line_px=18, min_px=64)
    for c in range(2, total_width + 1):
        ws.cell(row=row, column=c).border = border
    row += 1

    # ---- 第4行：空行 ----
    ws.row_dimensions[row].height = 14
    row += 1

    # ---- 维度统计表 ----
    if dimensions:
        # 表头
        headers = ["序号", "问题类别", "反馈条数", "占比", "归一化占比", "典型用户原文"]
        h_fill = make_fill(TR.SUBTITLE_BG)
        h_font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.WHITE)
        h_aligns = [ALIGN_CENTER, ALIGN_LEFT, ALIGN_CENTER, ALIGN_CENTER, ALIGN_CENTER, ALIGN_CENTER]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.fill = h_fill
            cell.font = h_font
            cell.alignment = h_aligns[ci - 1]
            cell.border = border
        ws.row_dimensions[row].height = 30
        row += 1

        # 数据行（干净斑马纹 + 统一靛蓝序号列 + 靛蓝占比条，对齐参考文件）
        first_data_row = row
        acc_fill = make_fill(TR.INDIGO_ACCENT_BG)
        for di, dim in enumerate(dimensions, 1):
            examples = dim.get("examples", [])
            example_text = "\n".join(f"·  {ex}" for ex in examples)

            is_other = dim.get("name", "") == OTHER_CANON
            pct_f = _pct_to_float(dim.get("percentage", "0%"))
            zebra = make_fill(TR.ZEBRA_ALT) if di % 2 == 0 else make_fill(TR.WHITE)

            name_color = TR.TEXT_MUTE if is_other else TR.TEXT_MAIN
            pct_color = TR.TEXT_MUTE if is_other else TR.INDIGO_MAIN

            # 序号（统一靛蓝强调列）
            cell = ws.cell(row=row, column=1, value=di)
            cell.fill = acc_fill
            cell.font = Font(name=Theme.FONT_NAME, size=11, bold=True, color=TR.INDIGO_MAIN)
            cell.alignment = ALIGN_CENTER
            cell.border = border

            # 问题类别
            cell = ws.cell(row=row, column=2, value=dim.get("name", ""))
            cell.fill = zebra
            cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=name_color)
            cell.alignment = ALIGN_LEFT
            cell.border = border

            # 反馈条数
            cell = ws.cell(row=row, column=3, value=dim.get("count", 0))
            cell.fill = zebra
            cell.font = Font(name=Theme.FONT_NAME, size=10, color=name_color)
            cell.alignment = ALIGN_CENTER
            cell.border = border

            # 占比（真数值 + 0.0% 格式 + DataBar）
            cell = ws.cell(row=row, column=4, value=pct_f)
            cell.number_format = '0.0%'
            cell.fill = zebra
            cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=pct_color)
            cell.alignment = ALIGN_RIGHT
            cell.border = border

            # 归一化占比（剔除「其他/未归类」后的有效样本占比；其他行留空）
            norm_val = dim.get("norm_percentage")
            cell = ws.cell(row=row, column=5, value=norm_val)
            cell.number_format = '0.0%'
            cell.fill = zebra
            cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True,
                             color=(TR.TEXT_MUTE if is_other else TR.INDIGO_MAIN))
            cell.alignment = ALIGN_RIGHT
            cell.border = border

            # 典型用户原文
            cell = ws.cell(row=row, column=6, value=example_text)
            cell.fill = zebra
            cell.font = Font(name=Theme.FONT_NAME, size=10, color=TR.TEXT_SUB)
            cell.alignment = ALIGN_TOP_LEFT
            cell.border = border

            ws.row_dimensions[row].height = _estimate_row_height(example_text, 68, line_px=17, min_px=56)
            row += 1

        # 占比列 DataBar（靛蓝条形，对齐参考文件）
        if row > first_data_row:
            rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='max',
                color=TR.INDIGO_CHIP,
                showValue=True, minLength=0, maxLength=100,
            )
            ws.conditional_formatting.add(f"D{first_data_row}:D{row - 1}", rule)
            # 归一化占比列 DataBar（深靛条形，剔除无效样本后的真实分布）
            norm_rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='max',
                color=TR.INDIGO_MAIN,
                showValue=True, minLength=0, maxLength=100,
            )
            ws.conditional_formatting.add(f"E{first_data_row}:E{row - 1}", norm_rule)

        # ---- 分析方法说明（表格下方注释，专业化） ----
        n_total = len(question_data.get("details", []))
        n_dims = sum(1 for d in dimensions if d.get("name", "") != OTHER_CANON)
        row += 1
        ws.row_dimensions[row].height = 12
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
        cell = ws.cell(row=row, column=1, value="  分析方法说明")
        cell.fill = make_fill(TR.SUBTITLE_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.WHITE)
        cell.alignment = ALIGN_LEFT
        cell.border = border
        ws.row_dimensions[row].height = 26
        for c in range(2, total_width + 1):
            ws.cell(row=row, column=c).border = border
        row += 1

        note_lines = [
            "· 方法：采用归纳式（自下而上）主题分析——先逐条阅读用户原文并开放编码，再将相近标签聚合为主题维度，未预设分类框架。",
            f"· 样本：有效抽样 N={n_total} 条；共归纳出 {n_dims} 个主题维度 + 1 个「其他/未归类」。",
            "· 多标签：一条反馈可同时命中多个维度，故各维度占比之和可能 > 100%（占比 = 该维度条数 ÷ N）。",
            "· 数据一致性：上表「反馈条数 / 占比」均由脚本从「逐条明细」页实时反算，两页数据完全一致，可逐条核对。",
            "· 归一化占比：在「占比」基础上剔除「其他/未归类」样本，按有效反馈数（N − 未归类条数）重新计算，更真实反映实质建议的分布；「其他/未归类」行不计算归一化占比（留空）。",
            "· 「其他/未归类」：无实质建议（如“很好”“没有”）或无法归入上述主题的反馈；占比越低说明主题覆盖越充分。",
            "· 典型用户原文：每个维度默认展示最多 5 条真实命中的反馈，均可在「逐条明细」页检索验证。",
        ]
        note_text = "\n".join(note_lines)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_width)
        cell = ws.cell(row=row, column=1, value=note_text)
        cell.fill = make_fill(TR.NOTE_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=9, color=TR.TEXT_SUB)
        cell.alignment = ALIGN_TOP_LEFT
        cell.border = border
        ws.row_dimensions[row].height = _estimate_row_height(note_text, 126, line_px=15, min_px=94)
        for c in range(2, total_width + 1):
            ws.cell(row=row, column=c).border = border
        row += 1

    # ---- 列宽 ----
    ws.column_dimensions['A'].width = 7    # 序号
    ws.column_dimensions['B'].width = 30   # 问题类别
    ws.column_dimensions['C'].width = 11   # 反馈条数
    ws.column_dimensions['D'].width = 12   # 占比
    ws.column_dimensions['E'].width = 14   # 归一化占比
    ws.column_dimensions['F'].width = 70   # 典型用户原文

    ws.sheet_properties.tabColor = TR.TITLE_BG
    ws.sheet_view.showGridLines = False

    return sheet_name


def _write_detail_sheet(writer, question_data: dict, sheet_idx: int):
    """
    写入单题的逐条明细 sheet（纯 openpyxl 手写，含序号列）。

    表头：序号 | 用户原文 | 归属类别 [ | 附加字段... ]
    附加字段来自 question_data["_meta_headers"]，每条 detail 的 item["_meta"]。
    """
    question = question_data.get("question", f"题目{sheet_idx}")
    details = question_data.get("details", [])
    meta_headers = question_data.get("_meta_headers", []) or []

    TR = TextReportTheme
    sheet_name = "逐条明细"

    if not details:
        ws = writer.book.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value="暂无明细数据")
        ws.sheet_properties.tabColor = TR.TEXT_SUB
        return sheet_name

    ws = writer.book.create_sheet(sheet_name)
    border = thin_border()

    # ---- 表头（基础 3 列 + 附加字段列）----
    headers = ["序号", "用户原文", "归属类别"] + list(meta_headers)
    h_fill = make_fill(TR.TITLE_BG)
    h_font = Font(name=Theme.FONT_NAME, size=10, bold=True, color=TR.WHITE)
    for ci, h in enumerate(headers, 1):
        align = ALIGN_LEFT if ci == 2 else ALIGN_CENTER
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[1].height = 34
    last_col = get_column_letter(len(headers))

    # ---- AutoFilter（表头筛选箭头）----
    ws.auto_filter.ref = f"A1:{last_col}{len(details) + 1}"

    meta_start_col = 4  # D 起为附加字段

    # ---- 数据行（干净斑马纹；类别淡靛标签条）----
    for ri, item in enumerate(details, 1):
        row_idx = ri + 1
        text = item.get("text", "")
        labels = item.get("labels", "")
        meta = item.get("_meta", {}) or {}
        zebra = make_fill(TR.ZEBRA_ALT) if ri % 2 == 0 else make_fill(TR.WHITE)

        # 序号
        cell = ws.cell(row=row_idx, column=1, value=ri)
        cell.fill = zebra
        cell.font = Font(name=Theme.FONT_NAME, size=9, color=TR.TEXT_MUTE)
        cell.alignment = ALIGN_CENTER
        cell.border = border

        # 用户原文
        cell = ws.cell(row=row_idx, column=2, value=text)
        cell.fill = zebra
        cell.font = Font(name=Theme.FONT_NAME, size=10, color=TR.TEXT_MAIN)
        cell.alignment = ALIGN_TOP_LEFT
        cell.border = border

        # 归属类别（淡靛"标签条"列，与原文区分层次）
        is_other = _canon_label((_split_labels(labels) or [""])[0]) == OTHER_CANON
        cell = ws.cell(row=row_idx, column=3, value=labels)
        cell.fill = zebra if is_other else make_fill(TR.INDIGO_BG)
        cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True,
                         color=(TR.TEXT_MUTE if is_other else TR.INDIGO_CHIP))
        cell.alignment = ALIGN_CENTER
        cell.border = border

        # 附加字段列（uid / 引擎版本 / 满意度等）
        for mi, h in enumerate(meta_headers):
            cell = ws.cell(row=row_idx, column=meta_start_col + mi, value=meta.get(h, ""))
            cell.fill = zebra
            cell.font = Font(name=Theme.FONT_NAME, size=9, color=TR.TEXT_SUB)
            cell.alignment = ALIGN_CENTER
            cell.border = border

        # 按原文长度自适应行高
        ws.row_dimensions[row_idx].height = _estimate_row_height(text, 80, min_px=34)

    # ---- 列宽 ----
    ws.column_dimensions['A'].width = 7    # 序号
    ws.column_dimensions['B'].width = 78   # 用户原文
    ws.column_dimensions['C'].width = 30   # 归属类别
    for mi, h in enumerate(meta_headers):
        col_letter = get_column_letter(meta_start_col + mi)
        if "UID" in h.upper() or "ID" in h.upper():
            w = 26
        elif "版本" in h or "engine" in h.lower():
            w = 16
        else:
            w = 20
        ws.column_dimensions[col_letter].width = w

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TR.TEXT_SUB

    return sheet_name


# ========================================================================= #
#                        主函数
# ========================================================================= #

def export_text_report(results: list, output_path: str,
                       file_path: str = None, sheet_name=0,
                       sample_n: int = 300,
                       attach_columns: list = None, attach_headers: list = None) -> dict:
    """
    将文本分析结果导出为专业 Excel 报告。

    数据源单一化：一律以逐条 details 为唯一数据源反算维度统计（count/percentage）
    并回填结论中的数字，确保总结概览与逐条明细完全一致、结论不出现幻觉数字。

    details 缺失时的策略（硬制）：
      - sample_n > 0（抽样模式，默认）：必须由 AI 逐条编码填入 details，否则直接
        返回 error（error_type=missing_details），不生成表格。
      - sample_n == 0（全量模式）：允许用 file_path + dimensions 关键词自动兜底
        （覆盖率较低，结果带 keyword_fallback_used 标记）。

    Args:
        results: 分析结果列表
        output_path: 输出文件路径
        file_path: 原始数据文件路径（全量兜底关键词标注 + 逐条明细附加字段回填）
        sheet_name: 工作表名或编号（默认 0）
        sample_n: 明细行数限制（默认 300，0=全量关键词兜底）
        attach_columns: 需要回填到「逐条明细」的原始列 token 列表（如 uid / 引擎版本 / 满意度）
        attach_headers: 与 attach_columns 一一对应的中文表头（可选，缺省自动命名）

    Returns:
        {"status": "success", "output_path": str, "sheets": [str]}
        或 {"status": "error", "error_type": "missing_details", ...}
    """
    if not results:
        return {"error": "results 不能为空"}

    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except PermissionError:
            raise PermissionError(f"请关闭正在使用的文件：{output_path}")

    # 如果需要自动标注，预加载原始数据
    source_df = None
    if file_path and os.path.exists(file_path):
        try:
            ext = file_path.rsplit('.', 1)[-1].lower()
            if ext == 'csv':
                source_df = pd.read_csv(file_path, encoding=_detect_csv_encoding(file_path), low_memory=False)
            else:
                source_df = pd.read_excel(file_path, sheet_name=sheet_name)
            source_df.columns = [str(c).strip() for c in source_df.columns]
        except Exception:
            source_df = None

    # ---- 硬制校验：抽样模式必须提供 details（逐条 LLM 编码），否则拒绝出表 ----
    missing_q = None
    for q in results:
        if q.get("details"):
            continue
        if sample_n == 0 and q.get("dimensions") and source_df is not None:
            continue  # 全量模式允许关键词兜底
        missing_q = q
        break
    if missing_q is not None:
        return {
            "status": "error",
            "error_type": "missing_details",
            "question": missing_q.get("question", ""),
            "message": (
                "抽样模式下必须先为每条抽样文本做逐条编码，再把结果填入 results JSON 的 details "
                "（格式 [{\"text\":\"用户原文\",\"labels\":\"维度A, 维度B\"}]）后才能导出。"
                "这样总结概览与逐条明细都以 AI 的编码为唯一数据源，准确度最高。"
                "如确需关键词自动兜底，请显式加 --sample_n 0 走全量标注（覆盖率较低）。"
            ),
        }

    sheets_created = []
    keyword_fallback_used = False
    warnings = []

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for idx, question_data in enumerate(results, 1):
            details = question_data.get("details", [])
            dimensions = question_data.get("dimensions", [])
            question_col = question_data.get("question", "")

            # 全量模式且 AI 未填 details → 关键词兜底（带告警）
            if not details and sample_n == 0 and dimensions and source_df is not None and question_col:
                # 尝试精确匹配列名，失败则模糊匹配
                matched_col = question_col if question_col in source_df.columns else None
                if not matched_col:
                    # 提取题号前缀（如 "Q20"）用于模糊匹配
                    q_prefix = question_col.split(".")[0] if "." in question_col else None
                    for c in source_df.columns:
                        if q_prefix and str(c).startswith(q_prefix + "."):
                            matched_col = c
                            break
                        elif question_col in str(c) or str(c) in question_col:
                            matched_col = c
                            break

                if matched_col:
                    extract_result = clean_column_texts(source_df, matched_col)
                    if "error" not in extract_result:
                        all_texts = extract_result.get("texts", [])
                        if all_texts:
                            details = _auto_label_texts(all_texts, dimensions)
                            question_data["details"] = details
                            keyword_fallback_used = True

            # ---- 单一数据源：一律从 details 反算维度统计 + 脚本回填结论数字 ----
            if details:
                rebuilt = _rebuild_dimensions_from_details(details, dimensions)
                question_data["dimensions"] = rebuilt
                question_data["conclusion"] = _augment_conclusion(
                    question_data.get("conclusion", ""), rebuilt, len(details))
                # 未归类占比告警
                other_dim = next((d for d in rebuilt if d["name"] == OTHER_CANON), None)
                if other_dim and len(details) > 0:
                    ratio = other_dim["count"] / len(details)
                    if ratio > 0.2:
                        warnings.append(
                            f"[{question_col}] 未归类占比 {ratio * 100:.1f}%（>20%），"
                            "建议补充维度或重新逐条编码以提升准确度。")

            # ---- 逐条明细附加字段：从原始数据回填 uid / 引擎版本 / 满意度等 ----
            if details and attach_columns and source_df is not None:
                try:
                    mh = _attach_meta_columns(details, source_df, question_col,
                                              attach_columns, attach_headers)
                    question_data["_meta_headers"] = mh
                    if not mh:
                        warnings.append(
                            f"[{question_col}] 附加字段回填失败：未能在原始数据中解析到指定列或文本列。")
                except Exception as e:
                    warnings.append(f"[{question_col}] 附加字段回填异常：{e}")

            # 总结概览
            summary_name = _write_summary_sheet(writer, question_data, idx)
            sheets_created.append(summary_name)

            # 逐条明细
            detail_name = _write_detail_sheet(writer, question_data, idx)
            sheets_created.append(detail_name)

    result = {
        "status": "success",
        "output_path": output_path,
        "sheets": sheets_created,
        "questions_count": len(results),
    }
    if keyword_fallback_used:
        result["keyword_fallback_used"] = True
    if warnings:
        result["warnings"] = warnings

    return result


# ========================================================================= #
#                        CLI 入口
# ========================================================================= #

def main():
    parser = argparse.ArgumentParser(description="问卷文本分析结果导出")
    parser.add_argument("--output_path", default=None,
                        help="输出 Excel 文件路径；不传时根据题目自动生成如 Q5_MC满意原因.xlsx")
    parser.add_argument("--results_file", default=None, help="分析结果 JSON 文件路径")
    parser.add_argument("--results_json", default=None, help="分析结果 JSON 字符串")
    parser.add_argument("--file_path", default=None,
                        help="原始数据文件路径（可选）。当 details 为空时，从此文件提取文本并自动标注")
    parser.add_argument("--sheet_name", default="0", help="工作表名或编号（默认 0）")
    parser.add_argument("--sample_n", type=int, default=300,
                        help="逐条明细行数限制（默认 300，0=全量标注并重新统计维度）")
    parser.add_argument("--attach_columns", nargs="*", default=None,
                        help="回填到逐条明细的原始列名/关键词（支持模糊匹配），如 Y1.uiduid Y2.engine_verengine_ver 'Q1.整体满意度'")
    parser.add_argument("--attach_headers", nargs="*", default=None,
                        help="与 --attach_columns 一一对应的中文表头（可选，缺省自动命名）")
    args = parser.parse_args()

    # 解析 sheet_name
    sheet_name = args.sheet_name
    try:
        sheet_name = int(sheet_name)
    except ValueError:
        pass

    # 读取分析结果
    results = None
    if args.results_file:
        try:
            with open(args.results_file, 'r', encoding='utf-8') as f:
                results = json.load(f)
        except Exception as e:
            print(json.dumps({"error": f"读取 results_file 失败: {e}"}, ensure_ascii=False), file=sys.stderr)
            sys.exit(1)
    elif args.results_json:
        try:
            results = json.loads(args.results_json)
        except json.JSONDecodeError as e:
            print(json.dumps({"error": f"results_json 解析失败: {e}"}, ensure_ascii=False), file=sys.stderr)
            sys.exit(1)
    else:
        print(json.dumps({"error": "请提供 --results_file 或 --results_json"}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)

    if not isinstance(results, list):
        results = [results]

    try:
        output_path = args.output_path or default_output_filename(results)
        result = export_text_report(results, output_path,
                                    file_path=args.file_path,
                                    sheet_name=sheet_name,
                                    sample_n=args.sample_n,
                                    attach_columns=args.attach_columns,
                                    attach_headers=args.attach_headers)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        if result.get("status") == "error":
            sys.exit(1)
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
