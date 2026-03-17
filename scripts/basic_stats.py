#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
问卷分析工具 - 基础统计分析 v2
================================

对问卷数据生成单 Sheet 竖排统计报告，完全对齐原 MCP 的输出格式。

输出布局（在一个 Sheet 中竖向排列所有题目）：
  - 文件头：问卷标题 + 元信息 + 时间范围
  - 隐含问题（Y 开头）：显示"略"
  - 量表题（1-5星 / 0-10分）：选项 / 数量 / 占比 + 总计行
  - NPS 题（0-10分）：额外输出 NPS 表（批评型/被动型/推荐型/净推荐分数）
  - 多选题：选项 / 数量 / 占比 + 总计行（总计=有效样本数）
  - 矩形量表题：子问题 / 1星~5星 / 平均分 / 总计
  - 矩形单选题：子问题 / 各选项 / 总计
  - 填空题：显示"略"

用法:
    python basic_stats.py --file_path "C:/xxx/data.xlsx" [--output_path "C:/xxx/out.xlsx"]
"""

import argparse
import json
import sys
import os
import re
import pandas as pd
import numpy as np
from collections import OrderedDict, defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _styles import (
    Theme, thin_border, header_fill, header_font,
    body_font, even_fill, odd_fill, total_fill, total_font,
    make_fill, ALIGN_CENTER, ALIGN_LEFT, ALIGN_RIGHT,
)
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ========================================================================= #
#                        题目类型检测
# ========================================================================= #

RE_Q = re.compile(r'^(Q\d+)\.')
RE_Y = re.compile(r'^(Y\d+)\.')


def _detect_question_type(col_name: str) -> str:
    """从列名中检测题型标签"""
    cn = str(col_name)
    if "多项填空" in cn:
        return "多项填空题"
    if "矩形量表" in cn or "矩阵量表" in cn:
        return "矩形量表题"
    if "矩形单选" in cn:
        return "矩形单选题"
    if "量表" in cn:
        return "量表题"
    if "多选" in cn:
        return "多选题"
    if "单选" in cn:
        return "单选题"
    if "填空" in cn or "非必填" in cn:
        return "填空题"
    return ""


def _extract_question_label(col_name: str) -> str:
    """从列名提取干净的题标签（去掉子问题部分）"""
    cn = str(col_name).strip()
    if ':' in cn:
        cn = cn.split(':', 1)[0]
    if '：' in cn:
        cn = cn.split('：', 1)[0]
    return cn


def _extract_sub_question(col_name: str) -> str:
    """从列名提取子问题文本"""
    cn = str(col_name).strip()
    for sep in [':', '：']:
        if sep in cn:
            return cn.split(sep, 1)[1].strip()
    return ""


# ========================================================================= #
#                     数据解析：构建题目结构
# ========================================================================= #

def _build_question_structure(df: pd.DataFrame) -> list:
    """
    扫描所有列，构建题目结构列表。
    """
    questions = OrderedDict()
    y_questions = []

    for col in df.columns:
        cn = str(col).strip()

        # 隐含问题
        ym = RE_Y.match(cn)
        if ym:
            yid = ym.group(1)
            if yid not in [q["qid"] for q in y_questions]:
                y_questions.append({
                    "qid": yid, "label": cn, "type": "隐含问题",
                    "columns": [col], "sub_questions": [],
                })
            continue

        # 非 Q 开头 → 跳过
        qm = RE_Q.match(cn)
        if not qm:
            continue

        # 个人信息 → 跳过
        if any(pk in cn for pk in ["姓名", "手机", "电话", "微信", "邮箱", "称呼", "联系到您", "个人信息"]):
            continue

        # 输入文本附属列 → 跳过
        if "输入文本" in cn:
            continue

        qid = qm.group(1)

        if qid not in questions:
            label = _extract_question_label(cn)
            qtype = _detect_question_type(cn)
            sub = _extract_sub_question(cn)
            questions[qid] = {
                "qid": qid, "label": label, "type": qtype,
                "columns": [col], "sub_questions": [sub] if sub else [],
            }
        else:
            questions[qid]["columns"].append(col)
            sub = _extract_sub_question(cn)
            if sub:
                questions[qid]["sub_questions"].append(sub)
            if not questions[qid]["type"]:
                qtype = _detect_question_type(cn)
                if qtype:
                    questions[qid]["type"] = qtype

    result = y_questions + list(questions.values())

    # ---- 后处理1：推断未标注类型 ----
    for q in result:
        if q["type"]:
            continue
        cols = q["columns"]
        if len(cols) == 1:
            col = cols[0]
            series = df[col].dropna()
            if len(series) == 0:
                q["type"] = "填空题"
                continue
            # 如果有子问题（列名含冒号）且含量表语义 → 矩形量表题（单子问题）
            has_sub = bool(q["sub_questions"])
            try:
                nums = pd.to_numeric(series)
                unique = set(nums.unique())
                col_label = str(col)
                scale_kw = ["满意度", "评价", "体验感受", "星表示", "推荐", "愿意"]
                is_scale_context = any(kw in col_label for kw in scale_kw)
                if unique.issubset({0, 1}):
                    q["type"] = "多选题"
                elif has_sub and is_scale_context and 1 <= min(unique) and max(unique) <= 5:
                    q["type"] = "矩形量表题"
                elif is_scale_context and 1 <= min(unique) and max(unique) <= 5 and len(unique) <= 6:
                    q["type"] = "量表题"
                elif is_scale_context and 0 <= min(unique) and max(unique) <= 10 and len(unique) <= 12:
                    q["type"] = "量表题"
                else:
                    q["type"] = "单选题"
            except (ValueError, TypeError):
                unique_rate = series.nunique() / len(series)
                avg_len = series.astype(str).str.len().mean()
                if unique_rate > 0.5 and avg_len > 8:
                    q["type"] = "填空题"
                else:
                    q["type"] = "单选题"
        else:
            # 多列：用整列 dropna 检测（不能只 head，因为条件跳转题前N行可能全为空）
            is_binary = True
            is_all_empty = True
            max_unique_count = 0
            for c in cols:
                if "输入文本" in str(c):
                    continue
                try:
                    vals = set(pd.to_numeric(df[c].dropna()).unique())
                except:
                    vals = set()
                if vals:
                    is_all_empty = False
                    max_unique_count = max(max_unique_count, len(vals))
                if vals and not vals.issubset({0, 1, 0.0, 1.0}):
                    is_binary = False

            label = q["label"]
            scale_matrix_kw = ["满意度", "评价", "体验感受", "星表示", "满意程度"]
            is_scale_matrix = any(kw in label for kw in scale_matrix_kw)

            if is_binary and not is_all_empty:
                q["type"] = "多选题"
            elif is_all_empty and q["sub_questions"]:
                if is_scale_matrix:
                    q["type"] = "矩形量表题"
                else:
                    q["type"] = "矩形单选题"
            elif is_scale_matrix and max_unique_count >= 4 and q["sub_questions"]:
                q["type"] = "矩形量表题"
            elif q["sub_questions"]:
                q["type"] = "矩形单选题"
            else:
                q["type"] = "多选题"

    # ---- 后处理2：在 label 中插入 [题型] 标记 ----
    for q in result:
        label = q["label"]
        qtype = q["type"]
        if re.search(r'\[.+题\]', label):
            continue
        m = re.match(r'^([QY]\d+\.)', label)
        if m and qtype:
            prefix = m.group(1)
            rest = label[len(prefix):]
            q["label"] = f"{prefix}[{qtype}]{rest}"

    return result


# ========================================================================= #
#                     统计计算
# ========================================================================= #

def _stat_scale(df, col, options_label="星"):
    """量表题统计（1-5星 或 0-10分）"""
    series = df[col].dropna()
    try:
        series = pd.to_numeric(series)
    except:
        pass
    total = len(series)
    counts = series.value_counts().sort_index()
    rows = []
    for val, cnt in counts.items():
        val_int = int(val)
        if options_label == "星":
            label = f"{val_int}星"
        else:
            label = str(val_int)
        pct = f"{cnt / total * 100:.2f}%" if total > 0 else "0%"
        rows.append({"选项": label, "数量": int(cnt), "占比": pct})
    rows.append({"选项": "总计（人）", "数量": total, "占比": "100.0%"})
    return rows, total


def _stat_nps(df, col):
    """NPS 净推荐值计算"""
    series = df[col].dropna()
    try:
        series = pd.to_numeric(series)
    except:
        return None
    total = len(series)
    if total == 0:
        return None
    detractors = int(((series >= 0) & (series <= 6)).sum())
    passives = int(((series >= 7) & (series <= 8)).sum())
    promoters = int(((series >= 9) & (series <= 10)).sum())
    d_pct = round(detractors / total * 100, 2)
    p_pct = round(passives / total * 100, 2)
    pr_pct = round(promoters / total * 100, 2)
    nps = round(pr_pct - d_pct, 2)
    return {
        "detractors": f"{d_pct}（{detractors}）",
        "passives": f"{p_pct}（{passives}）",
        "promoters": f"{pr_pct}（{promoters}）",
        "nps": nps, "total": total,
    }


def _stat_multi_choice(df, cols, root_label):
    """多选题统计"""
    valid_cols = [c for c in cols if "输入文本" not in str(c)]
    if not valid_cols:
        return [], 0
    valid_mask = pd.DataFrame()
    for c in valid_cols:
        valid_mask[c] = (df[c] == 1)
    valid_total = int(valid_mask.any(axis=1).sum())

    rows = []
    for c in valid_cols:
        cn = str(c).strip()
        sub = _extract_sub_question(cn)
        if not sub:
            parts = cn.split(root_label, 1)
            sub = parts[1].strip() if len(parts) > 1 else cn
        cnt = int((df[c] == 1).sum())
        pct = f"{cnt / valid_total * 100:.2f}%" if valid_total > 0 else "0.0%"
        rows.append({"选项": sub, "数量": cnt, "占比": pct})
    rows.append({"选项": "总计（人）", "数量": valid_total, "占比": "100.0%"})
    return rows, valid_total


def _stat_matrix_scale(df, cols, subs):
    """矩形量表题统计"""
    result_rows = []
    for c, sub_text in zip(cols, subs):
        if "输入文本" in str(c):
            continue
        series = df[c].dropna()
        try:
            series = pd.to_numeric(series)
        except:
            continue
        total = len(series)
        star_counts = {}
        for s in range(1, 6):
            star_counts[s] = int((series == s).sum())
        avg = round(series.mean(), 2) if total > 0 else 0
        result_rows.append({"sub": sub_text, "stars": star_counts, "avg": avg, "total": total})
    return result_rows


def _stat_matrix_single(df, cols, subs):
    """矩形单选题统计"""
    all_options = OrderedDict()
    data_rows = []
    for c, sub_text in zip(cols, subs):
        if "输入文本" in str(c):
            continue
        series = df[c].dropna()
        total = len(series)
        counts = series.value_counts()
        row_data = {"sub": sub_text, "counts": {}, "total": total}
        for opt, cnt in counts.items():
            opt_str = str(opt)
            all_options[opt_str] = True
            row_data["counts"][opt_str] = int(cnt)
        data_rows.append(row_data)
    return data_rows, list(all_options.keys())


def _stat_single_choice(df, col):
    """单选题统计"""
    series = df[col].dropna()
    total = len(series)
    counts = series.value_counts()
    try:
        sorted_idx = sorted(counts.index, key=lambda x: float(x))
    except:
        sorted_idx = list(counts.index)
    counts = counts.reindex(sorted_idx)
    rows = []
    for opt, cnt in counts.items():
        pct = f"{cnt / total * 100:.2f}%" if total > 0 else "0%"
        rows.append({"选项": str(opt), "数量": int(cnt), "占比": pct})
    rows.append({"选项": "总计（人）", "数量": total, "占比": "100.0%"})
    return rows, total


# ========================================================================= #
#                     Excel 写入
# ========================================================================= #

def _write_stat_report(df, questions, output_path,
                       questionnaire_title="", meta_info="", date_range=""):
    """将统计结果写入单 Sheet Excel"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    border = thin_border()
    row = 1

    # ---- 文件头 ----
    if questionnaire_title:
        ws.cell(row=row, column=1, value=questionnaire_title)
        ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=14, bold=True)
        row += 1
    row += 1
    if meta_info:
        ws.cell(row=row, column=1, value=meta_info)
        ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=9, color="666666")
        row += 1
    row += 1
    if date_range:
        ws.cell(row=row, column=1, value=date_range)
        ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=9, color="666666")
        row += 1
    row += 1

    # ---- 逐题输出 ----
    for q in questions:
        qid = q["qid"]
        label = q["label"]
        qtype = q["type"]
        cols = q["columns"]
        subs = q["sub_questions"]

        # ============ 隐含问题 / 填空题 / 多项填空题 ============
        if qtype in ("隐含问题", "填空题", "多项填空题"):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=10, bold=True)
            ws.cell(row=row, column=3, value="略")
            ws.cell(row=row, column=3).font = Font(name=Theme.FONT_NAME, size=9, color="999999")
            row += 2
            continue

        # ============ 量表题（1-5星 或 0-10分）============
        if qtype == "量表题" and len(cols) == 1:
            col = cols[0]
            series = df[col].dropna()
            try:
                vals = pd.to_numeric(series).unique()
                max_val = max(vals) if len(vals) > 0 else 5
            except:
                max_val = 5
            is_nps = (max_val >= 10)

            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=10, bold=True)
            row += 1

            if is_nps:
                stat_rows, total = _stat_scale(df, col, options_label="分")
            else:
                stat_rows, total = _stat_scale(df, col, options_label="星")

            # 表头
            for ci, h in enumerate(["选项", "数量", "占比"], 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True)
                cell.fill = make_fill("D9E2F3")
                cell.border = border
            row += 1

            for sr in stat_rows:
                is_total = "总计" in str(sr["选项"])
                ws.cell(row=row, column=1, value=sr["选项"]).border = border
                ws.cell(row=row, column=2, value=sr["数量"]).border = border
                ws.cell(row=row, column=3, value=sr["占比"]).border = border
                f = Font(name=Theme.FONT_NAME, size=9, bold=is_total)
                for ci in range(1, 4):
                    ws.cell(row=row, column=ci).font = f
                row += 1

            # NPS 计算
            if is_nps:
                nps_data = _stat_nps(df, col)
                if nps_data:
                    ws.cell(row=row, column=1, value="NPS净推荐值")
                    ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=9, bold=True, color="C00000")
                    row += 1
                    for ci, h in enumerate(["问题", "批评型", "被动型", "推荐型", "净推荐分数NPS"], 1):
                        cell = ws.cell(row=row, column=ci, value=h)
                        cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True)
                        cell.fill = make_fill("FCE4D6")
                        cell.border = border
                    row += 1
                    ws.cell(row=row, column=1, value=qid).border = border
                    ws.cell(row=row, column=2, value=nps_data["detractors"]).border = border
                    ws.cell(row=row, column=3, value=nps_data["passives"]).border = border
                    ws.cell(row=row, column=4, value=nps_data["promoters"]).border = border
                    ws.cell(row=row, column=5, value=nps_data["nps"]).border = border
                    for ci in range(1, 6):
                        ws.cell(row=row, column=ci).font = Font(name=Theme.FONT_NAME, size=9)
                    row += 1
                    ws.cell(row=row, column=1, value=f"受访人数：{nps_data['total']}")
                    ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=8, color="666666")
                    row += 1
            row += 1
            continue

        # ============ 单选题 ============
        if qtype == "单选题" and len(cols) == 1:
            col = cols[0]
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=10, bold=True)
            row += 1

            stat_rows, total = _stat_single_choice(df, col)

            for ci, h in enumerate(["选项", "数量", "占比"], 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True)
                cell.fill = make_fill("D9E2F3")
                cell.border = border
            row += 1

            for sr in stat_rows:
                is_total = "总计" in str(sr["选项"])
                ws.cell(row=row, column=1, value=sr["选项"]).border = border
                ws.cell(row=row, column=2, value=sr["数量"]).border = border
                ws.cell(row=row, column=3, value=sr["占比"]).border = border
                f = Font(name=Theme.FONT_NAME, size=9, bold=is_total)
                for ci in range(1, 4):
                    ws.cell(row=row, column=ci).font = f
                row += 1
            row += 1
            continue

        # ============ 多选题 ============
        if qtype == "多选题":
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=10, bold=True)
            row += 1

            stat_rows, valid_total = _stat_multi_choice(df, cols, label)

            for ci, h in enumerate(["选项", "数量", "占比"], 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True)
                cell.fill = make_fill("D9E2F3")
                cell.border = border
            row += 1

            for sr in stat_rows:
                is_total = "总计" in str(sr["选项"])
                ws.cell(row=row, column=1, value=sr["选项"]).border = border
                ws.cell(row=row, column=2, value=sr["数量"]).border = border
                ws.cell(row=row, column=3, value=sr["占比"]).border = border
                f = Font(name=Theme.FONT_NAME, size=9, bold=is_total)
                for ci in range(1, 4):
                    ws.cell(row=row, column=ci).font = f
                row += 1
            row += 1
            continue

        # ============ 矩形量表题 ============
        if qtype in ("矩形量表题", "矩阵量表题"):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=10, bold=True)
            row += 1

            valid_cols = [c for c in cols if "输入文本" not in str(c)]
            valid_subs = [_extract_sub_question(c) for c in valid_cols]
            matrix_rows = _stat_matrix_scale(df, valid_cols, valid_subs)

            if matrix_rows:
                m_headers = ["子问题", "1星", "2星", "3星", "4星", "5星", "平均分", "总计"]
                for ci, h in enumerate(m_headers, 1):
                    cell = ws.cell(row=row, column=ci, value=h)
                    cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True)
                    cell.fill = make_fill("D9E2F3")
                    cell.border = border
                row += 1
                for mr in matrix_rows:
                    ws.cell(row=row, column=1, value=mr["sub"]).border = border
                    ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=9)
                    for si in range(1, 6):
                        ws.cell(row=row, column=si + 1, value=mr["stars"].get(si, 0)).border = border
                        ws.cell(row=row, column=si + 1).font = Font(name=Theme.FONT_NAME, size=9)
                    ws.cell(row=row, column=7, value=mr["avg"]).border = border
                    ws.cell(row=row, column=7).font = Font(name=Theme.FONT_NAME, size=9)
                    ws.cell(row=row, column=8, value=mr["total"]).border = border
                    ws.cell(row=row, column=8).font = Font(name=Theme.FONT_NAME, size=9)
                    row += 1
            else:
                # 全空数据 → 只列出子问题名
                m_headers = ["子问题", "1星", "2星", "3星", "4星", "5星", "平均分", "总计"]
                for ci, h in enumerate(m_headers, 1):
                    cell = ws.cell(row=row, column=ci, value=h)
                    cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True)
                    cell.fill = make_fill("D9E2F3")
                    cell.border = border
                row += 1
                for sub_text in valid_subs:
                    ws.cell(row=row, column=1, value=sub_text).border = border
                    ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=9)
                    row += 1
            row += 1
            continue

        # ============ 矩形单选题 ============
        if qtype == "矩形单选题":
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=10, bold=True)
            row += 1

            valid_cols = [c for c in cols if "输入文本" not in str(c)]
            valid_subs = [_extract_sub_question(c) for c in valid_cols]
            data_rows, all_options = _stat_matrix_single(df, valid_cols, valid_subs)

            if data_rows:
                m_headers = ["子问题"] + all_options + ["总计"]
                for ci, h in enumerate(m_headers, 1):
                    cell = ws.cell(row=row, column=ci, value=h)
                    cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True)
                    cell.fill = make_fill("D9E2F3")
                    cell.border = border
                row += 1
                for dr in data_rows:
                    ws.cell(row=row, column=1, value=dr["sub"]).border = border
                    ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=9)
                    for oi, opt in enumerate(all_options):
                        val = dr["counts"].get(opt, 0)
                        ws.cell(row=row, column=oi + 2, value=val).border = border
                        ws.cell(row=row, column=oi + 2).font = Font(name=Theme.FONT_NAME, size=9)
                    ws.cell(row=row, column=len(all_options) + 2, value=dr["total"]).border = border
                    ws.cell(row=row, column=len(all_options) + 2).font = Font(name=Theme.FONT_NAME, size=9)
                    row += 1
            row += 1
            continue

        # ============ 其他未分类 → 当单选处理 ============
        if len(cols) == 1:
            col = cols[0]
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(name=Theme.FONT_NAME, size=10, bold=True)
            row += 1
            stat_rows, total = _stat_single_choice(df, col)
            for ci, h in enumerate(["选项", "数量", "占比"], 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = Font(name=Theme.FONT_NAME, size=9, bold=True)
                cell.fill = make_fill("D9E2F3")
                cell.border = border
            row += 1
            for sr in stat_rows:
                is_total = "总计" in str(sr["选项"])
                ws.cell(row=row, column=1, value=sr["选项"]).border = border
                ws.cell(row=row, column=2, value=sr["数量"]).border = border
                ws.cell(row=row, column=3, value=sr["占比"]).border = border
                f = Font(name=Theme.FONT_NAME, size=9, bold=is_total)
                for ci in range(1, 4):
                    ws.cell(row=row, column=ci).font = f
                row += 1
            row += 1
            continue

        row += 1

    # ---- 列宽 ----
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    for ci in range(4, 12):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.sheet_view.showGridLines = False
    wb.save(output_path)


# ========================================================================= #
#                        JSON 摘要
# ========================================================================= #

def _generate_summary(df, questions):
    """生成 JSON 摘要"""
    summary = {"total_rows": len(df), "questions": []}
    for q in questions:
        if q["type"] in ("隐含问题", "填空题", "多项填空题"):
            continue
        qi = {"qid": q["qid"], "label": q["label"], "type": q["type"]}
        if q["type"] in ("量表题", "单选题") and len(q["columns"]) == 1:
            col = q["columns"][0]
            series = df[col].dropna()
            top3 = series.value_counts().head(3)
            total = len(series)
            qi["valid_count"] = total
            qi["top3"] = {str(k): f"{v} ({v/total*100:.1f}%)" for k, v in top3.items()} if total > 0 else {}
        summary["questions"].append(qi)
    return summary


# ========================================================================= #
#                        主函数
# ========================================================================= #

def run_basic_stats(file_path: str, sheet_name=0, output_path: str = None) -> dict:
    """执行基础统计分析"""
    ext = file_path.rsplit('.', 1)[-1].lower()
    if ext == 'csv':
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    if output_path is None:
        base = os.path.splitext(file_path)[0]
        output_path = f"{base}_基础统计.xlsx"

    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except PermissionError:
            raise PermissionError(f"请关闭正在使用的文件：{output_path}")

    questions = _build_question_structure(df)

    date_range = ""
    for col in df.columns:
        if "开始答题时间" in str(col):
            try:
                dates = pd.to_datetime(df[col].dropna())
                start = dates.min().strftime("%Y-%m-%d")
                end = dates.max().strftime("%Y-%m-%d")
                date_range = f"{start}至{end}"
            except:
                pass
            break

    _write_stat_report(df, questions, output_path, date_range=date_range)

    summary = _generate_summary(df, questions)
    summary["output_path"] = output_path
    return summary


def main():
    parser = argparse.ArgumentParser(description="问卷基础统计分析")
    parser.add_argument("--file_path", required=True, help="数据文件的绝对路径")
    parser.add_argument("--sheet_name", default="0", help="工作表名称或编号")
    parser.add_argument("--output_path", default=None, help="输出 Excel 路径")
    args = parser.parse_args()

    sheet_name = args.sheet_name
    try:
        sheet_name = int(sheet_name)
    except ValueError:
        pass

    try:
        result = run_basic_stats(args.file_path, sheet_name, args.output_path)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()