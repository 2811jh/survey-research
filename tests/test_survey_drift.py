import importlib.util
from pathlib import Path
import math
import pandas as pd
import subprocess, sys, json, os

MODULE_PATH = Path(__file__).resolve().parents[1] / "scripts" / "survey_drift.py"
spec = importlib.util.spec_from_file_location("survey_drift", MODULE_PATH)
survey_drift = importlib.util.module_from_spec(spec)
spec.loader.exec_module(survey_drift)


def test_multi_choice_label_recovers_stem():
    root = "Q6."
    subcols = [
        "Q6.请问是什么原因让您回到了《我的世界》？:游戏版本更新",
        "Q6.请问是什么原因让您回到了《我的世界》？:其他",
    ]
    assert survey_drift._multi_choice_label(root, subcols) == "Q6.请问是什么原因让您回到了《我的世界》？"


def test_multi_choice_label_fullwidth_colon_and_fallback():
    assert survey_drift._multi_choice_label("Q9.", ["Q9.你为何离开？：选项A"]) == "Q9.你为何离开？"
    assert survey_drift._multi_choice_label("Q9.", ["Q9.没有冒号的列"]) == "Q9."


def test_two_prop_z_no_diff_gives_high_p():
    z, p = survey_drift.two_prop_z(50, 100, 50, 100)
    assert abs(z) < 1e-9
    assert p > 0.99


def test_two_prop_z_big_diff_is_significant():
    z, p = survey_drift.two_prop_z(80, 100, 40, 100)
    assert p < 0.01
    assert z > 0


def test_labels():
    dt = pd.Timestamp("2026-04-06")  # 周一, ISO 第15周
    assert survey_drift.week_label(dt) == "第15周（4.6-4.12）"
    assert survey_drift.month_label(dt) == "26年4月"
    assert survey_drift.day_label(dt) == "2026-04-06"


def test_bucketize_orders_chronologically():
    s = pd.to_datetime(pd.Series([
        "2026-04-06", "2026-04-13", "2026-04-06", "2026-04-20"
    ]))
    labels, ordered = survey_drift.bucketize(s, "week")
    assert ordered == ["第15周（4.6-4.12）", "第16周（4.13-4.19）", "第17周（4.20-4.26）"]
    assert list(labels) == [
        "第15周（4.6-4.12）", "第16周（4.13-4.19）",
        "第15周（4.6-4.12）", "第17周（4.20-4.26）",
    ]


def test_compute_nps():
    # 5 推荐者(9-10), 3 贬损者(0-6), 2 中立(7-8)
    s = pd.Series([10, 10, 9, 9, 9, 7, 8, 0, 3, 6])
    r = survey_drift.compute_nps(s)
    assert r["n"] == 10
    assert r["promoter"] == 5
    assert r["detractor"] == 3
    assert round(r["nps"], 1) == 20.0  # (5-3)/10 * 100


def test_evaluate_drift_double_threshold():
    assert survey_drift.evaluate_drift(6.0, 0.02, "pp") is True
    assert survey_drift.evaluate_drift(3.0, 0.01, "pp") is False
    assert survey_drift.evaluate_drift(8.0, 0.20, "pp") is False
    assert survey_drift.evaluate_drift(0.15, 0.03, "mean") is True
    assert survey_drift.evaluate_drift(0.05, 0.03, "mean") is False


def _demo_df():
    return pd.DataFrame({
        "结束答题时间": pd.to_datetime([
            "2026-04-06", "2026-04-06", "2026-04-13", "2026-04-13"]),
        "Q1.整体满意度": [5, 4, 3, 3],
        "Q7.活动评价": ["满意", "满意", "一般", "满意"],
    })


def test_single_choice_props():
    df = _demo_df()
    labels, ordered = survey_drift.bucketize(df["结束答题时间"], "week")
    by_bucket, sizes = survey_drift.single_choice_props(df, "Q7.活动评价", labels, ordered)
    b0, b1 = ordered
    assert sizes[b0] == 2 and sizes[b1] == 2
    assert round(by_bucket[b0]["满意"], 3) == 1.0
    assert round(by_bucket[b1]["满意"], 3) == 0.5


def test_scale_means():
    df = _demo_df()
    labels, ordered = survey_drift.bucketize(df["结束答题时间"], "week")
    by_bucket, sizes = survey_drift.scale_means(df, "Q1.整体满意度", labels, ordered)
    b0, b1 = ordered
    assert round(by_bucket[b0], 2) == 4.5
    assert round(by_bucket[b1], 2) == 3.0


def test_multi_choice_rates():
    df = pd.DataFrame({
        "结束答题时间": pd.to_datetime(["2026-04-06", "2026-04-06"]),
        "Q9.喜欢的模式:生存": ["生存", None],
        "Q9.喜欢的模式:创造": ["创造", "创造"],
    })
    labels, ordered = survey_drift.bucketize(df["结束答题时间"], "week")
    subcols = ["Q9.喜欢的模式:生存", "Q9.喜欢的模式:创造"]
    by_bucket, sizes = survey_drift.multi_choice_rates(df, subcols, "Q9.", labels, ordered)
    b0 = ordered[0]
    assert round(by_bucket[b0]["生存"], 2) == 0.5
    assert round(by_bucket[b0]["创造"], 2) == 1.0


def test_multi_choice_base_excludes_unanswered():
    """多选题样本量=答过此题的人(至少勾选一项)，跳过未触达该题的空行(逻辑门控题)。"""
    df = pd.DataFrame({
        "结束答题时间": pd.to_datetime(["2026-04-06"] * 4),
        "Q25.MOBA:王者荣耀": [1, 1, None, None],   # 前2人答过；后2人未触达
        "Q25.MOBA:英雄联盟": [1, 0, None, None],
    })
    labels, ordered = survey_drift.bucketize(df["结束答题时间"], "week")
    subcols = ["Q25.MOBA:王者荣耀", "Q25.MOBA:英雄联盟"]
    by_bucket, sizes = survey_drift.multi_choice_rates(df, subcols, "Q25.", labels, ordered)
    b0 = ordered[0]
    assert sizes[b0] == 2                           # 基数=2（非 4）
    assert round(by_bucket[b0]["王者荣耀"], 2) == 1.0   # 2/2
    assert round(by_bucket[b0]["英雄联盟"], 2) == 0.5   # 1/2


def test_adjacent_prop_tests_flags_drift():
    # b0 满意=100%(n=60), b1 满意=50%(n=60) → 相邻期 z 检验显著且 >5pp
    by_bucket = {"b0": {"满意": 1.0}, "b1": {"满意": 0.5}}
    sizes = {"b0": 60, "b1": 60}
    ordered = ["b0", "b1"]  # 旧→新
    res = survey_drift.adjacent_prop_tests(by_bucket, sizes, ordered, min_n=30)
    row = [r for r in res if r["option"] == "满意"][0]
    assert row["from"] == "b0" and row["to"] == "b1"
    assert row["significant"] is True
    assert row["drift"] is True
    assert row["direction"] == "down"


def test_identify_metric_cols():
    single = ["Q1.请问您对本赛季的满意度如何？", "Q51.您有多大可能将本游戏推荐给朋友？", "Q3.性别"]
    nps, sat = survey_drift.identify_metric_cols(single)
    assert nps == "Q51.您有多大可能将本游戏推荐给朋友？"
    assert "Q1.请问您对本赛季的满意度如何？" in sat


def test_build_findings_structure(tmp_path):
    df = pd.DataFrame({
        "结束答题时间": pd.to_datetime(
            ["2026-04-06"] * 40 + ["2026-04-13"] * 40),
        "Q1.整体满意度": [5] * 40 + [3] * 40,
        "Q7.活动评价（单选）": (["满意"] * 40) + (["满意"] * 20 + ["一般"] * 20),
    })
    classification = {
        "single_choice": ["Q1.整体满意度", "Q7.活动评价（单选）"],
        "multi_choice": {}, "matrix_scale": {}, "text": [], "meta": ["结束答题时间"],
    }
    findings = survey_drift.build_findings(
        df, classification, granularity="week", time_col="结束答题时间",
        nps_col=None, satisfaction_cols=["Q1.整体满意度"], min_n=30)
    assert findings["granularity"] == "week"
    assert len(findings["buckets"]) == 2
    q_names = [q["question"] for q in findings["questions"]]
    assert "Q7.活动评价（单选）" in q_names
    assert any(m["type"] == "satisfaction_mean" for m in findings["metrics"])


def test_build_findings_auto_includes_five_point_scale_metric():
    """五点量表单选题即使未被关键词识别为满意度，也应自动进指标做均分显著性检验。"""
    scale = [5, 4, 3, 2, 1] * 8  # 40 行，取值覆盖 1~5
    df = pd.DataFrame({
        "结束答题时间": pd.to_datetime(["2026-04-06"] * 40 + ["2026-04-13"] * 40),
        "Q4.回归意愿": scale + scale,
        "Q3.性别": (["男"] * 20 + ["女"] * 20) * 2,
    })
    classification = {
        "single_choice": ["Q4.回归意愿", "Q3.性别"],
        "multi_choice": {}, "matrix_scale": {}, "text": [], "meta": ["结束答题时间"],
    }
    findings = survey_drift.build_findings(
        df, classification, granularity="week", time_col="结束答题时间",
        nps_col=None, satisfaction_cols=[], min_n=30)  # 不传满意度关键词
    metric_cols = [m["source_col"] for m in findings["metrics"]]
    assert "Q4.回归意愿" in metric_cols           # 五点量表被自动纳入
    assert "Q3.性别" not in metric_cols            # 非量表单选不纳入
    m = next(m for m in findings["metrics"] if m["source_col"] == "Q4.回归意愿")
    assert m["type"] == "satisfaction_mean"
    assert len(m["adjacent"]) == 1                # 相邻期做了均分显著性检验
    assert "Q4.回归意愿" in findings["satisfaction_cols"]


def test_analyze_cli_end_to_end(tmp_path):
    csv = tmp_path / "demo.csv"
    pd.DataFrame({
        "结束答题时间": (["2026-04-06 10:00:00"] * 40 + ["2026-04-13 10:00:00"] * 40),
        "Q1.整体满意度": [5] * 40 + [3] * 40,
        "Q51.您有多大可能将本游戏推荐给朋友？": [10] * 40 + [5] * 40,
    }).to_csv(csv, index=False, encoding="utf-8-sig")
    out = tmp_path / "findings.json"
    r = subprocess.run(
        [sys.executable, str(MODULE_PATH), "analyze",
         "--file_path", str(csv), "--granularity", "week",
         "--findings_out", str(out)],
        capture_output=True, text=True, encoding="utf-8")
    assert r.returncode == 0, r.stderr
    payload = json.loads(r.stdout)
    assert payload["status"] == "success"
    assert out.exists()
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["granularity"] == "week"


def test_analyze_cli_multi_choice_json_safe(tmp_path):
    csv = tmp_path / "multi.csv"
    pd.DataFrame({
        "结束答题时间": (["2026-04-06 10:00:00"] * 40 + ["2026-04-13 10:00:00"] * 40),
        "Q1.整体满意度": [5] * 40 + [3] * 40,
        "Q9.喜欢的模式:生存": (["生存"] * 40) + (["生存"] * 10 + [None] * 30),
        "Q9.喜欢的模式:创造": (["创造"] * 40) + (["创造"] * 40),
    }).to_csv(csv, index=False, encoding="utf-8-sig")
    out = tmp_path / "findings.json"
    r = subprocess.run(
        [sys.executable, str(MODULE_PATH), "analyze",
         "--file_path", str(csv), "--granularity", "week",
         "--satisfaction_cols", "Q1.整体满意度",
         "--findings_out", str(out)],
        capture_output=True, text=True, encoding="utf-8")
    assert r.returncode == 0, r.stderr
    payload = json.loads(r.stdout)
    assert payload["status"] == "success"
    data = json.loads(out.read_text(encoding="utf-8"))
    # 多选题应作为一道 question 出现
    q_types = [q["type"] for q in data["questions"]]
    assert "multi_choice" in q_types
    # 每个 drift 字段应为合法 JSON bool
    for q in data["questions"]:
        for t in q["adjacent_option_tests"]:
            assert isinstance(t["drift"], bool)


from openpyxl import load_workbook


def test_export_creates_four_sheets(tmp_path):
    findings = {
        "granularity": "week", "time_col": "结束答题时间",
        "buckets": ["第15周（4.6-4.12）", "第16周（4.13-4.19）"],
        "bucket_sizes": {"第15周（4.6-4.12）": 40, "第16周（4.13-4.19）": 40},
        "low_n_buckets": [],
        "metrics": [{
            "name": "Q1.整体满意度 均分", "type": "satisfaction_mean", "source_col": "Q1.整体满意度",
            "by_bucket": {"第15周（4.6-4.12）": 4.5, "第16周（4.13-4.19）": 3.0},
            "sizes": {"第15周（4.6-4.12）": 38, "第16周（4.13-4.19）": 40},
            "adjacent": [{"from": "第15周（4.6-4.12）", "to": "第16周（4.13-4.19）",
                          "delta": -1.5, "test": "t_test", "p": 0.001,
                          "significant": True, "drift": True, "low_n": False, "direction": "down"}],
        }],
        "questions": [{
            "question": "Q7.活动评价（单选）", "type": "single_choice",
            "options": ["满意", "一般"],
            "by_bucket": {"第15周（4.6-4.12）": {"满意": 1.0, "一般": 0.0},
                          "第16周（4.13-4.19）": {"满意": 0.5, "一般": 0.5}},
            "sizes": {"第15周（4.6-4.12）": 40, "第16周（4.13-4.19）": 40},
            "overall_test": {"test": "chi_square", "p": 0.001, "significant": True},
            "adjacent_option_tests": [{"option": "满意", "from": "第15周（4.6-4.12）",
                "to": "第16周（4.13-4.19）", "delta_pp": -50.0, "test": "two_prop_z",
                "p": 0.001, "significant": True, "drift": True, "low_n": False, "direction": "down"}],
            "drift": True, "low_n": False,
        }],
        "nps_col": None, "satisfaction_cols": ["Q1.整体满意度"],
    }
    conclusions = {"Q7.活动评价（单选）": "满意占比从100%骤降至50%，显著恶化，需排查活动体验。"}
    out = tmp_path / "report.xlsx"
    r = survey_drift.export_excel(findings, conclusions, str(out))
    assert r["status"] == "success"
    wb = load_workbook(out)
    assert "📊 指标总览" in wb.sheetnames
    assert "📈 逐题异动明细" in wb.sheetnames
    assert "⚠️ 异动汇总" in wb.sheetnames
    assert "ℹ️ 方法与样本" in wb.sheetnames
    # 指标总览：每个指标下方紧跟「样本量」行，格式=整数，值=该指标各期有效 n
    ws1 = wb["📊 指标总览"]
    assert ws1.cell(2, 1).value == "Q1.整体满意度 均分"
    assert ws1.cell(3, 1).value == "样本量"
    assert ws1.cell(3, 2).value == 38 and ws1.cell(3, 3).value == 40


def test_detail_sheet_question_order_by_qnum(tmp_path):
    """明细表题目按题号 Q1→Q35 升序排列（不受 findings 中原始顺序影响）。"""
    def _q(name, opt="A"):
        return {"question": name, "type": "single_choice", "question_label": name,
                "options": [opt], "by_bucket": {"W1": {opt: 1.0}}, "sizes": {"W1": 50},
                "overall_test": None, "adjacent_option_tests": [], "drift": False, "low_n": False}
    findings = {
        "granularity": "week", "time_col": "结束答题时间",
        "buckets": ["W1"], "bucket_sizes": {"W1": 50}, "low_n_buckets": [], "metrics": [],
        # 乱序输入：Q35, Q2, Q10
        "questions": [_q("Q35.职业"), _q("Q2.渠道"), _q("Q10.时长")],
        "nps_col": None, "satisfaction_cols": [],
    }
    out = tmp_path / "report.xlsx"
    survey_drift.export_excel(findings, {}, str(out))
    ws = load_workbook(out)["📈 逐题异动明细"]
    titles = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value]
    assert titles == ["Q2.渠道", "Q10.时长", "Q35.职业"]


def test_detail_sheet_overall_column_and_sample_row(tmp_path):
    """明细表 C 列=整体加权占比基线；每题末行=样本量（整体N + 各期n）。"""
    findings = {
        "granularity": "week", "time_col": "结束答题时间",
        "buckets": ["W1", "W2"],
        "bucket_sizes": {"W1": 60, "W2": 40},
        "low_n_buckets": [], "metrics": [],
        "questions": [{
            "question": "Q7.活动评价（单选）", "type": "single_choice",
            "question_label": "Q7.活动评价（单选）",
            "options": ["满意", "一般"],
            "by_bucket": {"W1": {"满意": 1.0, "一般": 0.0},
                          "W2": {"满意": 0.5, "一般": 0.5}},
            "sizes": {"W1": 60, "W2": 40},
            "overall_test": {"test": "chi_square", "p": 0.001, "significant": True},
            "adjacent_option_tests": [{"option": "满意", "from": "W1", "to": "W2",
                "delta_pp": -50.0, "test": "two_prop_z", "p": 0.001,
                "significant": True, "drift": True, "low_n": False, "direction": "down"}],
            "drift": True, "low_n": False,
        }],
        "nps_col": None, "satisfaction_cols": [],
    }
    out = tmp_path / "report.xlsx"
    survey_drift.export_excel(findings, {}, str(out))
    ws = load_workbook(out)["📈 逐题异动明细"]
    # 表头：A题目 B选项 C整体 D..桶 …
    assert [ws.cell(1, c).value for c in range(1, 4)] == ["题目", "选项", "整体"]
    assert ws.cell(1, 4).value == "W1" and ws.cell(1, 5).value == "W2"
    # C 列整体占比（加权）：满意 = (60*1.0 + 40*0.5)/100 = 0.8
    assert round(ws.cell(2, 3).value, 4) == 0.8
    # 每题末行 = 样本量：整体 100，W1=60，W2=40
    last = ws.max_row
    assert ws.cell(last, 2).value == "样本量"
    assert ws.cell(last, 3).value == 100
    assert ws.cell(last, 4).value == 60 and ws.cell(last, 5).value == 40


def test_five_point_scale_helpers():
    pairs = survey_drift._five_point_scale_opts(["1", "2", "3", "4", "5"])
    assert pairs is not None
    assert dict(pairs) == {"1": 1, "2": 2, "3": 3, "4": 4, "5": 5}
    # NPS 0~10、二元、文本 → 非五点量表
    assert survey_drift._five_point_scale_opts([str(i) for i in range(11)]) is None
    assert survey_drift._five_point_scale_opts(["1", "5"]) is None
    assert survey_drift._five_point_scale_opts(["满意", "一般"]) is None
    # 加权满意度 = Σ(分值×占比)
    props = {"1": 0.0, "2": 0.0, "3": 0.0, "4": 0.5, "5": 0.5}
    assert round(survey_drift._weighted_satisfaction(props, pairs), 2) == 4.5


def test_detail_sheet_weighted_satisfaction_row(tmp_path):
    """五点量表题在样本量行下自动加一行加权满意度（1~5 均分）。"""
    findings = {
        "granularity": "week", "time_col": "结束答题时间",
        "buckets": ["W1", "W2"], "bucket_sizes": {"W1": 40, "W2": 60},
        "low_n_buckets": [], "metrics": [],
        "questions": [{
            "question": "Q1.整体满意度", "type": "single_choice",
            "question_label": "Q1.整体满意度",
            "options": ["1", "2", "3", "4", "5"],
            "by_bucket": {
                "W1": {"1": 0.0, "2": 0.0, "3": 0.0, "4": 0.0, "5": 1.0},  # 均分 5.0
                "W2": {"1": 0.0, "2": 0.0, "3": 0.0, "4": 1.0, "5": 0.0},  # 均分 4.0
            },
            "sizes": {"W1": 40, "W2": 60},
            "overall_test": None, "adjacent_option_tests": [],
            "drift": False, "low_n": False,
        }],
        "nps_col": None, "satisfaction_cols": [],
    }
    out = tmp_path / "report.xlsx"
    survey_drift.export_excel(findings, {}, str(out))
    ws = load_workbook(out)["📈 逐题异动明细"]
    last = ws.max_row
    assert ws.cell(last, 2).value == "加权满意度"
    # 整体 = (40*5 + 60*4)/100 = 4.4；W1=5.0；W2=4.0
    assert round(ws.cell(last, 3).value, 2) == 4.4
    assert round(ws.cell(last, 4).value, 2) == 5.0
    assert round(ws.cell(last, 5).value, 2) == 4.0
    # 上一行应为样本量
    assert ws.cell(last - 1, 2).value == "样本量"


def test_detail_sheet_sorts_choice_by_overall(tmp_path):
    """单选/多选按「整体」占比降序排；五点量表题保持 1~5 自然序。"""
    findings = {
        "granularity": "week", "time_col": "结束答题时间",
        "buckets": ["W1"], "bucket_sizes": {"W1": 100},
        "low_n_buckets": [], "metrics": [],
        "questions": [
            {  # 多选题：整体 C > A > B，应排序为 C, A, B
                "question": "Q9.喜欢的模式", "type": "multi_choice",
                "question_label": "Q9.喜欢的模式",
                "options": ["A", "B", "C"],
                "by_bucket": {"W1": {"A": 0.5, "B": 0.2, "C": 0.8}},
                "sizes": {"W1": 100},
                "overall_test": None, "adjacent_option_tests": [],
                "drift": False, "low_n": False,
            },
            {  # 五点量表：保持 1..5 顺序，不因占比重排
                "question": "Q1.满意度", "type": "single_choice",
                "question_label": "Q1.满意度",
                "options": ["1", "2", "3", "4", "5"],
                "by_bucket": {"W1": {"1": 0.4, "2": 0.1, "3": 0.1, "4": 0.1, "5": 0.3}},
                "sizes": {"W1": 100},
                "overall_test": None, "adjacent_option_tests": [],
                "drift": False, "low_n": False,
            },
            {  # 人口题(职业)：保持原顺序，不按占比重排
                "question": "Q3.职业", "type": "single_choice",
                "question_label": "Q3.您的职业是？",
                "options": ["学生", "上班族", "自由职业"],
                "by_bucket": {"W1": {"学生": 0.2, "上班族": 0.7, "自由职业": 0.1}},
                "sizes": {"W1": 100},
                "overall_test": None, "adjacent_option_tests": [],
                "drift": False, "low_n": False,
            },
        ],
        "nps_col": None, "satisfaction_cols": [],
    }
    out = tmp_path / "report.xlsx"
    survey_drift.export_excel(findings, {}, str(out))
    ws = load_workbook(out)["📈 逐题异动明细"]
    # 题目按题号升序：Q1 → Q3 → Q9
    labels = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1)]
    # Q1 五点量表块（首块）：保持 1,2,3,4,5，其后为样本量+加权满意度
    assert labels[:5] == ["1", "2", "3", "4", "5"]
    # Q9 多选题（末块）：按整体占比降序 C, A, B
    assert labels[-4:-1] == ["C", "A", "B"]
    # Q3 人口题(职业)：保持原顺序 学生/上班族/自由职业（不按占比重排）
    assert "学生" in labels
    demo_idx = labels.index("学生")
    assert labels[demo_idx:demo_idx + 3] == ["学生", "上班族", "自由职业"]


def test_detail_sheet_value_labels_and_normalized_block(tmp_path):
    """人口题：编码→标签映射生效、按编码升序、补剔除「不愿意透露」归一化占比+样本量，且归一化行也做异动检验。"""
    findings = {
        "granularity": "month", "time_col": "结束答题时间",
        "buckets": ["M1", "M2"], "bucket_sizes": {"M1": 1000, "M2": 1000},
        "low_n_buckets": [], "metrics": [],
        "questions": [{
            "question": "Q33.请问您的性别是？", "type": "single_choice",
            "question_label": "Q33.请问您的性别是？",
            "options": ["1", "2", "3"],
            # M1: 男40 女40 拒20 → 归一 男/女=50/50；M2: 男20 女60 拒20 → 归一 25/75
            "by_bucket": {"M1": {"1": 0.4, "2": 0.4, "3": 0.2},
                          "M2": {"1": 0.2, "2": 0.6, "3": 0.2}},
            "sizes": {"M1": 1000, "M2": 1000},
            "overall_test": None, "adjacent_option_tests": [],
            "drift": False, "low_n": False,
        }],
        "nps_col": None, "satisfaction_cols": [],
    }
    value_labels = {"Q33.请问您的性别是？": {"1": "男", "2": "女", "3": "不愿意透露"}}
    out = tmp_path / "report.xlsx"
    survey_drift.export_excel(findings, {}, str(out), value_labels=value_labels)
    ws = load_workbook(out)["📈 逐题异动明细"]
    col2 = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1)]
    # 编码替换为标签、按编码升序
    assert col2[:3] == ["男", "女", "不愿意透露"]
    assert col2[3] == "样本量"
    # 归一化子区块小标题 + 剔除「不愿意透露」后的选项 + 剔除后样本量
    assert col2[4] == "剔除「不愿意透露」后归一化"
    assert col2[5:7] == ["男", "女"]
    assert col2[7] == "样本量"
    # 归一化占比：男(整体) = 0.3/(0.3+0.5) = 0.375（norm 行为 row7/row8）
    assert round(ws.cell(7, 3).value, 4) == 0.375
    assert round(ws.cell(8, 3).value, 4) == 0.625
    # M2 归一化：男 = 0.2/0.8 = 0.25
    assert round(ws.cell(7, 5).value, 4) == 0.25
    # 剔除「不愿意透露」后样本量：整体 1600、各期 800
    assert ws.cell(9, 3).value == 1600
    assert ws.cell(9, 4).value == 800 and ws.cell(9, 5).value == 800
    # 归一化行也做异动检验：男 50%→25% 大幅下降，异动周列(第6列)应标注
    assert "▼" in str(ws.cell(7, 6).value) and "pp" in str(ws.cell(7, 6).value)


def test_summary_scope_all_includes_historical_drift(tmp_path):
    """scope=latest 只收录最新相邻期异动；scope=all 收录任意相邻期历史异动并标注时段。"""
    b = ["W1", "W2", "W3"]
    findings = {
        "granularity": "week", "time_col": "结束答题时间",
        "buckets": b, "bucket_sizes": {"W1": 40, "W2": 40, "W3": 40},
        "low_n_buckets": [], "metrics": [],
        "questions": [{
            "question": "Q7.活动评价（单选）", "type": "single_choice",
            "options": ["满意", "一般"],
            "by_bucket": {"W1": {"满意": 1.0, "一般": 0.0},
                          "W2": {"满意": 0.5, "一般": 0.5},
                          "W3": {"满意": 0.5, "一般": 0.5}},
            "sizes": {"W1": 40, "W2": 40, "W3": 40},
            "overall_test": {"test": "chi_square", "p": 0.001, "significant": True},
            "adjacent_option_tests": [
                {"option": "满意", "from": "W1", "to": "W2", "delta_pp": -50.0,
                 "test": "two_prop_z", "p": 0.001, "significant": True,
                 "drift": True, "low_n": False, "direction": "down"},
                {"option": "满意", "from": "W2", "to": "W3", "delta_pp": 0.0,
                 "test": "two_prop_z", "p": 1.0, "significant": False,
                 "drift": False, "low_n": False, "direction": "flat"},
            ],
            "drift": True, "low_n": False,
        }],
        "nps_col": None, "satisfaction_cols": [],
    }

    # scope=latest：最新相邻期(→W3)无异动 → 汇总为"无异动"提示
    out_latest = tmp_path / "latest.xlsx"
    survey_drift.export_excel(findings, {}, str(out_latest), summary_scope="latest")
    ws = load_workbook(out_latest)["⚠️ 异动汇总"]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    assert any("无显著异动" in str(row[0]) for row in rows)
    assert not any(row[2] == "满意" for row in rows if row[2])

    # scope=all：收录 W1→W2 历史异动，时段列标注 "W2"
    out_all = tmp_path / "all.xlsx"
    survey_drift.export_excel(findings, {}, str(out_all), summary_scope="all")
    ws2 = load_workbook(out_all)["⚠️ 异动汇总"]
    data = [[c.value for c in r] for r in ws2.iter_rows(min_row=2)]
    hit = [row for row in data if row[2] == "满意"]
    assert len(hit) == 1, data
    assert hit[0][0] == "Q7.活动评价（单选）"
    assert hit[0][1] == "W2"      # 时段
    assert hit[0][3] == "▼"       # 方向


def test_default_output_filename():
    name = survey_drift.default_output_filename("week")
    assert name.startswith("回流异动诊断_按周_")
    assert name.endswith(".xlsx")


def test_full_pipeline_smoke(tmp_path):
    csv = tmp_path / "reflow.csv"
    pd.DataFrame({
        "结束答题时间": (["2026-04-06 10:00"] * 50 + ["2026-04-13 10:00"] * 50),
        "Q1.整体满意度": [5] * 50 + [3] * 50,
        "Q51.您有多大可能将本游戏推荐给朋友？": [10] * 50 + [4] * 50,
        "Q7.活动评价（单选）": (["满意"] * 50) + (["满意"] * 25 + ["一般"] * 25),
    }).to_csv(csv, index=False, encoding="utf-8-sig")

    findings_out = tmp_path / "drift_findings.json"
    r1 = subprocess.run(
        [sys.executable, str(MODULE_PATH), "analyze",
         "--file_path", str(csv), "--granularity", "week",
         "--findings_out", str(findings_out)],
        capture_output=True, text=True, encoding="utf-8")
    assert r1.returncode == 0, r1.stderr
    p1 = json.loads(r1.stdout)
    assert p1["status"] == "success"
    assert p1["questions_with_drift"] >= 1

    findings = json.loads(findings_out.read_text(encoding="utf-8"))
    conclusions = {q["question"]: "自动结论" for q in findings["questions"]}
    conclusions_out = tmp_path / "conclusions.json"
    conclusions_out.write_text(json.dumps(conclusions, ensure_ascii=False), encoding="utf-8")

    out_xlsx = tmp_path / "report.xlsx"
    r2 = subprocess.run(
        [sys.executable, str(MODULE_PATH), "export",
         "--findings", str(findings_out), "--conclusions", str(conclusions_out),
         "--output_path", str(out_xlsx)],
        capture_output=True, text=True, encoding="utf-8")
    assert r2.returncode == 0, r2.stderr
    assert json.loads(r2.stdout)["status"] == "success"
    assert out_xlsx.exists()


def test_compare_means_zero_variance_no_warning():
    import warnings
    a = pd.Series([5] * 40)
    b = pd.Series([3] * 40)
    with warnings.catch_warnings():
        warnings.simplefilter("error")  # any RuntimeWarning becomes an error
        r = survey_drift.compare_means(a, b)
    assert r["test"] == "degenerate"
    assert r["p"] == 0.0
    assert r["delta"] == 2.0
    # equal constants → not significant
    r2 = survey_drift.compare_means(pd.Series([4] * 40), pd.Series([4] * 40))
    assert r2["p"] == 1.0
